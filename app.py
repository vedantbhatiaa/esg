"""
TIP ESG Platform -- Streamlit Frontend
========================================
Run: streamlit run app.py

Changes from original:
  - formula_engine now imported from formula_engine.py (was .ipynb -- crashed on startup)
  - Removed dead load_consolidated_data() function (used local_storage but was never called)
  - analysis page now reads LONG_DATA from real sector CSV when available
  - data_loader now auto-finds master CSV in data_storage/raw/ (no manual path fix needed)
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import html as _html
import logging
import os
from pathlib import Path
from datetime import datetime, date
from filelock import FileLock

import config as cfg

# ── Logging setup ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
_log = logging.getLogger("esg_app")

from formula_engine import (
    TemplateInputs, calculate, validate_submission,
    get_benchmarks, build_template_dataframe, fmt_num,
    yoy_change, ValidationFlag, BenchmarkResult
)

import data_loader as dl

# Chatbot is embedded in page_readiness only — no global import needed

# Load fresh from disk on every Streamlit rerun (Streamlit reruns the full
# script on every user interaction, so this is always up-to-date after a save).
# data_loader checks data_storage/master/ first, then falls back to raw/ etc.
_CONSOLIDATED_DF = dl.load_consolidated()
_COMPANIES       = dl.get_companies(_CONSOLIDATED_DF)
_SECTOR_DF       = dl.load_sector_aggregated(_CONSOLIDATED_DF)


def _write_verification_status(company: str, year: int, status: str) -> None:
    """
    Persist DSS+ verification status for a company+year to a CSV file.
    Status values: 'Verified', 'Pending', 'Flagged'.
    Client home page reads this file to show the verification chip.
    """
    from pathlib import Path
    import csv, os

    vcsv = Path("data_storage/verifications.csv")
    vcsv.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    if vcsv.exists():
        with open(vcsv, newline="") as f:
            for row in csv.DictReader(f):
                if not (row.get("Company","").strip() == company and
                        str(row.get("Year","")).strip() == str(year)):
                    rows.append(row)   # keep other company/year rows

    rows.append({"Company": company, "Year": str(year), "Status": status,
                 "UpdatedBy": "dss+ Analyst"})

    with open(vcsv, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["Company","Year","Status","UpdatedBy"])
        w.writeheader()
        w.writerows(rows)


def _reload_consolidated_df() -> bool:
    """
    Force-reload master CSV + rebuild ALL dependent globals.
    Uses dl.load_consolidated() dynamic scan so newly-named files
    (e.g. _2024.csv, _2025.csv) are always found without hardcoded filenames.
    Sector aggregation is computed live from master so new years appear immediately.
    """
    global _CONSOLIDATED_DF, _COMPANIES, _SECTOR_DF, _USING_FALLBACK_DATA
    global HIST_YEARS, CURR_YEAR, LONG_YEARS, LONG_DATA, FUEL_MIX
    try:
        fresh = pd.DataFrame()
        # NOTE: SharePoint load_master() not yet implemented in StorageClient.
        # We fall through directly to the local CSV path below.
        if fresh.empty:
            fresh = dl.load_consolidated()
        if not fresh.empty and "Company" in fresh.columns and "Year" in fresh.columns:
            _CONSOLIDATED_DF     = fresh
            _COMPANIES           = dl.get_companies(fresh)
            # Compute sector live — always covers the latest submitted year
            _SECTOR_DF           = dl.load_sector_aggregated(fresh)
            _USING_FALLBACK_DATA = False
            try:
                cfg.refresh_year_bounds(fresh)
                HIST_YEARS = cfg.hist_years()
                CURR_YEAR  = cfg.curr_year()
                LONG_YEARS = cfg.long_years()
            except Exception:
                pass
            try:
                LONG_DATA, FUEL_MIX = _build_long_data()
            except Exception:
                pass
            st.session_state["_df_version"] = st.session_state.get("_df_version", 0) + 1
            return True
    except Exception:
        pass
    return False

# ─────────────────────────────────────────────────────────
# PAGE CONFIG & GLOBAL CSS
# ─────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TIP ESG Platform",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
/* -- Sidebar -- */
[data-testid="stSidebar"] { background: #0A2240 !important; }
[data-testid="stSidebar"] * { color: rgba(255,255,255,0.75) !important; }
[data-testid="stSidebar"] h1,[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,[data-testid="stSidebar"] strong
{ color: #ffffff !important; }
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.1) !important; }
[data-testid="stSidebarNav"] { display: none; }

/* -- Sidebar nav buttons -- */
[data-testid="stSidebar"] .stButton > button {
    background: transparent !important;
    color: rgba(255,255,255,0.75) !important;
    border: none !important;
    border-radius: 8px !important;
    font-size: 13.5px !important;
    font-weight: 500 !important;
    text-align: left !important;
    padding: 9px 12px !important;
    width: 100% !important;
    transition: background .15s, color .15s !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.10) !important;
    color: #ffffff !important;
    border: none !important;
}
[data-testid="stSidebar"] .stButton > button:focus {
    box-shadow: none !important; outline: none !important;
    border: none !important; color: #ffffff !important;
}

/* -- Main buttons -- */
.stButton > button {
    border-radius: 7px; font-weight: 500; font-size: 13px;
    border: 1.5px solid #D1D5DB; transition: all .15s;
}
.stButton > button:hover { border-color: #6B7280; }

/* -- Form inputs -- */
.stNumberInput input, .stTextInput input, .stSelectbox select {
    border-radius: 7px; border: 1.5px solid #D1D5DB; font-size: 14px !important;
}

/* -- KPI cards -- */
.kpi-card {
    background: #fff; border: 1px solid #E5E7EB;
    border-radius: 10px; padding: 16px 18px; text-align: left;
}
.kpi-card .label { font-size: 11px; color: #6B7280;
    text-transform: uppercase; letter-spacing: .5px; font-weight: 500; }
.kpi-card .value { font-size: 26px; font-weight: 700; color: #111827; margin: 5px 0 2px; }
.kpi-card .unit  { font-size: 12px; color: #9CA3AF; }
.kpi-card .delta { font-size: 12px; font-weight: 600; margin-top: 4px; }
.delta-pos { color: #059669; }
.delta-neg { color: #DC2626; }

/* -- Stepper -- */
.step-bar { display:flex; align-items:center; gap:0;
    background:#fff; border:1px solid #E5E7EB; border-radius:10px;
    padding:16px 20px; margin-bottom:20px; }
.step-item { display:flex; align-items:center; flex:1; min-width:0; }
.step-circle { width:28px; height:28px; border-radius:50%;
    display:flex; align-items:center; justify-content:center;
    font-size:12px; font-weight:700; flex-shrink:0; }
.sc-done   { background:#00916E; color:#fff; }
.sc-active { background:#1D4ED8; color:#fff; }
.sc-todo   { background:#F3F4F6; color:#9CA3AF; border:2px solid #E5E7EB; }
.step-label { font-size:11.5px; font-weight:500; margin-left:7px;
    white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.sl-done   { color:#00916E; }
.sl-active { color:#1D4ED8; }
.sl-todo   { color:#9CA3AF; }
.step-line { flex:1; height:2px; background:#E5E7EB; margin:0 6px; min-width:8px; }
.sl-done-line { background:#00916E; }

/* -- Table legend -- */
.tbl-legend { display:flex; gap:14px; padding:10px 16px;
    background:#F9FAFB; border-top:1px solid #E5E7EB;
    border-radius:0 0 8px 8px; flex-wrap:wrap; }
.tl { display:flex; align-items:center; gap:5px; font-size:11px; color:#6B7280; }
.tl-sw { width:14px; height:14px; border-radius:3px;
    border:1px solid #D1D5DB; display:inline-block; }

/* -- Band chart -- */
.band-container { margin: 6px 0 12px; }
.band-row-wrap { display:flex; align-items:center; gap:10px; margin-bottom:10px; }
.band-lbl { font-size:12px; font-weight:500; color:#374151; width:170px; flex-shrink:0; }
.band-track { flex:1; height:18px; border-radius:4px; background:#F3F4F6; position:relative; }
.band-seg { position:absolute; top:0; height:100%; }
.band-pin { position:absolute; width:4px; height:28px;
    background:#0A2240; border-radius:2px; top:-5px; transform:translateX(-50%); }
.band-pin-val { position:absolute; font-size:10px; font-weight:700;
    color:#0A2240; top:-18px; transform:translateX(-50%);
    white-space:nowrap; background:#fff; padding:0 2px; }
.band-chip { font-size:11px; font-weight:600; padding:3px 9px;
    border-radius:10px; flex-shrink:0; }
.chip-top  { background:#D1FAE5; color:#065F46; }
.chip-mid  { background:#FEF3C7; color:#92400E; }
.chip-bot  { background:#FEE2E2; color:#991B1B; }

/* -- Flag cards -- */
.flag-card { display:flex; align-items:flex-start; gap:10px;
    padding:12px 14px; border-radius:8px; border:1px solid; margin-bottom:10px; }
.fc-warn  { background:#FFF7ED; border-color:#FCD34D; }
.fc-error { background:#FEF2F2; border-color:#FECACA; }
.fc-ok    { background:#ECFDF5; border-color:#6EE7B7; }
.fc-icon  { width:20px; height:20px; border-radius:50%;
    display:flex; align-items:center; justify-content:center;
    font-size:10px; font-weight:700; flex-shrink:0; }
.fi-warn  { background:#D97706; color:#fff; }
.fi-error { background:#DC2626; color:#fff; }
.fi-ok    { background:#00916E; color:#fff; }
.fc-title { font-size:13px; font-weight:600; color:#111827; }
.fc-detail{ font-size:12px; color:#6B7280; margin-top:3px; line-height:1.6; }

/* -- AI card -- */
.ai-card { background:#fff; border:1px solid #E5E7EB;
    border-radius:10px; overflow:hidden; margin-bottom:12px; }
.ai-head  { display:flex; align-items:center; gap:8px;
    padding:10px 14px; background:#F9FAFB; border-bottom:1px solid #E5E7EB; }
.ai-pulse { width:8px; height:8px; border-radius:50%;
    background:#00916E; flex-shrink:0; }
.ai-title { font-size:13px; font-weight:600; color:#111827; }
.ai-badge { margin-left:auto; background:#E6F5F1; color:#007A5C;
    font-size:11px; padding:2px 9px; border-radius:10px;
    border:1px solid #6EE7B7; font-weight:500; }
.ai-body  { padding:12px 14px; font-size:13px; color:#374151; line-height:1.8; }
</style>
""", unsafe_allow_html=True)

# ── Animation system & design tokens ─────────────────────────────────────────
from ui_components import (
    inject_global_css, kpi_card_html, skeleton_card_html, skeleton_chart_html,
    status_chip_html, section_header_html, empty_state_html, co_card_html,
    apply_chart_animation, chart_layout_defaults, sparkline_html,
    GREEN, AMBER, RED, NAVY, BG, BORDER, TEXT, MUTED,
    CAT_CO2, CAT_ENERGY, CAT_WATER, CAT_WASTE, CAT_RENEW,
)
inject_global_css()

# ─────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────
HIST_YEARS = list(range(2009, 2023))
CURR_YEAR  = 2023
LONG_YEARS = list(range(2009, 2024))

# ── All 31 electricity-by-country names (matches UI editor row order) ─────────
ELEC_ALL_COUNTRIES = [
    "Canada", "Chile", "Mexico", "United States",
    "Australia", "Japan", "Korea", "New Zealand",
    "Austria", "Belgium", "Czech Republic", "Denmark", "Finland", "France",
    "Germany", "Hungary", "Iceland", "Ireland", "Italy", "Luxembourg",
    "Netherlands", "Norway", "Poland", "Portugal", "Spain", "Sweden",
    "Switzerland", "Turkey", "United Kingdom",
    "China", "India",
]

def _elec_col(country: str) -> str:
    """Canonical master CSV column name for a country's electricity (GJ)."""
    return "Elec_" + country.replace(" ", "_") + "_GJ"

# Dict: country name → master CSV column name
ELEC_COUNTRY_COLS = {c: _elec_col(c) for c in ELEC_ALL_COUNTRIES}

COMPANIES = _COMPANIES if _COMPANIES else [
    "VerdaTyres Corp", "AlphaTread Ltd", "BetaRubber Inc", "GammaTire SA",
    "DeltaGrip GmbH", "EpsilonWheel Co", "ZetaTrac LLC", "EtaRoad AG",
    "ThetaDrive NV", "IotaTire PLC",
]

# ── Year bounds — dynamic from real data, config-driven fallback ─────────────
# cfg.refresh_year_bounds() updates cfg.DATA_YEAR_START / DATA_YEAR_END in-place
cfg.refresh_year_bounds(_CONSOLIDATED_DF)
HIST_YEARS = cfg.hist_years()   # e.g. [2009..2022]
CURR_YEAR  = cfg.curr_year()    # e.g. 2023
LONG_YEARS = cfg.long_years()   # e.g. [2009..2023]

# ── Client auth mapping — loaded from config (secrets.toml or env var) ───────
# In production set CLIENTS_JSON in secrets.toml. Demo fallback in config.py.
CLIENTS = cfg.load_clients()
DSS_DOMAIN = cfg.DSS_EMAIL_DOMAIN

# -- Static fallback HIST_RAW (used only if no company selected yet) -----------
HIST_RAW: dict[str, list] = {
    "total_sites":   [38,39,40,40,41,42,43,44,46,48,51,51,52,52],
    "iso_sites":     [36,38,39,39,40,41,42,43,45,47,51,51,52,52],
    "production":    [2_840_000,3_510_000,3_770_000,3_520_000,3_640_000,3_620_000,
                      3_540_000,3_630_000,3_700_000,3_910_000,3_860_000,3_050_000,3_320_000,3_580_000],
    "water_withdrawals": [18_000_000,19_200_000,20_100_000,19_700_000,20_500_000,21_000_000,
                          21_500_000,22_000_000,22_800_000,23_100_000,23_100_000,20_300_000,21_100_000,20_900_000],
    "renew_elec_purchased": [0,0,0,0,0,0,0,300_000,250_000,1_100_000,706_562,1_528_836,2_557_561,4_082_923],
    "nonrenew_elec_purchased": [9_500_000,9_400_000,9_600_000,9_300_000,9_400_000,9_500_000,
                                9_600_000,9_400_000,9_300_000,9_100_000,12_271_131,9_667_437,10_297_758,9_037_549],
    "nat_gas": [13_000_000,14_000_000,15_000_000,14_500_000,15_000_000,15_200_000,
                15_500_000,15_700_000,15_900_000,16_100_000,16_210_969,14_040_397,15_939_109,15_927_554],
    "coal_sub": [500_000,490_000,480_000,470_000,460_000,450_000,
                 440_000,430_000,420_000,410_000,456_997,337_992,360_848,395_006],
    "lpg": [1_000_000,1_050_000,1_100_000,1_100_000,1_100_000,1_150_000,
            1_150_000,1_180_000,1_200_000,1_220_000,1_237_839,1_124_479,1_271_422,1_329_571],
    "waste_total":    [330_000,330_000,335_000,335_000,340_000,342_000,
                       344_000,346_000,348_000,350_000,352_000,295_000,320_000,335_000],
    "waste_recovery": [280_000,281_000,283_000,284_000,286_000,287_000,
                       289_000,292_000,295_000,298_000,299_200,253_700,275_200,284_750],
}

# -- LONG_DATA: built from real sector CSV if available, else static fallback --
def _build_long_data() -> tuple[dict, dict]:
    """
    Build LONG_DATA and FUEL_MIX from the real consolidated wide DataFrame.
    Falls back to static values only for missing years/fields.
    """
    static_long = {
        "energy":     [28.1,32.3,33.6,32.4,33.0,32.4,31.8,32.1,33.0,34.2,33.2,28.5,32.3,32.5,32.4],
        "co2":        [2.41,2.69,2.88,2.80,2.87,2.86,2.73,2.72,2.80,2.85,2.76,2.22,2.27,2.06,2.05],
        "water":      [22.4,23.8,24.9,24.4,23.9,22.9,22.9,23.5,23.5,23.2,23.1,20.3,21.1,20.9,21.5],
        "scope1":     [1.08,1.19,1.21,1.17,1.20,1.15,1.09,1.08,1.12,1.15,1.11,0.94,1.06,1.05,1.03],
        "scope2":     [1.33,1.50,1.67,1.63,1.67,1.71,1.64,1.64,1.68,1.70,1.65,1.27,1.21,1.01,1.02],
        "energy_kpi": [9.9,9.2,8.9,9.2,9.1,8.9,8.9,8.8,8.9,8.8,8.6,9.3,9.7,9.1,8.7],
        "co2_kpi":    [0.850,0.765,0.764,0.795,0.789,0.791,0.771,0.748,0.758,0.729,0.715,0.729,0.684,0.576,0.551],
        "renew_pct":  [0,0,0,0,0,0,0,2.3,2.2,9.7,10.6,21.8,31.4,40.6,48.3],
        "waste_recov":[83,83,84,84,84,84,84,85,85,85,85,86,86,85,86],
        "prod":       [2.84,3.51,3.77,3.52,3.64,3.62,3.54,3.63,3.70,3.91,3.86,3.05,3.32,3.58,3.72],
    }
    static_fuel = {
        "Natural Gas": [46,46,47,47,47,46,47,47,48,49,49,49,49,49,50],
        "Electricity": [34.7,34.2,34.9,34.6,35.3,36.7,37.1,37.9,38.2,38.8,39.1,39.3,39.8,40.4,40.7],
        "Fuel Oil":    [8.5,6.7,6,5.8,5.1,4.8,3.7,3.2,3,2.6,2.4,1.8,1.4,0.5,0.5],
        "LPG":         [2.4,2.4,2.3,3.6,3.5,3.5,3.5,3.6,3.5,3.6,3.7,3.9,3.9,4.1,4.2],
        "Coal":        [3.2,3.1,2.8,2.8,3.7,3.6,2.3,2,2.1,1.6,1.4,1.2,1.2,1.2,1.2],
        "Other":       [5.2,7.6,7,5.4,5.4,5.4,6.4,5.3,4.7,4.4,4,4.8,4.3,4.8,3.4],
    }

    def _safe_list(series, fallback):
        result = []
        for i, v in enumerate(series):
            try:
                f = float(v)
                result.append(f if not np.isnan(f) else fallback[i])
            except Exception:
                result.append(fallback[i])
        return result

    df = _CONSOLIDATED_DF
    if df.empty or "Row_Label" in df.columns:
        # Long format or no data -- use static
        if not _SECTOR_DF.empty:
            try:
                s = _SECTOR_DF.set_index("Year").reindex(LONG_YEARS)
                live = {
                    "energy":     _safe_list((s["Total_Energy"]/1e6), static_long["energy"]),
                    "co2":        _safe_list((s["Total_CO2"]/1e6),    static_long["co2"]),
                    "water":      _safe_list((s["Total_Water"]/1e6),  static_long["water"]),
                    "energy_kpi": _safe_list(s["Avg_Energy_KPI"],     static_long["energy_kpi"]),
                    "co2_kpi":    _safe_list(s["Avg_CO2_KPI"],        static_long["co2_kpi"]),
                    "renew_pct":  _safe_list(s["Avg_Renewable_Share"],static_long["renew_pct"]),
                    "prod":       _safe_list((s["Total_Production"]/1e6), static_long["prod"]),
                    "scope1":     static_long["scope1"],
                    "scope2":     static_long["scope2"],
                    "waste_recov":static_long["waste_recov"],
                }
                return live, static_fuel
            except Exception as e:
                _log.warning("[app] Sector DF error: %s", e)
        return static_long, static_fuel

    # Wide format -- compute directly from master DataFrame
    try:
        grp = df.groupby("Year")

        def _col_sum(col, divisor=1):
            if col in df.columns:
                return grp[col].sum() / divisor
            return None

        def _col_mean(col):
            if col in df.columns:
                return grp[col].mean()
            return None

        def _col_sum_norm(col, divisor=1):
            """Sector sum normalised by n_submitting / n_total_companies."""
            if col not in df.columns:
                return None
            raw = grp[col].sum() / divisor
            n_sub = grp["Company"].count()
            n_all = df["Company"].nunique()
            return raw / n_sub * n_all

        energy_s  = _col_sum_norm("Total energy", 1e6)
        co2_s     = _col_sum_norm("Total CO2", 1e6)
        scope1_s  = _col_sum_norm("Total CO2 - Scope 1", 1e6)
        scope2_s  = _col_sum_norm("Total CO2 - Scope 2", 1e6)
        water_s   = _col_sum_norm("Water intake", 1e6)
        # Use MEAN × n_submitting_companies so partial-year submissions
        # don't cause a false cliff-drop in sector production.
        prod_s_raw  = _col_sum("Production", 1e6)
        n_companies = df["Company"].nunique()
        if prod_s_raw is not None:
            n_submitting = grp["Company"].count()
            prod_s       = prod_s_raw / n_submitting * n_companies
        else:
            prod_s = prod_s_raw
        ekpi_m    = _col_mean("Total energy - KPI")
        co2kpi_m  = _col_mean("Total CO2 - KPI")
        # For renewable %, only average companies that have submitted for that year
        # (NaN values from non-submitting companies would drag the mean down)
        renew_m   = df.groupby("Year")["Renewable_Electricity_Share_%"].apply(
            lambda x: x.dropna().mean() if x.dropna().size > 0 else float("nan")
        ) if "Renewable_Electricity_Share_%" in df.columns else None

        def _to_list(series, fallback):
            if series is None:
                return fallback
            s = series.reindex(LONG_YEARS)
            return _safe_list(s.values, fallback)

        live = {
            "energy":     _to_list(energy_s,  static_long["energy"]),
            "co2":        _to_list(co2_s,     static_long["co2"]),
            "scope1":     _to_list(scope1_s,  static_long["scope1"]),
            "scope2":     _to_list(scope2_s,  static_long["scope2"]),
            "water":      _to_list(water_s,   static_long["water"]),
            "prod":       _to_list(prod_s,    static_long["prod"]),
            "energy_kpi": _to_list(ekpi_m,    static_long["energy_kpi"]),
            "co2_kpi":    _to_list(co2kpi_m,  static_long["co2_kpi"]),
            "renew_pct":  _to_list(renew_m,   static_long["renew_pct"]),
            "waste_recov":static_long["waste_recov"],
        }

        # Fuel mix as % of total energy per year
        fuel_cols = {
            "Natural Gas": "Natural Gas",
            "Electricity": "Total Electricity",
            "Fuel Oil":    "Fuel Oil",
            "LPG":         "LPG",
            "Coal":        "Coal",
            "Other":       "Other",
        }
        total_e_by_yr = grp["Total energy"].sum() if "Total energy" in df.columns else None
        live_fuel = {}
        n_yrs = len(LONG_YEARS)
        for label, col in fuel_cols.items():
            _fb = static_fuel.get(label, [])
            _fb_ext = (_fb + [_fb[-1] if _fb else 0.0] * max(0, n_yrs - len(_fb)))
            if col in df.columns and total_e_by_yr is not None:
                fuel_sum = grp[col].sum().reindex(LONG_YEARS)
                total_e  = total_e_by_yr.reindex(LONG_YEARS)
                pct = (fuel_sum / total_e.replace(0, np.nan) * 100).fillna(0)
                live_fuel[label] = _safe_list(pct.values, _fb_ext)
            else:
                live_fuel[label] = _fb_ext[:n_yrs]

        return live, live_fuel if any(sum(v) > 0 for v in live_fuel.values()) else static_fuel

    except Exception as e:
        _log.warning("[app] Wide DF live computation error: %s", e)
        return static_long, static_fuel


LONG_DATA, FUEL_MIX = _build_long_data()

# L2 FIX: track whether analysis charts are showing real or fallback data.
# Surfaced as a warning banner in page_analysis() so analysts never mistake
# synthetic demo numbers for real client submissions.
_USING_FALLBACK_DATA = _CONSOLIDATED_DF.empty

CLIENTS = {
    "verdatyres@tip-reporting.com":   "VerdaTyres Corp",
    "alphatread@tip-reporting.com":   "AlphaTread Ltd",
    "betarubber@tip-reporting.com":   "BetaRubber Inc",
    "gammatire@tip-reporting.com":    "GammaTire SA",
    "deltagrip@tip-reporting.com":    "DeltaGrip GmbH",
    "epsilonwheel@tip-reporting.com": "EpsilonWheel Co",
    "zetatrac@tip-reporting.com":     "ZetaTrac LLC",
    "etaroad@tip-reporting.com":      "EtaRoad AG",
    "thetadrive@tip-reporting.com":   "ThetaDrive NV",
    "iotatire@tip-reporting.com":     "IotaTire PLC",
}

STEP_META = [
    ("ISO 14001",  "Certified sites and facility coverage"),
    ("Production", "Annual production volume"),
    ("Water",      "Water withdrawals by source"),
    ("Energy",     "Electricity and fuel consumption"),
    ("CO2",        "Emission inputs and auto-calculated totals"),
    ("Waste",      "Waste generated, recovered and eliminated"),
]

# ─────────────────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "authenticated":      False,
        "user_name":          "",
        "user_company":       "",
        "user_email":         "",
        "is_dss":             False,
        "page":               "login",
        "step":               0,
        "template_done":      False,
        "company_setup_done": False,
        "reporting_company":  "",
        "reporting_year":     2023,
        "employee_name":      "",
        "company_hist":       {},
        "live_hist_raw":      {},
        "kpi_hints":          {},
        "step_data": {
            "total_sites": 54, "iso_sites": 54,
            "production": 3_720_000,
            "water_withdrawals": 21_500_000,
            "renew_elec_purchased": 5_200_000,
            "nonrenew_elec_purchased": 8_500_000,
            "self_gen_elec": 45_000,
            "purchased_steam": 1_050_000,
            "sold_electricity": 8_000,
            "sold_steam": 0,
            "nat_gas": 16_100_000, "coal_sub": 380_000,
            "propane": 340_000, "fuel_oil_heavy_a": 150_000,
            "diesel": 190_000, "petrol": 0, "biomass": 0,
            "waste_tires_mt": 0, "lpg": 1_350_000, "other_fuels": 0,
            "co2_scope2_steam": 60_000,
            "waste_total": 338_000, "waste_recovery": 290_000,
        },
        "flags_resolved": set(),
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────
def get_current_outputs():
    sd = st.session_state.step_data
    inp = TemplateInputs(
        company=st.session_state.get("reporting_company") or st.session_state.user_company,
        year=st.session_state.get("reporting_year", CURR_YEAR),
        **{k: float(sd.get(k, 0)) for k in [
            "total_sites", "iso_sites", "production", "water_withdrawals",
            "renew_elec_purchased", "nonrenew_elec_purchased", "self_gen_elec",
            "purchased_steam", "sold_electricity", "sold_steam",
            "nat_gas", "coal_sub", "propane", "fuel_oil_heavy_a",
            "diesel", "petrol", "biomass", "waste_tires_mt", "lpg", "other_fuels",
            "co2_scope2_steam", "waste_total", "waste_recovery",
        ]}
    )
    return inp, calculate(inp)


# Valid TemplateInputs fields -- used to guard against unexpected keys from consolidated data
_VALID_TEMPLATE_FIELDS = {
    "total_sites", "iso_sites", "production", "water_withdrawals",
    "renew_elec_purchased", "nonrenew_elec_purchased", "self_gen_elec",
    "purchased_steam", "sold_electricity", "sold_steam",
    "nat_gas", "coal_sub", "propane", "fuel_oil_heavy_a",
    "diesel", "petrol", "biomass", "waste_tires_mt", "lpg", "other_fuels",
    "co2_scope2_steam", "waste_total", "waste_recovery",
}

def _get_fresh_hist(company: str = None) -> dict:
    """
    Load company historical data for ALL available years from _CONSOLIDATED_DF.
    Includes the current year if a row exists (e.g. after a save).
    Falls back to HIST_RAW (static demo data) when company is unknown.
    """
    co = company or st.session_state.get("reporting_company") or st.session_state.get("user_company") or ""
    if co and not _CONSOLIDATED_DF.empty:
        hist = dl.get_company_hist(_CONSOLIDATED_DF, co)
        if hist:
            # Use ALL years present in the DB, not just the pre-2023 window
            all_years = sorted(dl.get_years(_CONSOLIDATED_DF, co) or [])
            return dl.get_hist_raw(hist, all_years) if all_years else dl.get_hist_raw(hist, HIST_YEARS)
    return st.session_state.get("live_hist_raw") or HIST_RAW


def get_hist_outputs():
    """
    Return list of (year, TemplateInputs, TemplateOutputs) for ALL years in the DB.
    Uses year-keyed dict lookup — avoids positional list drift when fields missing.
    Always reads from _CONSOLIDATED_DF so any saved update is immediately visible.
    """
    company = (st.session_state.get("reporting_company") or
               st.session_state.get("user_company") or "")
    if company and not _CONSOLIDATED_DF.empty:
        all_years  = sorted(dl.get_years(_CONSOLIDATED_DF, company) or [])
        comp_hist  = dl.get_company_hist(_CONSOLIDATED_DF, company)
    else:
        all_years  = list(HIST_YEARS)
        comp_hist  = {}
    outs = []
    for yr in all_years:
        step  = dl.get_step_data(comp_hist, yr) if comp_hist else {}
        clean = {k: v for k, v in step.items() if k in _VALID_TEMPLATE_FIELDS}
        inp   = TemplateInputs(company=company, year=yr, **clean)
        outs.append((yr, inp, calculate(inp)))
    return outs



# kpi_card_html is imported from ui_components (see import block above).
# The old local definition has been removed to avoid shadowing the import.


def band_html(label, val, q25, median, q75, unit, lower_better=True):
    span    = q75 - q25
    bmin    = q25 - span * 0.4
    bmax    = q75 + span * 0.4
    rng     = bmax - bmin if (bmax - bmin) != 0 else 1
    pos_pct = max(2, min(98, (val - bmin) / rng * 100))
    if lower_better:
        top_cls = "chip-top" if val <= q25 else "chip-mid" if val <= median else "chip-bot"
        top_lbl = "Top 25%"  if val <= q25 else "Average"  if val <= median else "Below avg"
    else:
        top_cls = "chip-top" if val >= q75 else "chip-mid" if val >= median else "chip-bot"
        top_lbl = "Top 25%"  if val >= q75 else "Average"  if val >= median else "Below avg"
    val_str = f"{val:.3f}" if isinstance(val, float) and val < 10 else fmt_num(val)
    # H3 FIX: escape label before injecting into HTML
    e = _html.escape
    return f"""
    <div class="band-row-wrap">
      <div class="band-lbl">{e(str(label))}</div>
      <div class="band-track">
        <div class="band-seg" style="left:0%;width:25%;background:#D1FAE5;border-radius:4px 0 0 4px"></div>
        <div class="band-seg" style="left:25%;width:50%;background:#FEF3C7"></div>
        <div class="band-seg" style="left:75%;width:25%;background:#FEE2E2;border-radius:0 4px 4px 0"></div>
        <div class="band-pin" style="left:{pos_pct}%">
          <div class="band-pin-val" style="left:0">{val_str} {unit}</div>
        </div>
      </div>
      <span class="band-chip {top_cls}">{top_lbl}</span>
    </div>"""


# ─────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────
def show_login():
    # ── Dark page background ───────────────────────────────────────────────────
    st.markdown("""<style>
    [data-testid="stApp"]  { background:#1a1b6b !important; }
    [data-testid="stHeader"]{ display:none !important; }
    .main                  { background:transparent !important; }
    .block-container       { max-width:520px !important; padding:80px 24px 40px !important; margin:auto !important; }
    /* Card appearance for the whole block-container */
    .block-container > div {
        background:#f0f1f7 !important;
        border-radius:16px !important;
        padding:38px 40px 32px !important;
        box-shadow:0 25px 70px rgba(0,0,0,.45) !important;
    }
    /* Input field styling */
    [data-testid="stTextInput"] input {
        background:#fff !important; border:1px solid #dde0f0 !important;
        border-radius:7px !important; padding:11px 14px !important;
        font-size:14px !important; color:#1a1b6b !important;
    }
    [data-testid="stTextInput"] label {
        font-size:10px !important; font-weight:700 !important;
        letter-spacing:.9px !important; color:#8b90a0 !important;
        text-transform:uppercase !important;
    }
    /* Sign-in button */
    [data-testid="stButton"] > button[kind="primary"] {
        background:#111827 !important; color:#fff !important;
        border:none !important; border-radius:8px !important;
        font-size:15px !important; font-weight:500 !important;
        padding:13px !important; letter-spacing:.1px !important;
    }
    [data-testid="stButton"] > button[kind="primary"]:hover {
        background:#1f2937 !important;
    }
    /* Radio tabs look */
    [data-testid="stRadio"] > div {
        background:#e5e6ef; border-radius:8px; padding:4px;
        display:flex; gap:4px;
    }
    [data-testid="stRadio"] label {
        flex:1; text-align:center; padding:7px 12px;
        border-radius:6px; font-size:13px; font-weight:500;
        color:#6b7280; cursor:pointer;
    }
    [data-testid="stRadio"] [aria-checked="true"] + div {
        background:#fff !important;
    }
    </style>""", unsafe_allow_html=True)

    # ── Logo ──────────────────────────────────────────────────────────────────
    st.markdown("""
    <div style="margin-bottom:4px">
      <span style="font-size:34px;font-weight:800;color:#dc2626;font-style:italic;letter-spacing:-1px">dss</span><span
            style="font-size:34px;font-weight:800;color:#1a1b6b;font-style:italic;letter-spacing:-1px">+</span>
    </div>
    <div style="font-size:10px;font-weight:600;color:#9ca3af;letter-spacing:2.5px;margin-bottom:28px">
      PROTECT · TRANSFORM · SUSTAIN
    </div>""", unsafe_allow_html=True)

    # ── Role tabs ─────────────────────────────────────────────────────────────
    role = st.radio("", ["TIP Client Company", "dss+ Analyst"],
                    horizontal=True, key="login_role",
                    label_visibility="collapsed")
    st.markdown('<div style="height:14px"></div>', unsafe_allow_html=True)

    # Auto-switch email to correct default when role changes
    _prev_role = st.session_state.get("_login_prev_role")
    if _prev_role != role:
        st.session_state["_login_prev_role"] = role
        st.session_state["login_email"] = (
            "verdatyres@tip-reporting.com"
            if role == "TIP Client Company"
            else "employee@consultdss.com"
        )

    # ── Fields ────────────────────────────────────────────────────────────────
    email    = st.text_input("EMAIL ADDRESS", key="login_email")
    password = st.text_input("PASSWORD", type="password", value="demo1234", key="login_pw")
    st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)

    # ── Sign-in button ─────────────────────────────────────────────────────────
    if st.button("Sign in to workspace", type="primary",
                 use_container_width=True, key="login_btn"):
        email_l   = email.strip().lower()
        is_dss    = DSS_DOMAIN in email_l
        is_client = email_l in CLIENTS
        if not is_dss and not is_client:
            st.error("Email not recognised. Use the demo credentials below.")
        else:
            name_parts = email.split("@")[0].replace(".", " ").split()
            name       = " ".join(p.capitalize() for p in name_parts)
            st.session_state.authenticated = True
            st.session_state.user_email    = email_l
            st.session_state.user_name     = name
            st.session_state.is_dss        = is_dss
            st.session_state.user_company  = "All Companies" if is_dss else CLIENTS[email_l]
            st.session_state.page          = "portfolio" if is_dss else "home"
            st.rerun()

    # ── Demo credentials ──────────────────────────────────────────────────────
    st.markdown("""
    <div style="text-align:center;margin-top:18px;font-size:11px;color:#9ca3af;line-height:1.7">
      Demo: verdatyres@tip-reporting.com (Client) ·<br>
      analyst@consultdss.com (dss+)
    </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# SIDEBAR — two-shell navigation
# ─────────────────────────────────────────────────────────
def _nav_item(page_id: str, label: str) -> None:
    """Render one sidebar nav item. Active item is highlighted green."""
    active = st.session_state.page == page_id
    if active:
        st.markdown(
            f'<div style="background:rgba(22,163,74,0.85);border-radius:7px;'
            f'padding:8px 14px;margin-bottom:2px;color:#fff;font-size:13px;'
            f'font-weight:600">{label}</div>',
            unsafe_allow_html=True,
        )
    else:
        if st.sidebar.button(label, key=f"nav_{page_id}", use_container_width=True):
            st.session_state.page = page_id
            st.rerun()


def show_sidebar():
    with st.sidebar:
        # ── Logo ─────────────────────────────────────────────────────────────
        st.markdown("""
        <div style="padding:16px 14px 12px;border-bottom:1px solid rgba(255,255,255,.08)">
          <div style="display:flex;align-items:center;gap:8px">
            <div style="width:10px;height:10px;border-radius:50%;background:#16A34A;flex-shrink:0"></div>
            <div>
              <div style="color:#fff;font-size:14px;font-weight:700;letter-spacing:-.2px">TIP ESG Platform</div>
              <div style="color:rgba(255,255,255,.35);font-size:10px;margin-top:1px">dss+ · Tire Industry Project</div>
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Company badge (client only) ───────────────────────────────────────
        if not st.session_state.is_dss:
            _safe_co = _html.escape(st.session_state.user_company)
            st.markdown(f"""
            <div style="margin:10px 10px 0;padding:8px 12px;background:rgba(255,255,255,.06);
                border-radius:8px;border:1px solid rgba(255,255,255,.08)">
              <div style="color:rgba(255,255,255,.35);font-size:9px;text-transform:uppercase;
                  letter-spacing:.6px">Your Company</div>
              <div style="color:#fff;font-size:13px;font-weight:500;margin-top:2px">{_safe_co}</div>
            </div>""", unsafe_allow_html=True)

        # ── CLIENT navigation ─────────────────────────────────────────────────
        if not st.session_state.is_dss:
            _nav_item("home",         "Home")
            _nav_item("dashboard",    "My Dashboard")
            _nav_item("my_records",   "My Records")
            _nav_item("benchmarking", "Benchmarks")
            _nav_item("reports",      "Reports")
            _nav_item("entry",        "Submit Data")
            _nav_item("settings",     "Settings")

        # ── DSS+ INTERNAL navigation ──────────────────────────────────────────
        else:
            _nav_item("portfolio",      "Portfolio")
            _nav_item("company_data",   "Company Data")
            _nav_item("verification",   "Verification Queue")
            _nav_item("analysis",       "Analysis")
            _nav_item("benchmarking",   "Benchmarks")
            _nav_item("readiness",      "AI Assistant")
            _nav_item("doc_library",    "Document Library")
            _nav_item("sector_reports", "Sector Reports")
            _nav_item("admin",          "Admin")
            _nav_item("settings",       "Settings")
            # Submit Data hidden for now — restore by uncommenting:
            # _nav_item("entry", "Submit Data")

        # ── User footer ───────────────────────────────────────────────────────
        st.markdown("---")
        name_init  = _html.escape(
            "".join(p[0].upper() for p in st.session_state.user_name.split()[:2])
        )
        _safe_name = _html.escape(st.session_state.user_name)
        role_lbl   = "dss+ Analyst" if st.session_state.is_dss else f"Client · {st.session_state.user_company}"
        _safe_role = _html.escape(role_lbl)
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:9px;padding:0 2px">
          <div style="width:30px;height:30px;border-radius:50%;background:#16A34A;color:#fff;
              font-size:11px;font-weight:700;display:flex;align-items:center;
              justify-content:center;flex-shrink:0">{name_init}</div>
          <div>
            <div style="color:#fff;font-size:13px;font-weight:500">{_safe_name}</div>
            <div style="color:rgba(255,255,255,.4);font-size:10px">{_safe_role}</div>
          </div>
        </div>""", unsafe_allow_html=True)
        st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
        if st.button("Sign out", use_container_width=True, key="signout_btn"):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()


# ─────────────────────────────────────────────────────────
# STEPPER BAR
# ─────────────────────────────────────────────────────────
def render_stepper_bar():
    items = ""
    for i, (name, _) in enumerate(STEP_META):
        if i < st.session_state.step:
            sc, sl, icon = "sc-done",   "sl-done",   "v"
        elif i == st.session_state.step:
            sc, sl, icon = "sc-active", "sl-active", str(i+1)
        else:
            sc, sl, icon = "sc-todo",   "sl-todo",   str(i+1)
        line = f'<div class="step-line {"sl-done-line" if i>0 and i<=st.session_state.step else ""}"></div>' if i > 0 else ""
        items += f'{line}<div class="step-item"><div class="step-circle {sc}">{icon}</div><span class="step-label {sl}">{name}</span></div>'
    st.markdown(f'<div class="step-bar">{items}</div>', unsafe_allow_html=True)


STEP_FIELDS = [
    # Step 0: ISO 14001
    [("total_sites", "Total no. of sites",           "All facilities globally (no.)",              None),
     ("iso_sites",   "ISO 14001 certified sites",    "Sites with active ISO 14001 certification",  None)],
    # Step 1: Production
    [("production",  "Annual production",            "Total weight of finished products (metric t)", None)],
    # Step 2: Water
    [("water_withdrawals", "Total water withdrawals","All sources: surface, ground, municipal (m³)",None)],
    # Step 3: Energy
    [("renew_elec_purchased",    "Renewable electricity purchased",              "Grid-purchased certified renewable electricity (GJ)", None),
     ("nonrenew_elec_purchased", "Non-renewable electricity purchased",          "Standard grid electricity (GJ)",                      None),
     ("self_gen_elec",           "Self-generated renewable electricity on-site", "On-site solar, wind, hydro (GJ)",                     None),
     ("purchased_steam",         "Purchased steam",                              "GJ",                                                  None),
     ("sold_electricity",        "Sold electricity",                             "GJ (enter 0 if none)",                                None),
     ("sold_steam",              "Sold steam",                                   "GJ (enter 0 if none)",                                None),
     ("nat_gas",                 "Natural gas",                                  "GJ LHV",                                              None),
     ("coal_sub",                "Coal (all types)",                             "GJ LHV",                                              None),
     ("propane",                 "Propane",                                      "GJ LHV",                                              None),
     ("fuel_oil_heavy_a",        "Fuel oil",                                     "GJ LHV",                                              None),
     ("diesel",                  "Diesel",                                       "GJ LHV",                                              None),
     ("petrol",                  "Petrol",                                       "GJ LHV",                                              None),
     ("biomass",                 "Biomass",                                      "GJ LHV (biogenic CO2 excluded)",                      None),
     ("waste_tires_mt",          "Waste tires",                                  "metric t (converted to GJ internally)",               None),
     ("lpg",                     "LPG",                                          "GJ LHV",                                              None),
     ("other_fuels",             "Other fuels",                                  "GJ LHV",                                              None)],
    # Step 4: CO2
    [("co2_scope2_steam", "Scope 2 CO2 from purchased steam", "T.CO2 -- company-provided figure from steam supplier", None)],
    # Step 5: Waste
    [("waste_total",    "Total waste generated",  "metric t — all waste streams",                  None),
     ("waste_recovery", "Waste sent to recovery", "metric t — recycling, composting, energy rec.", None)],
]


# ─────────────────────────────────────────────────────────
# PAGE 1 -- KPI DATA ENTRY
# ─────────────────────────────────────────────────────────
def _build_master_row(inp, out) -> dict:
    """
    Build a dict whose keys exactly match the master wide CSV column names.
    Ensures no duplicate columns when appended to the master CSV.
    Includes waste KPIs and electricity-by-country columns.
    """
    renew_share  = (inp.renew_elec_purchased + inp.self_gen_elec) / max(out.total_electricity, 1) * 100
    scope1_share = out.total_co2_scope1 / max(out.total_co2, 1) * 100
    scope2_share = out.total_co2_scope2 / max(out.total_co2, 1) * 100
    fuel_total   = (inp.nat_gas + inp.coal_sub + inp.propane + inp.fuel_oil_heavy_a + inp.diesel + inp.petrol)
    fossil_share = fuel_total / max(out.total_energy, 1) * 100
    prod         = max(inp.production, 1)

    # Waste derived
    waste_total    = float(getattr(inp, "waste_total", 0) or 0)
    waste_recovery = float(getattr(inp, "waste_recovery", 0) or 0)
    recovery_rate  = round(waste_recovery / waste_total * 100, 4) if waste_total else 0.0
    waste_elim     = round(waste_total - waste_recovery, 4)

    row = {
        "Company": inp.company, "Year": inp.year,
        "Total no. of sites": int(round(inp.total_sites)),
        "ISO 14001 sites":    int(round(inp.iso_sites)),
        "% certified sites":  round(out.pct_certified, 6),
        "Production":         round(inp.production, 4),
        "Water intake":       round(inp.water_withdrawals, 4),
        "Water intake - KPI": round(out.water_kpi, 6),
        "Total Electricity":                               round(out.total_electricity, 4),
        "Renewable Electricity Purchased":                 round(inp.renew_elec_purchased, 4),
        "Non-Renewable Electricity Purchased":             round(inp.nonrenew_elec_purchased, 4),
        "Self-generated AND consumed electricity on-site": round(inp.self_gen_elec, 4),
        "Purchased Steam":   round(inp.purchased_steam, 4),
        "Sold Electricity":  round(inp.sold_electricity, 4),
        "Sold Steam":        round(inp.sold_steam, 4),
        "Natural Gas":       round(inp.nat_gas, 4),
        "Coal":              round(inp.coal_sub, 4),
        "Propane":           round(inp.propane, 4),
        "Fuel Oil":          round(inp.fuel_oil_heavy_a, 4),
        "Diesel":            round(inp.diesel, 4),
        "Petrol":            round(inp.petrol, 4),
        "Biomass":           round(inp.biomass, 4),
        "Waste tires":       round(inp.waste_tires_mt, 4),
        "LPG":               round(inp.lpg, 4),
        "Other":             round(inp.other_fuels, 4),
        "Total energy":          round(out.total_energy, 4),
        "Total energy - KPI":    round(out.energy_kpi, 6),
        "Total CO2 - Scope 1":   round(out.total_co2_scope1, 4),
        "Total CO2 - Scope 2":   round(out.total_co2_scope2, 4),
        "Total CO2":             round(out.total_co2, 4),
        "Total CO2 - KPI":       round(out.co2_kpi, 6),
        # ── Waste fields ──────────────────────────────────────────────────────
        "Total Waste":           round(waste_total, 4),
        "Waste Recovered":       round(waste_recovery, 4),
        "Recovery Rate":         recovery_rate,
        # ── Country electricity placeholders (filled by _save_electricity_to_master) ─
        **{_elec_col(c): None for c in ELEC_ALL_COUNTRIES},
        # ── Derived KPIs ──────────────────────────────────────────────────────
        "Renewable_Electricity_Share_%": round(renew_share, 4),
        "Scope1_Share_%":                round(scope1_share, 4),
        "Scope2_Share_%":                round(scope2_share, 4),
        "Fossil_Energy_Share_%":         round(fossil_share, 4),
        "Water_per_ton":                 round(inp.water_withdrawals / prod, 4),
        "CO2_per_ton":                   round(out.total_co2 / prod, 4),
        "Energy_per_ton":                round(out.total_energy / prod, 4),
        "ISO_Certification_%":           round(out.pct_certified * 100, 2),
        "Waste_Recovery_Rate_%":         recovery_rate,
        "Total_Electricity_by_Country_GJ": None,  # filled after country save
    }
    return row


def _save_version_parquet(inp, combined_df: pd.DataFrame) -> str:
    """
    Save the ENTIRE company template (all years) as a Parquet snapshot.
    Stored in data_storage/versions/{CompanyName}/ — subfolder only, never flat.
    Filename: CompanyName_Year_YYYYMMDD_HHMMSS.parquet (year = the year just edited).
    NEVER overwritten — each save event creates a new file (full audit trail).
    Reading this file shows the complete state of all years for that company
    at the exact moment the save was made.
    """
    from pathlib import Path
    from datetime import datetime
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    co_safe = inp.company.replace(" ", "_").replace("/", "_")
    # Extract ALL rows for this company from the combined master DataFrame
    company_all_years = combined_df[combined_df["Company"] == inp.company].copy()
    filename = f"{co_safe}_{inp.year}_{ts}.parquet"
    # Subfolder only — no flat file
    ver_dir  = Path("data_storage") / "versions" / co_safe
    ver_dir.mkdir(parents=True, exist_ok=True)
    try:
        company_all_years.to_parquet(ver_dir / filename, index=False)
        return f"{co_safe}/{filename}"
    except Exception as e:
        return f"[version save failed: {e}]"



def _drop_zero_elec_cols(df: "pd.DataFrame") -> "pd.DataFrame":
    """
    Return df with Elec_*_GJ country columns removed if every value in that
    column is zero or null across all rows.  Non-electricity columns are
    never touched.  Used so files only carry countries with actual consumption.
    """
    elec_cols = [c for c in df.columns if c.startswith("Elec_") and c.endswith("_GJ")
                 and c != "Total_Electricity_by_Country_GJ"]
    zero_cols = [c for c in elec_cols
                 if df[c].fillna(0).eq(0).all()]
    return df.drop(columns=zero_cols) if zero_cols else df

def _sync_consolidate_excel(master_df: "pd.DataFrame") -> None:
    """
    Sync the CONSOLIDATED_DUMMY Excel (Raw Dummy data sheet) from the master wide CSV.

    The Raw Dummy data sheet stores data in long format: one row per
    (Company, Year, Row_Label). This function overwrites it completely from
    the current master wide DataFrame so that the consolidate stays in sync
    after any save from the platform.
    """
    from pathlib import Path
    from openpyxl import load_workbook

    xl_path = Path("data_storage/master/CONSOLIDATED_DUMMY_2009_2023.xlsx")
    if not xl_path.exists():
        return  # nothing to sync yet

    # Mapping: wide-CSV column  →  (Section, Row_Label)
    COL_MAP = {
        "Total no. of sites":                              ("ISO 14001",    "Total no. of sites"),
        "ISO 14001 sites":                                 ("ISO 14001",    "ISO 14001 sites"),
        "% certified sites":                               ("ISO 14001",    "% certified sites"),
        "Production":                                      ("Production",   "Production"),
        "Water intake":                                    ("Water",        "Water intake"),
        "Water intake - KPI":                              ("Water",        "Water intake - KPI"),
        "Total Electricity":                               ("Energy",       "Total Electricity"),
        "Renewable Electricity Purchased":                 ("Energy",       "Renewable Electricity Purchased"),
        "Non-Renewable Electricity Purchased":             ("Energy",       "Non-Renewable Electricity Purchased"),
        "Self-generated AND consumed electricity on-site": ("Energy",       "Self-generated AND consumed electricity on-site"),
        "Purchased Steam":                                 ("Energy",       "Purchased Steam"),
        "Sold Electricity":                                ("Energy",       "Sold Electricity"),
        "Sold Steam":                                      ("Energy",       "Sold Steam"),
        "Natural Gas":                                     ("Energy",       "Natural Gas"),
        "Coal":                                            ("Energy",       "Coal"),
        "Propane":                                         ("Energy",       "Propane"),
        "Fuel Oil":                                        ("Energy",       "Fuel Oil"),
        "Diesel":                                          ("Energy",       "Diesel"),
        "Petrol":                                          ("Energy",       "Petrol"),
        "Biomass":                                         ("Energy",       "Biomass"),
        "Waste tires":                                     ("Energy",       "Waste tires"),
        "LPG":                                             ("Energy",       "LPG"),
        "Other":                                           ("Energy",       "Other"),
        "Total energy":                                    ("Energy",       "Total energy"),
        "Total energy - KPI":                              ("Energy",       "Total energy - KPI"),
        "Total CO2 - Scope 1":                             ("CO2 emissions","Total CO2 - Scope 1"),
        "Total CO2 - Scope 2":                             ("CO2 emissions","Total CO2 - Scope 2"),
        "Total CO2":                                       ("CO2 emissions","Total CO2"),
        "Total CO2 - KPI":                                 ("CO2 emissions","Total CO2 - KPI"),
        "Total Waste":                                     ("Waste",        "Total Waste"),
        "Waste Recovered":                                 ("Waste",        "Waste Recovered"),
        "Recovery Rate":                                   ("Waste",        "Recovery Rate"),
        **{_elec_col(c): ("Energy", f"Electricity - {c}") for c in ELEC_ALL_COUNTRIES},
    }

    # Build long rows from master_df
    long_rows = []  # list of dicts: Company, Row, Year, Data, Section, Row_Label, Notes, Consistency test
    row_order = list(COL_MAP.keys())
    # Assign fixed row numbers to match what build_esg_master.py uses
    ROW_NUM = {col: i + 1 for i, col in enumerate(row_order)}

    # Pre-compute which electricity country columns have any non-zero value
    # across the whole master — only those countries get rows in the consolidate.
    active_elec_cols = {
        col for col in COL_MAP
        if col.startswith("Elec_") and col.endswith("_GJ")
        and col in master_df.columns
        and master_df[col].fillna(0).ne(0).any()
    }

    for _, wrow in master_df.sort_values(["Company", "Year"]).iterrows():
        company = wrow["Company"]
        year    = int(wrow["Year"]) if pd.notna(wrow.get("Year")) else None
        if not company or not year:
            continue
        for col, (section, label) in COL_MAP.items():
            # Skip electricity country columns that are all-zero across the dataset
            is_elec_country = col.startswith("Elec_") and col.endswith("_GJ")
            if is_elec_country and col not in active_elec_cols:
                continue
            val = wrow.get(col)
            # For an active electricity country, skip rows where this company-year is zero
            if is_elec_country and (pd.isna(val) or float(val) == 0):
                continue
            long_rows.append({
                "Company": company,
                "Row":     ROW_NUM[col],
                "Year":    year,
                "Data":    float(val) if pd.notna(val) else None,
                "Section": section,
                "Row_Label": label,
                "Notes":   None,
                "Consistency test": None,
            })

    if not long_rows:
        return

    try:
        wb = load_workbook(xl_path)
        ws = wb["Raw Dummy data"]
        # Clear existing data rows (keep header row 1)
        for r in range(2, ws.max_row + 1):
            for c in range(1, 9):
                ws.cell(r, c).value = None
        # Write new rows
        cols = ["Company", "Row", "Year", "Data", "Section", "Row_Label", "Notes", "Consistency test"]
        for i, row in enumerate(long_rows):
            for j, col in enumerate(cols):
                ws.cell(i + 2, j + 1).value = row[col]
        wb.save(xl_path)
    except (PermissionError, OSError):
        pass  # file locked — skip, master CSV is the source of truth
    except Exception:
        pass  # any other error is also non-fatal


def _sync_company_member_files(master_df: "pd.DataFrame") -> list:
    """
    Write per-company CSVs in data_storage/members/TIP/<CompanyName>/<CompanyName>_latest.csv
    from the current master wide DataFrame.
    Skips any file that is locked (e.g. open in Excel) instead of crashing.
    Returns list of company names that were skipped.
    """
    from pathlib import Path
    members_tip = Path("data_storage/members/TIP")
    skipped = []
    for company, grp in master_df.groupby("Company"):
        co_safe   = str(company).replace(" ", "_")
        co_folder = members_tip / co_safe
        co_folder.mkdir(parents=True, exist_ok=True)
        try:
            # Drop electricity country columns that are all zero for this company
            grp_clean = _drop_zero_elec_cols(grp.reset_index(drop=True))
            grp_clean.to_csv(co_folder / f"{co_safe}_latest.csv", index=False)
        except (PermissionError, OSError):
            skipped.append(str(company))
    return skipped


def _save_submission_to_csv(inp, out) -> str:
    """
    Three independent operations:

    1. MASTER CSV (data_storage/master/) — overwrite the row for this company+year.
       The master always holds the LATEST values. Second save for same company+year
       replaces the first row.

    2. VERSION Parquet (data_storage/versions/) — always ADD a new timestamped file.
       Never overwritten. Full audit trail of every save event.

    3. SYNC (after master is written):
       - CONSOLIDATED_DUMMY Excel Raw Dummy data sheet (long format)
       - Per-company CSVs in data_storage/members/TIP/<Company>/
       - TIP members aggregate CSV
    """
    import os, tempfile
    from pathlib import Path
    from datetime import datetime

    _master_cands = dl._get_csv_candidates()
    csv_path = next((p for p in _master_cands if p.exists()
                     and p.name.startswith("ESG_MASTER_WIDE_ALL_COMPANIES_")),
                    None) or Path("data_storage/master/ESG_MASTER_WIDE_ALL_COMPANIES_2009_2023.csv")
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    new_row     = pd.DataFrame([_build_master_row(inp, out)])
    master_cols = list(new_row.columns)

    def _align(df):
        """Align DataFrame to master column order: strip extras, fill missing."""
        if df.empty:
            return pd.DataFrame(columns=master_cols)
        extra = [c for c in df.columns if c not in master_cols]
        if extra:
            df = df.drop(columns=extra)
        for col in master_cols:
            if col not in df.columns:
                df[col] = None
        return df[master_cols]

    def _load_best_existing():
        """
        Load the most complete existing master DataFrame.
        Checks all candidate paths and picks the one with the most rows.
        """
        candidates = [
            csv_path,
            Path("data_storage/raw/ESG_MASTER_WIDE_ALL_COMPANIES_2009_2023.csv"),
        ]
        best = pd.DataFrame(columns=master_cols)
        for p in candidates:
            if p.exists():
                try:
                    df = pd.read_csv(p)
                    if "Company" in df.columns and "Year" in df.columns and len(df) > len(best):
                        best = df
                except PermissionError:
                    pass
                except Exception:
                    pass
        return _align(best)

    # ── 1. Build combined DataFrame ──────────────────────────────────────────
    # H1 FIX: advisory file lock held for the full read-modify-write cycle.
    # Prevents data corruption if two analysts save simultaneously.
    # Timeout=10s: if another process holds the lock and crashes, we don't
    # block forever.  The PermissionError branch below still handles Excel locks.
    lock_path = csv_path.with_suffix(".lock")
    with FileLock(str(lock_path), timeout=cfg.FILELOCK_TIMEOUT):
        existing = _load_best_existing()
        mask     = ~((existing["Company"] == inp.company) & (existing["Year"] == inp.year))
        existing = existing[mask]
        combined = pd.concat([existing, new_row], ignore_index=True)
        combined = combined.sort_values(["Company", "Year"]).reset_index(drop=True)

        n_records   = len(combined)
        n_companies = combined["Company"].nunique()

        # ── 2. Save version Parquet BEFORE touching master (audit trail first) ───
        version_filename = _save_version_parquet(inp, combined)

        # ── 3. Write master CSV, then sync all dependent files ───────────────────
        try:
            # Master CSV keeps all country columns (even all-zero) as the full schema.
            # Derived outputs (member files, TIP aggregate) strip all-zero country cols.
            combined.to_csv(csv_path, index=False)
            # Sync TIP members aggregate
            tip_master_path = Path("data_storage/members/TIP/ESG_MASTER_WIDE_TIP_MEMBERS_2009_2023.csv")
            _update_tip_members_file(csv_path, tip_master_path)
            # Sync per-company member files
            _sync_company_member_files(combined)
            # Sync CONSOLIDATED_DUMMY Excel
            _sync_consolidate_excel(combined)

            # ── Auto-add electricity-by-country year columns for new submission ──
            # If company just submitted for a year that isn't in the electricity
            # editor yet, initialize all country columns to 0 in the master CSV.
            _new_yr = inp.year
            _elec_cols_all = [c for c in combined.columns
                              if c.startswith("Elec_") and c.endswith("_GJ")]
            _co_yr_mask = (combined["Company"] == inp.company) & (combined["Year"] == _new_yr)
            if _co_yr_mask.any() and _elec_cols_all:
                for _ec in _elec_cols_all:
                    if pd.isna(combined.loc[_co_yr_mask, _ec]).all():
                        combined.loc[_co_yr_mask, _ec] = 0.0
                # Also ensure Total_Electricity_by_Country_GJ exists
                if "Total_Electricity_by_Country_GJ" not in combined.columns:
                    combined["Total_Electricity_by_Country_GJ"] = 0.0
                elif pd.isna(combined.loc[_co_yr_mask, "Total_Electricity_by_Country_GJ"]).all():
                    combined.loc[_co_yr_mask, "Total_Electricity_by_Country_GJ"] = 0.0
                # Re-write master with the zero-initialized electricity columns
                combined.to_csv(csv_path, index=False)

            # ── CRITICAL: update the in-memory globals so all pages in this
            #    session immediately see the new data without requiring a restart.
            global _CONSOLIDATED_DF, _COMPANIES, _SECTOR_DF, _USING_FALLBACK_DATA
            global HIST_YEARS, CURR_YEAR, LONG_YEARS, LONG_DATA, FUEL_MIX
            _CONSOLIDATED_DF     = combined.copy()
            _COMPANIES           = dl.get_companies(combined)
            _USING_FALLBACK_DATA = False
            try:
                cfg.refresh_year_bounds(combined)
                HIST_YEARS = cfg.hist_years()
                CURR_YEAR  = cfg.curr_year()
                LONG_YEARS = cfg.long_years()
            except Exception:
                pass
            try:
                _SECTOR_DF = dl.load_sector_aggregated(combined)
            except Exception:
                pass
            try:
                LONG_DATA, FUEL_MIX = _build_long_data()
            except Exception:
                pass
            st.cache_data.clear()

            return (f"Saved {inp.company} — {inp.year}. "
                    f"Master: {n_records} records across {n_companies} companies. "
                    f"Version: {version_filename}")
        except PermissionError:
            ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"ESG_MASTER_{inp.company.replace(' ','_')}_{inp.year}_{ts}.csv"
            backup_path = csv_path.parent / backup_name
            try:
                combined.to_csv(backup_path, index=False)
                return (
                    f"⚠️ Master file open in Excel — saved backup: **{backup_name}**\n"
                    f"Version snapshot: {version_filename}\n"
                    f"Close Excel and click Save again."
                )
            except Exception as e2:
                return f"❌ Save failed (file locked AND backup failed): {e2}"
        except Exception as e:
            return f"❌ Save failed: {e}"


def page_entry():
    """
    Submit Data — single scrolling page with all 6 sections.
    For the current/latest year shows blank fields (new entry).
    For previous years shows existing records pre-filled.
    After Submit → redirects to My Records.
    CLIENT SIDE ONLY.
    """
    company   = st.session_state.user_company
    comp_hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    all_yrs   = sorted(dl.get_years(_CONSOLIDATED_DF, company) or [])

    # ── Header: company name + year dropdown ──────────────────────────────────
    h1, h2, _ = st.columns([2, 1, 2])
    with h1:
        st.markdown(f"""
        <div style="font-size:22px;font-weight:800;color:{TEXT};margin-top:4px">
          {_html.escape(company)}</div>
        <div style="font-size:12px;color:{MUTED}">ESG KPI Data Entry</div>
        """, unsafe_allow_html=True)
    with h2:
        yr_options = sorted(list(set(all_yrs + [CURR_YEAR])), reverse=True)
        sel_yr = st.selectbox("", yr_options, key="entry_year_sel",
                              label_visibility="collapsed")

    is_new = sel_yr not in all_yrs
    if is_new:
        st.info(f"Entering new data for **{sel_yr}** — pre-filled with projected values based on last year's trend")
    else:
        st.info(f"Editing existing data for **{sel_yr}** (pre-filled from database)")

    # ── Pre-fill: existing data or projection from prior year ─────────────────
    existing = dl.get_step_data(comp_hist, sel_yr) if (comp_hist and not is_new) else {}

    if is_new and comp_hist:
        # Use most recent available year as projection baseline
        prior_yr   = max(all_yrs) if all_yrs else sel_yr - 1
        prior_data = dl.get_step_data(comp_hist, prior_yr)
        # Compute a simple linear projection: prior + avg YoY change over last 3 years
        def _projected(key, default=0.0):
            recent_yrs = sorted([y for y in all_yrs if y >= prior_yr - 3], reverse=True)
            if len(recent_yrs) >= 2:
                vals = [float(dl.get_step_data(comp_hist, y).get(key, 0)) for y in recent_yrs]
                vals = [v for v in vals if v > 0]
                if len(vals) >= 2:
                    avg_change = (vals[0] - vals[-1]) / max(len(vals) - 1, 1)
                    return max(0.0, float(vals[0]) + avg_change)
            return float(prior_data.get(key, default))
        _num = _projected
    else:
        def _num(key, default=0.0):
            return float(existing.get(key, default))

    # Use year-scoped keys so switching year always reloads from DB (no stale form state)
    _yk = f"_{sel_yr}"  # key suffix per year

    # ── Live input sections — no st.form, calculations update as you type ──────
    st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

    # ── Section 1: ISO 14001 ──────────────────────────────────────────────────
    st.markdown(f"""
    <div style="border-left:3px solid {GREEN};padding:4px 0 4px 12px;margin-bottom:12px">
      <div style="font-size:14px;font-weight:700;color:{TEXT}">
        1.  ISO 14001 Certification</div>
      <div style="font-size:11px;color:{MUTED}">Number of sites and certified sites</div>
    </div>""", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    total_sites = c1.number_input("Total no. of sites", min_value=0,
                                   value=int(_num("total_sites")), step=1, key=f"e_sites{_yk}")
    iso_sites   = c2.number_input("ISO 14001 certified sites", min_value=0,
                                   value=int(_num("iso_sites")), step=1, key=f"e_iso{_yk}")
    iso_pct = round(iso_sites / max(total_sites, 1) * 100, 1)
    c3.metric("% Certified (live)", f"{iso_pct}%")

    st.divider()

    # ── Section 2: Production ─────────────────────────────────────────────────
    st.markdown(f"""
    <div style="border-left:3px solid {CAT_CO2};padding:4px 0 4px 12px;margin-bottom:12px">
      <div style="font-size:14px;font-weight:700;color:{TEXT}">
        2.  Production Volume</div>
      <div style="font-size:11px;color:{MUTED}">Annual tire/rubber production in metric tonnes</div>
    </div>""", unsafe_allow_html=True)
    production = st.number_input("Production (metric t)", min_value=0.0,
                                  value=_num("production"), step=1000.0,
                                  format="%.0f", key=f"e_prod{_yk}")

    st.divider()

    # ── Section 3: Water ──────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="border-left:3px solid {CAT_WATER};padding:4px 0 4px 12px;margin-bottom:12px">
      <div style="font-size:14px;font-weight:700;color:{TEXT}">
        3.  Water Withdrawals</div>
      <div style="font-size:11px;color:{MUTED}">Total water intake from all sources (m³)</div>
    </div>""", unsafe_allow_html=True)
    _wc1, _wc2 = st.columns([3, 1])
    water_withdrawals = _wc1.number_input("Water withdrawals (m³)", min_value=0.0,
                                          value=_num("water_withdrawals"), step=100.0,
                                          format="%.0f", key=f"e_water{_yk}")
    _wc2.metric("Water KPI (m³/t)", f"{round(water_withdrawals / max(production, 1), 2):.2f}")

    st.divider()

    # ── Section 4: Energy ─────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="border-left:3px solid {CAT_ENERGY};padding:4px 0 4px 12px;margin-bottom:12px">
      <div style="font-size:14px;font-weight:700;color:{TEXT}">
        4.  Energy Consumption</div>
      <div style="font-size:11px;color:{MUTED}">Electricity and fuel consumption (GJ)</div>
    </div>""", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    renew_elec    = c1.number_input("Renewable elec. purchased (GJ)", min_value=0.0,
                                     value=_num("renew_elec_purchased"), step=100.0, format="%.0f", key=f"e_re{_yk}")
    nonrenew_elec = c2.number_input("Non-renewable elec. (GJ)", min_value=0.0,
                                     value=_num("nonrenew_elec_purchased"), step=100.0, format="%.0f", key=f"e_nre{_yk}")
    self_gen      = c3.number_input("Self-generated elec. (GJ)", min_value=0.0,
                                     value=_num("self_gen_elec"), step=100.0, format="%.0f", key=f"e_sg{_yk}")
    c1b, c2b, c3b = st.columns(3)
    nat_gas       = c1b.number_input("Natural Gas (GJ LHV)", min_value=0.0,
                                      value=_num("nat_gas"), step=100.0, format="%.0f", key=f"e_ng{_yk}")
    coal_sub      = c2b.number_input("Coal (GJ LHV)", min_value=0.0,
                                      value=_num("coal_sub"), step=100.0, format="%.0f", key=f"e_coal{_yk}")
    diesel        = c3b.number_input("Diesel (GJ LHV)", min_value=0.0,
                                      value=_num("diesel"), step=100.0, format="%.0f", key=f"e_diesel{_yk}")
    c1c, c2c, c3c = st.columns(3)
    biomass       = c1c.number_input("Biomass (GJ LHV)", min_value=0.0,
                                      value=_num("biomass"), step=100.0, format="%.0f", key=f"e_bio{_yk}")
    purchased_steam  = c2c.number_input("Purchased Steam (GJ)", min_value=0.0,
                                         value=_num("purchased_steam"), step=100.0, format="%.0f", key=f"e_ps{_yk}")
    sold_electricity = c3c.number_input("Sold Electricity (GJ)", min_value=0.0,
                                         value=_num("sold_electricity"), step=100.0, format="%.0f", key=f"e_se{_yk}")
    # Live energy summary
    _telec = renew_elec + nonrenew_elec + self_gen
    _te    = _telec + purchased_steam + nat_gas + coal_sub + diesel + biomass - sold_electricity
    _ekpi  = round(_te / max(production, 1), 2)
    _rs    = round((renew_elec + self_gen) / max(_telec, 1) * 100, 1)
    _ea, _eb, _ec, _ed = st.columns(4)
    _ea.metric("Total Electricity (GJ)", f"{_telec:,.0f}")
    _eb.metric("Total Energy (GJ)", f"{_te:,.0f}")
    _ec.metric("Energy KPI (GJ/t)", f"{_ekpi:.2f}")
    _ed.metric("Renewable Share", f"{_rs:.1f}%")

    st.divider()

    # ── Section 5: CO₂ ────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="border-left:3px solid {RED};padding:4px 0 4px 12px;margin-bottom:12px">
      <div style="font-size:14px;font-weight:700;color:{TEXT}">
        5.  CO₂ Emissions — Scope 2 Steam</div>
      <div style="font-size:11px;color:{MUTED}">
        Scope 1 auto-calculated from fuel inputs. Enter Scope 2 steam separately.</div>
    </div>""", unsafe_allow_html=True)
    _cc1, _cc2 = st.columns([2, 2])
    co2_scope2_steam = _cc1.number_input("CO₂ Scope 2 Steam (tCO₂)", min_value=0.0,
                                          value=_num("co2_scope2_steam"), step=10.0,
                                          format="%.0f", key=f"e_s2s{_yk}")
    from formula_engine import EF as _EF, GJ_TO_MWH as _G2M, _DEFAULT_SCOPE2_ELEC_EF as _S2EF
    _s1 = (nat_gas*_EF["Natural Gas"] + coal_sub*_EF["Coal"]
           + diesel*_EF["Diesel"] + biomass*_EF["Biomass"])
    _s2 = co2_scope2_steam + (nonrenew_elec * _G2M) * _S2EF
    _co2t = _s1 + _s2
    _cc2.metric("CO₂ KPI (t/t)", f"{round(_co2t / max(production, 1), 3):.3f}")
    _ca, _cb, _ccc = st.columns(3)
    _ca.metric("Scope 1 (tCO₂)", f"{_s1:,.0f}")
    _cb.metric("Scope 2 (tCO₂)", f"{_s2:,.0f}")
    _ccc.metric("Total CO₂ (tCO₂)", f"{_co2t:,.0f}")

    st.divider()

    # ── Section 6: Waste ──────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="border-left:3px solid {CAT_WASTE};padding:4px 0 4px 12px;margin-bottom:12px">
      <div style="font-size:14px;font-weight:700;color:{TEXT}">
        6.  Waste Management</div>
      <div style="font-size:11px;color:{MUTED}">Total waste generated and recovered (metric t)</div>
    </div>""", unsafe_allow_html=True)
    c1w, c2w = st.columns(2)
    waste_total    = c1w.number_input("Total waste (metric t)", min_value=0.0,
                                       value=_num("waste_total"), step=10.0, format="%.0f", key=f"e_wt{_yk}")
    waste_recovery = c2w.number_input("Waste recovered (metric t)", min_value=0.0,
                                       value=_num("waste_recovery"), step=10.0, format="%.0f", key=f"e_wr{_yk}")
    _wrpct = round(waste_recovery / max(waste_total, 1) * 100, 1)
    _wa, _wb, _wc = st.columns(3)
    _wa.metric("Eliminated (T)", f"{waste_total - waste_recovery:,.0f}")
    _wb.metric("Recovery Rate", f"{_wrpct:.1f}%")
    _wc.metric("Waste KPI (kg/t)", f"{round(waste_total / max(production, 1) * 1000, 1):.1f}")
    if waste_recovery > waste_total > 0:
        st.error("⚠ Waste recovered cannot exceed total waste.")

    # ── Submit button ─────────────────────────────────────────────────────────
    st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)
    submitted = st.button("✅  Submit & Save Data", type="primary",
                          use_container_width=True, key=f"entry_submit_btn{_yk}")

    if submitted:
        inp = TemplateInputs(
            company=company, year=sel_yr,
            total_sites=total_sites, iso_sites=iso_sites,
            production=production, water_withdrawals=water_withdrawals,
            renew_elec_purchased=renew_elec, nonrenew_elec_purchased=nonrenew_elec,
            self_gen_elec=self_gen, purchased_steam=purchased_steam,
            sold_electricity=sold_electricity, nat_gas=nat_gas,
            coal_sub=coal_sub, diesel=diesel, biomass=biomass,
            co2_scope2_steam=co2_scope2_steam,
            waste_total=waste_total, waste_recovery=waste_recovery,
        )
        out = calculate(inp)

        # 1. Update session state fully — both old step_data dict AND new _codata_inp
        new_step_data = {
            fld: getattr(inp, fld)
            for fld in _VALID_TEMPLATE_FIELDS
        }
        st.session_state.step_data          = new_step_data   # keeps get_current_outputs() in sync
        st.session_state["_codata_inp"]     = inp             # used by render_template_table
        st.session_state["_codata_out"]     = out
        st.session_state.reporting_company  = company
        st.session_state.reporting_year     = sel_yr
        st.session_state.template_done      = True
        st.session_state.company_setup_done = True
        st.session_state.step               = 6
        for fld in _VALID_TEMPLATE_FIELDS:
            st.session_state[fld] = getattr(inp, fld)

        # 2. Save → master CSV + parquet version + sync TIP files + updates _CONSOLIDATED_DF global
        msg = _save_submission_to_csv(inp, out)
        st.session_state["_last_save_msg"] = msg
        st.session_state.page = "my_records"
        # Clear persisted selectbox key so My Records opens on the submitted year
        st.session_state.pop("myrec_year", None)
        st.rerun()


def page_my_records():
    """
    My Records — view and save all historical KPI data.
    Shows the full template table with all 5 sheets.
    Has Submit & Save button top-right.
    Versioning (parquet) + master CSV sync on every save.
    CLIENT SIDE ONLY.
    """
    company   = st.session_state.user_company
    comp_hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    all_yrs   = sorted(dl.get_years(_CONSOLIDATED_DF, company) or [CURR_YEAR], reverse=True)

    # ── Header: title + year dropdown + Save button ───────────────────────────
    h_title, h_yr, h_btn = st.columns([3, 1, 1])
    with h_title:
        st.markdown(section_header_html(
            "My Records",
            f"{company} · Historical KPI data",
        ), unsafe_allow_html=True)
    with h_yr:
        _def_yr  = st.session_state.get("reporting_year", all_yrs[0] if all_yrs else CURR_YEAR)
        _def_idx = all_yrs.index(_def_yr) if _def_yr in all_yrs else 0
        sel_yr = st.selectbox("Year", all_yrs, index=_def_idx, key="myrec_year",
                               label_visibility="collapsed")
    with h_btn:
        save_clicked = st.button("💾  Submit & Save", type="primary",
                                  use_container_width=True, key="myrec_save_btn")

    # Show message from Submit Data redirect
    if "_last_save_msg" in st.session_state:
        st.success(f"✅ {st.session_state.pop('_last_save_msg')}")

    # ── Load data for selected year — use in-memory _CONSOLIDATED_DF  ──────────
    # _CONSOLIDATED_DF is updated in-memory by _save_submission_to_csv so it
    # always reflects the latest saved data without needing a disk re-read.
    st.session_state.reporting_company  = company
    st.session_state.reporting_year     = sel_yr

    fresh_hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    step_data  = dl.get_step_data(fresh_hist, sel_yr) if fresh_hist else {}
    valid_flds = {f.name for f in TemplateInputs.__dataclass_fields__.values()}
    clean      = {k: v for k, v in step_data.items() if k in valid_flds}

    if clean:
        for k, v in clean.items():
            st.session_state[k] = v
        inp = TemplateInputs(company=company, year=sel_yr, **clean)
        out = calculate(inp)
    else:
        inp = TemplateInputs(company=company, year=sel_yr)
        out = calculate(inp)

    # Keep both step_data dict AND _codata_inp in sync
    st.session_state.step_data          = {fld: getattr(inp, fld) for fld in _VALID_TEMPLATE_FIELDS}
    st.session_state["_codata_inp"]     = inp
    st.session_state["_codata_out"]     = out
    st.session_state.template_done      = True
    st.session_state.company_setup_done = True
    st.session_state.step               = 6

    # ── Save & sync on button click ───────────────────────────────────────────
    if save_clicked:
        msg = _save_submission_to_csv(inp, out)   # updates globals in-place
        st.success(f"✅ {msg}")
        st.rerun()   # force re-render so table shows updated values

    # ── All 5 template sheets as tabs ─────────────────────────────────────────
    tab_main, tab_elec, tab_waste, tab_qual, tab_conv = st.tabs([
        "Main Data Input",
        "Electricity by Country",
        "Waste",
        "Qualitative Data",
        "Conversion Tables",
    ])
    with tab_main: render_template_table()
    with tab_elec: render_electricity_tab()
    with tab_waste: render_waste_tab()
    with tab_qual: render_qualitative_tab()
    with tab_conv: render_conversion_tab()


# ─────────────────────────────────────────────────────────
def render_template_table():
    company  = st.session_state.get("reporting_company") or st.session_state.get("user_company") or "TIP Member Company"
    if company == "All Companies": company = "TIP Member Company"
    rep_year = st.session_state.get("reporting_year", CURR_YEAR)
    # Always reload from _CONSOLIDATED_DF so updates to any year are visible
    _hist    = _get_fresh_hist(company)

    st.markdown(f"""
    <div style="display:flex;align-items:center;justify-content:space-between;
        background:#fff;border:1px solid #E5E7EB;border-radius:10px;
        padding:18px 24px;margin-bottom:14px">
      <div>
        <div style="font-size:17px;font-weight:700;color:#0A2240;letter-spacing:-.2px">
          Tire Industry Project — Key Performance Indicators
        </div>
        <div style="font-size:26px;font-weight:800;color:#00916E;margin-top:5px;letter-spacing:-.4px">
          {_html.escape(company)}
        </div>
        <div style="font-size:12px;color:#9CA3AF;margin-top:4px">Corporate units · ESG KPI Template — {rep_year}</div>
      </div>
      <div style="text-align:right">
        <div style="font-size:11px;color:#6B7280;text-transform:uppercase;letter-spacing:.5px">Reporting year</div>
        <div style="font-size:36px;font-weight:800;color:#0A2240;line-height:1">{rep_year}</div>
        <div style="font-size:11px;color:#9CA3AF;margin-top:3px">Data range: 2009–{rep_year}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.info("Template generated from your inputs. Blue cells = company input, grey italic = auto-calculated formula.")

    # ── Data sources — all from _CONSOLIDATED_DF so saves are immediately visible ──
    hist = get_hist_outputs()   # ALL years including current year

    # For the current reporting year prefer _codata_inp (freshest — set right after save)
    if (st.session_state.get("_codata_inp") is not None and
            getattr(st.session_state["_codata_inp"], "year", None) == rep_year):
        inp = st.session_state["_codata_inp"]
        out = st.session_state["_codata_out"]
    else:
        inp, out = _load_company_year_outputs(company, rep_year)

    # Always ensure current year is in hist with the freshest values
    hist = sorted(
        [(yr, hi, ho) for yr, hi, ho in hist if yr != rep_year] + [(rep_year, inp, out)],
        key=lambda t: t[0],
    )

    ROWS = [
        ("section","ISO 14001",None,None,None),
        ("input","Total no. of sites","no.","total_sites",None),
        ("input","ISO 14001 certified sites","no.","iso_sites",None),
        ("calc","% certified sites","%",None,lambda i,o:f"{o.pct_certified*100:.1f}%"),
        ("section","Production",None,None,None),
        ("input","Production","metric t","production",None),
        ("section","Water",None,None,None),
        ("input","Water withdrawals","m³","water_withdrawals",None),
        ("calc","Water intensity KPI","m³/t",None,lambda i,o:f"{o.water_kpi:.2f}"),
        ("section","Energy",None,None,None),
        ("calc","Total Electricity","GJ",None,lambda i,o:f"{o.total_electricity:,.0f}"),
        ("input","— Renewable electricity purchased","GJ","renew_elec_purchased",None),
        ("input","— Non-renewable electricity purchased","GJ","nonrenew_elec_purchased",None),
        ("input","— Self-generated renewable on-site","GJ","self_gen_elec",None),
        ("input","Purchased Steam","GJ","purchased_steam",None),
        ("input","Sold Electricity","GJ","sold_electricity",None),
        ("input","Sold Steam","GJ","sold_steam",None),
        ("input","Natural Gas","GJ LHV","nat_gas",None),
        ("input","Coal (all types)","GJ LHV","coal_sub",None),
        ("input","Propane","GJ LHV","propane",None),
        ("input","Fuel Oil","GJ LHV","fuel_oil_heavy_a",None),
        ("input","Diesel","GJ LHV","diesel",None),
        ("input","Petrol","GJ LHV","petrol",None),
        ("input","Biomass","GJ LHV","biomass",None),
        ("input","Waste tires","metric t","waste_tires_mt",None),
        ("input","LPG","GJ LHV","lpg",None),
        ("input","Other fuels","GJ LHV","other_fuels",None),
        ("calc","TOTAL ENERGY","GJ",None,lambda i,o:f"{o.total_energy:,.0f}"),
        ("calc","Energy intensity KPI","GJ/t",None,lambda i,o:f"{o.energy_kpi:.2f}"),
        ("section","CO2 Emissions",None,None,None),
        ("input","Scope 2 — Steam","T.CO2","co2_scope2_steam",None),
        ("calc","CO2 — Natural Gas","T.CO2",None,lambda i,o:f"{o.co2_nat_gas:,.0f}"),
        ("calc","CO2 — Coal","T.CO2",None,lambda i,o:f"{o.co2_coal:,.0f}"),
        ("calc","CO2 — Propane","T.CO2",None,lambda i,o:f"{o.co2_propane:,.0f}"),
        ("calc","CO2 — Fuel Oil","T.CO2",None,lambda i,o:f"{o.co2_fuel_oil:,.0f}"),
        ("calc","CO2 — Diesel","T.CO2",None,lambda i,o:f"{o.co2_diesel:,.0f}"),
        ("calc","CO2 — Petrol","T.CO2",None,lambda i,o:f"{o.co2_petrol:,.0f}"),
        ("calc","CO2 — LPG","T.CO2",None,lambda i,o:f"{o.co2_lpg:,.0f}"),
        ("calc","TOTAL CO2 Scope 1","T.CO2",None,lambda i,o:f"{o.total_co2_scope1:,.0f}"),
        ("calc","TOTAL CO2 Scope 2","T.CO2",None,lambda i,o:f"{o.total_co2_scope2:,.0f}"),
        ("calc","TOTAL CO2 (S1+S2)","T.CO2",None,lambda i,o:f"{o.total_co2:,.0f}"),
        ("calc","CO2 intensity KPI","T.CO2/T",None,lambda i,o:f"{o.co2_kpi:.3f}"),
        ("section","Waste",None,None,None),
        ("input","Total waste generated","metric t","waste_total",None),
        ("input","Waste sent to recovery","metric t","waste_recovery",None),
        ("calc","Waste sent to elimination","metric t",None,lambda i,o:f"{o.waste_elimination:,.0f}"),
        ("calc","Recovery rate","%",None,lambda i,o:f"{o.waste_recovery_pct*100:.1f}%"),
        ("calc","Waste intensity KPI","kg/T",None,lambda i,o:f"{i.waste_total/i.production*1000:.1f}" if i.production else "—"),
    ]

    data = []
    for rdef in ROWS:
        rtype, label, unit, key, fn = rdef
        if rtype == "section":
            row = {"Indicator": f"▸ {label}", "Unit": ""}
            for yr, hi, ho in hist: row[str(yr)] = ""
            row["YoY %"] = ""
            data.append({"_type": "section", "_row": row})
            continue

        row = {"Indicator": label, "Unit": unit or ""}
        prev_num = None
        for yr, hi, ho in hist:
            v = getattr(hi, key, None) if key else None
            if v is None and fn:
                v = fn(hi, ho)
            try:
                row[str(yr)] = f"{int(round(float(v))):,}"
            except (TypeError, ValueError):
                row[str(yr)] = str(v) if v else "—"
            try:
                prev_num = float(str(v).replace(",", "").replace("%", "").replace("—", "0"))
            except:
                pass

        # YoY %: compare raw floats — avoids string formatting artifacts
        def _rv(hi, ho, k, f):
            if k:
                v = getattr(hi, k, None)
                if v is not None:
                    try: return float(v)
                    except: pass
            if f:
                raw = f(hi, ho)
                try: return float(str(raw).replace(",","").replace("%","")
                                          .replace("—","0") or "0")
                except: pass
            return None
        curr_num = _rv(inp, out, key, fn)
        prev_num = None
        if len(hist) >= 2:
            _, ph, po = hist[-2]
            prev_num  = _rv(ph, po, key, fn)
        try:
            if curr_num is not None and prev_num is not None and prev_num != 0:
                row["YoY %"] = f"{(curr_num-prev_num)/abs(prev_num)*100:+.1f}%"
            else:
                row["YoY %"] = "—"
        except:
            row["YoY %"] = "—"

        data.append({"_type": rtype, "_row": row, "_key": key, "_label": label})

    all_rows  = [d["_row"]  for d in data]
    all_types = [d["_type"] for d in data]
    df_tbl    = pd.DataFrame(all_rows)
    curr_col  = str(rep_year)

    def style_row(row, idx):
        rt = all_types[idx]
        return [
            ("background-color:#E8F5F0;color:#065F46;font-weight:800;font-size:13px;"
             "border-top:2px solid #6EE7B7;padding-top:8px;padding-bottom:8px;"
             "letter-spacing:.3px;text-transform:uppercase") if rt == "section"
            else "background-color:#DBEAFE;color:#1E40AF;font-weight:700" if (col == curr_col and rt == "input")
            else "background-color:#EFF6FF;color:#1D4ED8;font-style:italic" if (col == curr_col and rt == "calc")
            else "background-color:#F8FAFC;color:#6B7280;font-style:italic" if rt == "calc"
            else "background-color:#F0F9FF;"
            for col in df_tbl.columns
        ]

    styled     = df_tbl.style.apply(lambda row: style_row(row, row.name), axis=1)
    tbl_height = min(900, max(400, len(all_rows)*36+60))
    st.dataframe(styled, hide_index=True, height=tbl_height, use_container_width=True)
    st.markdown(f"""<div class="tbl-legend">
      <div class="tl"><div class="tl-sw" style="background:#F0F9FF;border-color:#BAE6FD"></div>Company input (historical)</div>
      <div class="tl"><div class="tl-sw" style="background:#DBEAFE;border-color:#93C5FD"></div>Company input ({rep_year})</div>
      <div class="tl"><div class="tl-sw" style="background:#EFF6FF;border-color:#A5B4FC"></div>Auto-calculated ({rep_year})</div>
      <div class="tl"><div class="tl-sw" style="background:#F8FAFC;border-color:#E5E7EB"></div>Auto-calculated (historical)</div>
    </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# ELECTRICITY TAB
# ─────────────────────────────────────────────────────────
def _update_tip_members_file(master_path: "Path", tip_master_path: "Path") -> None:
    """Rebuild the TIP members aggregate strictly from the latest master on disk.

    This prevents mismatches where the in-memory combined_df used during save
    (bootstrap/reconstruction) differs from the finally-written master CSV.
    """
    import pandas as pd
    tip_master_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        master_df = pd.read_csv(master_path)
    except Exception as e:
        _log.error("[tip_members] Could not read master to rebuild tip members: %s", e)
        return

    try:
        _drop_zero_elec_cols(master_df).to_csv(tip_master_path, index=False)
    except Exception as e:
        _log.error("[tip_members] Could not write TIP members file: %s", e)



def _save_electricity_to_master(company: str, year: int) -> str:
    """
    Save electricity-by-country data (from the Electricity tab editor) into:
      1. Master wide CSV  — columns Elec_<Country>_GJ  (GJ = MWh x 3.6)
      2. TIP members aggregate CSV
      3. Per-company member CSVs in data_storage/members/TIP/<Company>/
      4. CONSOLIDATED_DUMMY Excel (Raw Dummy data sheet, long format)
      5. Parquet snapshot of the complete company+year row

    Updates ALL years that have non-zero values in the electricity editor.
    Only the 7 countries already in the master schema are written:
        Canada, Mexico, United States, Japan, France, Hungary, Italy
    Any other country rows in the editor UI are displayed but not persisted.
    """
    from pathlib import Path
    from datetime import datetime

    COUNTRY_COL = ELEC_COUNTRY_COLS  # all 31 countries
    MWH_TO_GJ = 3.6

    elec_df = st.session_state.get("elec_data", pd.DataFrame())
    if elec_df.empty:
        return "No electricity data entered yet."

    _ecands = dl._get_csv_candidates()
    csv_path = next((p for p in _ecands if p.exists()
                     and p.name.startswith("ESG_MASTER_WIDE_ALL_COMPANIES_")), None)
    if csv_path is None:
        return "Master CSV not found. Save KPI data first."
    try:
        master = pd.read_csv(csv_path)
    except PermissionError:
        return "Master CSV is open in Excel — close it and try again."

    # Ensure all country columns exist in master (add if missing)
    for col in COUNTRY_COL.values():
        if col not in master.columns:
            master[col] = None
    if "Total_Electricity_by_Country_GJ" not in master.columns:
        master["Total_Electricity_by_Country_GJ"] = None

    yr_cols = [c for c in elec_df.columns if str(c).isdigit() and 2000 < int(c) < 2030]

    updated_years = []
    for yr_str in yr_cols:
        yr   = int(yr_str)
        mask = (master["Company"] == company) & (master["Year"] == yr)
        if not mask.any():
            # KPI row not found for this year — create a minimal stub row so
            # electricity data is not lost; user can submit KPIs later.
            stub = pd.DataFrame([{
                "Company": company, "Year": yr,
                **{c: 0.0 for c in COUNTRY_COL.values()},
                "Total_Electricity_by_Country_GJ": 0.0,
            }])
            # Align to master columns
            for col in master.columns:
                if col not in stub.columns:
                    stub[col] = None
            stub = stub[master.columns]
            master = pd.concat([master, stub], ignore_index=True)
            master = master.sort_values(["Company", "Year"]).reset_index(drop=True)
            mask = (master["Company"] == company) & (master["Year"] == yr)

        year_series = elec_df.set_index("Country")[yr_str]
        for country, mwh_val in year_series.items():
            col_name = COUNTRY_COL.get(str(country))
            if col_name is None:
                continue  # country not in master schema
            gj_val = float(mwh_val) * MWH_TO_GJ if pd.notna(mwh_val) else 0.0
            master.loc[mask, col_name] = round(gj_val, 4)

        # Recompute total-by-country for this row
        country_vals = [master.loc[mask, c].values[0]
                        for c in COUNTRY_COL.values() if c in master.columns]
        master.loc[mask, "Total_Electricity_by_Country_GJ"] = round(
            sum(v for v in country_vals if pd.notna(v)), 4)

        updated_years.append(yr)

    if not updated_years:
        return f"No KPI rows found for {company}. Save KPI data first."

    try:
        # H1 FIX: use the same advisory lock as the KPI save path
        lock_path = csv_path.with_suffix(".lock")
        with FileLock(str(lock_path), timeout=cfg.FILELOCK_TIMEOUT):
            master.to_csv(csv_path, index=False)
    except PermissionError:
        return "Master CSV is open in Excel — close it and try again."

    # Sync all dependent files
    try:
        _sp = csv_path.stem.split("_"); _ys, _ye = _sp[-2], _sp[-1]
    except Exception:
        _ys, _ye = str(cfg.DATA_YEAR_START), str(cfg.DATA_YEAR_END)
    tip_master_path = Path(f"data_storage/members/TIP/ESG_MASTER_WIDE_TIP_MEMBERS_{_ys}_{_ye}.csv")
    _update_tip_members_file(csv_path, tip_master_path)
    _sync_company_member_files(master)
    _sync_consolidate_excel(master)

    # Parquet snapshot
    co_safe  = company.replace(" ", "_").replace("/", "_")
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    ver_dir  = Path("data_storage") / "versions" / co_safe
    ver_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{co_safe}_{year}_elec_{ts}.parquet"
    single_row = master[(master["Company"] == company) & (master["Year"] == year)].copy()
    try:
        single_row.to_parquet(ver_dir / filename, index=False)
    except Exception:
        filename = "[parquet skipped]"

    return (f"Electricity saved — {len(updated_years)} year(s) updated "
            f"({min(updated_years)}-{max(updated_years)}) converted MWh to GJ. "
            f"Consolidate + member files synced. "
            f"Snapshot: versions/{co_safe}/{filename}")


def render_electricity_tab():
    """
    Electricity-by-country editor.

    Fix for two bugs:
    1. VALUES RESET BUG — st.data_editor with a static key causes Streamlit to
       discard edits on the first rerun. Fix: never use a static key on the
       data_editor when its underlying data comes from session_state. Instead
       read the widget result back via `on_change` / direct assignment and
       give the editor a key that is stable only within one company+year session,
       so it re-initialises exactly when the company or year changes.

    2. PRE-LOAD BUG — elec_data was always initialised to zeros even when the
       master CSV already had non-zero Elec_*_GJ values for this company.
       Fix: on first load (or when company/year changes) read Elec_*_GJ cols
       from _CONSOLIDATED_DF, convert GJ→MWh, and populate the editor.
    """
    # Countries shown in the UI (all 31 — display-only for countries not in master schema)
    ELEC_COUNTRIES = ELEC_ALL_COUNTRIES  # module-level list of all 31
    COUNTRY_COL_GJ = ELEC_COUNTRY_COLS  # all 31 countries stored in master
    GJ_TO_MWH = 1.0 / 3.6

    company  = st.session_state.get("reporting_company") or st.session_state.get("user_company", "")
    rep_year = st.session_state.get("reporting_year", CURR_YEAR)

    # Build YEARS dynamically: always include the latest submitted year + any
    # year the company has data for. This ensures 2024/2025 columns appear
    # automatically when the company submits for those years.
    _co_yrs = []
    if not _CONSOLIDATED_DF.empty and company:
        _co_yrs = dl.get_years(_CONSOLIDATED_DF, company) or []
    # Always go up to the reporting year (rep_year)
    _max_yr = max([rep_year] + (_co_yrs or [2023]))
    YEARS = list(range(2009, _max_yr + 1))

    # ── Key that tracks which company+year the editor was last initialised for ──
    # When this changes we rebuild elec_data from the master so the editor
    # always shows what is actually stored in the DB.
    load_key = f"{company}|{rep_year}"
    needs_reload = st.session_state.get("_elec_load_key") != load_key

    if needs_reload:
        # Build base DataFrame of zeros
        rows = [{"Country": c, "Unit": "MWh", **{str(yr): 0.0 for yr in YEARS}}
                for c in ELEC_COUNTRIES]
        df = pd.DataFrame(rows)

        # Pre-populate from master CSV for countries that are stored
        if not _CONSOLIDATED_DF.empty and company:
            for country, col_gj in COUNTRY_COL_GJ.items():
                if col_gj not in _CONSOLIDATED_DF.columns:
                    continue
                co_df = _CONSOLIDATED_DF[_CONSOLIDATED_DF["Company"] == company]
                for _, mrow in co_df.iterrows():
                    yr = int(mrow["Year"]) if pd.notna(mrow.get("Year")) else None
                    if yr is None or yr < 2009:
                        continue
                    # Extend YEARS list if master has data for a year not yet in YEARS
                    if yr not in YEARS:
                        YEARS.append(yr)
                        df[str(yr)] = 0.0
                    gj_val = mrow.get(col_gj)
                    if pd.notna(gj_val) and float(gj_val) != 0:
                        mwh_val = round(float(gj_val) * GJ_TO_MWH, 2)
                        idx = df.index[df["Country"] == country]
                        if len(idx):
                            df.loc[idx[0], str(yr)] = mwh_val

        # Ensure all year columns are numeric (avoid object dtype after assignment)
        for yr in YEARS:
            df[str(yr)] = pd.to_numeric(df[str(yr)], errors="coerce").fillna(0.0)

        st.session_state.elec_data     = df
        st.session_state._elec_load_key = load_key
        # Drop the old widget key so Streamlit re-renders a fresh editor
        if "_elec_editor_key_idx" not in st.session_state:
            st.session_state._elec_editor_key_idx = 0
        st.session_state._elec_editor_key_idx += 1

    # ── Editor key: unique per company+year so Streamlit does not reuse ───────
    # the old internal widget state (which is what causes edits to be lost).
    editor_key = f"elec_editor_{st.session_state.get('_elec_editor_key_idx', 0)}"

    st.markdown("#### Non-Renewable Electricity Purchased by Country")


    col_cfg = {
        "Country": st.column_config.TextColumn("Country", disabled=True, width="medium"),
        "Unit":    st.column_config.TextColumn("Unit",    disabled=True, width="small"),
    }
    for yr in YEARS:
        col_cfg[str(yr)] = st.column_config.NumberColumn(
            str(yr), min_value=0, format="%.2f", width="small"
        )

    # Render the editor — DO NOT write its return value back to session_state
    # here; instead use the on-change callback approach via a separate Save button.
    # The data_editor return value IS the live edited state on every rerun.
    edited = st.data_editor(
        st.session_state.elec_data,
        column_config=col_cfg,
        hide_index=True,
        use_container_width=True,
        height=900,
        key=editor_key,
        # num_rows="fixed" so no row add/delete accidentally resets things
        num_rows="fixed",
    )
    # Always keep session_state in sync with what the editor returns this frame
    st.session_state.elec_data = edited

    # ── Save button ───────────────────────────────────────────────────────────
    col_a, col_b = st.columns([4, 1])
    with col_b:
        if st.button("💾 Save electricity data", type="primary", key="elec_save_btn"):
            msg = _save_electricity_to_master(company, rep_year)
            if "saved" in msg.lower() or "synced" in msg.lower():
                st.success("✅ Saved successfully — added to your database.")
            else:
                st.warning(msg)

    # ── Summary metrics ───────────────────────────────────────────────────────
    rep_yr_str = str(rep_year)
    total_rep  = edited[rep_yr_str].sum() if rep_yr_str in edited.columns else 0
    total_all  = sum(edited[str(yr)].sum() for yr in YEARS if str(yr) in edited.columns)
    c1, c2 = st.columns(2)
    c1.metric(f"Total — {rep_yr_str} (all countries)", f"{total_rep:,.0f} MWh")
    c2.metric("Grand total all years", f"{total_all:,.0f} MWh")




# ─────────────────────────────────────────────────────────
# WASTE TAB
# ─────────────────────────────────────────────────────────
def render_waste_tab():
    inp, out = get_current_outputs()
    hist     = get_hist_outputs()
    rep_year = st.session_state.get("reporting_year", CURR_YEAR)
    st.markdown("#### Waste KPIs — Corporate Units")
    st.caption("Total waste must equal Recovery + Elimination. The consistency check validates this.")

    WASTE_ROWS = [
        ("section","Global Information",None,None,None),
        ("input","Total no. of sites","no.","total_sites",None),
        ("input","Production","metric t","production",None),
        ("section","Waste",None,None,None),
        ("input","Total amount of waste","metric t","waste_total",None),
        ("input","Amount of waste sent to recovery","metric t","waste_recovery",None),
        ("calc","Amount of waste sent to elimination","metric t",None,lambda i,o:f"{o.waste_elimination:,.0f}"),
        ("calc","Consistency check","—",None,lambda i,o:"OK" if o.check_waste else "Error"),
        ("calc","Recovery rate","%",None,lambda i,o:f"{o.waste_recovery_pct*100:.1f}%"),
        ("calc","Waste intensity","kg/T prod",None,lambda i,o:f"{i.waste_total/i.production*1000:.2f}" if i.production else "—"),
    ]
    data = []
    for rtype, label, unit, key, fn in WASTE_ROWS:
        if rtype == "section":
            row = {"Indicator": f"▸ {label}", "Unit": ""}
            for yr, hi, ho in hist: row[str(yr)] = ""
            row[str(rep_year)] = ""; row["YoY %"] = ""
            data.append({"_type":"section","_row":row}); continue
        row = {"Indicator": label, "Unit": unit or ""}
        hist_nums = []
        for yr, hi, ho in hist:
            v = getattr(hi, key, None) if key else None
            if v is None and fn: v = fn(hi, ho)
            try:
                row[str(yr)] = f"{int(round(float(v))):,}" if isinstance(v,(int,float)) else (str(v) if v else "—")
            except (TypeError, ValueError):
                row[str(yr)] = str(v) if v is not None else "—"
            try: hist_nums.append(float(str(v).replace(",","").replace("%","").replace("—","0")))
            except: hist_nums.append(0)
        cv = getattr(inp, key, None) if key else None
        if cv is None and fn: cv = fn(inp, out)
        row[str(rep_year)] = str(cv) if cv is not None else "—"
        try:
            cn = float(str(cv).replace(",","").replace("%",""))
            pn = hist_nums[-1] if hist_nums else 0
            row["YoY %"] = f"{(cn-pn)/abs(pn)*100:+.1f}%" if pn else "—"
        except: row["YoY %"] = "—"
        data.append({"_type":rtype,"_row":row})

    all_rows  = [d["_row"]  for d in data]
    all_types = [d["_type"] for d in data]
    df_w = pd.DataFrame(all_rows)
    curr_col = str(rep_year)

    def _style_waste(row, idx):
        rt = all_types[idx]
        return [
            "background:#F0FDF8;font-weight:700;color:#065F46" if rt == "section"
            else "background:#DBEAFE;font-weight:600" if (rt == "input" and col == curr_col)
            else "background:#F0F9FF" if rt == "input"
            else "background:#F8FAFC;font-style:italic;color:#6B7280"
            for col in df_w.columns
        ]

    st.dataframe(df_w.style.apply(lambda row: _style_waste(row, row.name), axis=1),
                 hide_index=True, use_container_width=True, height=400)
    st.divider()
    c1, c2, c3 = st.columns(3)
    c1.metric("Total Waste", f"{inp.waste_total:,.0f} T")
    c2.metric("Recovery Rate", f"{out.waste_recovery_pct*100:.1f}%")
    c3.metric("Consistency", "OK" if out.check_waste else "Error")


# ─────────────────────────────────────────────────────────
# QUALITATIVE TAB
# ─────────────────────────────────────────────────────────
def render_qualitative_tab():
    st.markdown("""
    <div style="background:#F9FAFB;border:1px solid #E5E7EB;border-radius:8px;padding:14px 18px;margin-bottom:16px;font-size:13px;color:#374151;line-height:1.7">
    This section gathers qualitative data to help gain additional insights for better interpretation of your
    quantitative data. Please report your company's main programs, trends, or actions that are already
    implemented, under implementation or planned.<br>
    <span style="color:#9CA3AF;font-size:12px">Non-public information will be kept confidential and only used at an aggregated level.</span>
    </div>
    """, unsafe_allow_html=True)

    def qual_section(icon, title, questions):
        st.markdown(f"""
        <div style="background:#0A2240;color:#fff;font-size:13px;font-weight:700;
            padding:9px 16px;border-radius:8px 8px 0 0;margin-top:18px;margin-bottom:0">
          {icon} {title}
        </div>
        """, unsafe_allow_html=True)
        with st.container(border=True):
            for q_label, q_hint, q_key in questions:
                st.markdown(f"**{q_label}**")
                if q_hint: st.caption(q_hint)
                c1, c2, c3 = st.columns([2,2,1])
                with c1: st.text_area("Public information",   key=f"pub_{title}_{q_key}",   height=90, placeholder="Information for the Global KPIs Report...")
                with c2: st.text_area("Non-public (confidential)", key=f"nonpub_{title}_{q_key}", height=90, placeholder="Used only at aggregated level...")
                with c3: st.text_area("Other comments",       key=f"cmt_{title}_{q_key}",   height=90, placeholder="Any additional remarks...")
                st.divider()

    qual_section("", "Energy", [
        ("Program — Management approach", "Explain how your organization manages the energy topic: policies, commitments, ISO 50001 certifications, goals & targets.", "program"),
        ("Impacts", "Include the expected impacts related to the program initiatives. Do you expect efforts to impact the Energy KPI?", "impacts"),
        ("Specific projects completed / underway", "Report specific projects related to energy that you are currently running, implementing or planning.", "projects"),
    ])
    qual_section("", "CO2 Emissions", [
        ("Program — Management approach", "Explain how your organization manages CO2: policies, commitments, goals & targets.", "program"),
        ("Impacts", "Do you expect the efforts to positively or negatively impact the CO2 KPI?", "impacts"),
        ("Specific projects completed / underway", "Report specific projects related to CO2 reduction.", "projects"),
    ])
    qual_section("", "Water", [
        ("Program — Management approach", "Explain how your organization manages water: policies, commitments, goals & targets.", "program"),
        ("Specific projects completed / underway", "Report specific projects related to water management.", "projects"),
    ])
    qual_section("", "Waste", [
        ("Program — Management approach", "Explain how your organization manages waste: policies, commitments, goals & targets.", "program"),
        ("Specific projects completed / underway", "Report the specific projects related to waste that you are currently running.", "projects"),
    ])

    st.markdown("""<div style="background:#0A2240;color:#fff;font-size:13px;font-weight:700;
        padding:9px 16px;border-radius:8px 8px 0 0;margin-top:18px;margin-bottom:0">
      Additional Information</div>""", unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown("**Other information that may affect the five environmental KPIs**")
        st.text_area("Additional comments", key="qual_additional", height=120,
                     placeholder="e.g. major plant closures, acquisitions, production restructuring...")


# ─────────────────────────────────────────────────────────
# CONVERSION TABLES TAB
# ─────────────────────────────────────────────────────────
def render_conversion_tab():
    st.markdown("#### Unit Conversion Tables")
    st.caption("Reference factors used to normalise data to corporate units. Do not edit.")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Energy conversion factors**")
        st.dataframe(pd.DataFrame({
            "Energy Type": ["Natural Gas","Propane","LPG","Diesel","Petrol","Fuel Oil","Coal","Biomass","Waste Tires"],
            "Unit":        ["GJ LHV"]*9,
            "CO2 EF (T.CO2/GJ)": [0.0561,0.0631,0.0561,0.0741,0.0693,0.0774,0.0950,0.0,0.0475],
        }), hide_index=True, use_container_width=True)
    with col2:
        st.markdown("**Unit conversion factors**")
        st.dataframe(pd.DataFrame({
            "Indicator": ["Production","Production","Water","Energy (electric)","Energy (electric)","Waste","Waste"],
            "From unit": ["kg","lb","m³","MWh","TJ","kg","lb"],
            "To unit":   ["metric t","metric t","m³","GJ","GJ","metric t","metric t"],
            "Factor":    [0.001,0.000454,1.0,3.6,1000.0,0.001,0.000454],
        }), hide_index=True, use_container_width=True)
    st.divider()
    st.markdown("**Source:** WBCSD TIP methodology · IEA country factors (Scope 2) · IPCC 2006 Guidelines")


# ─────────────────────────────────────────────────────────
# PAGE 2 -- ANALYSIS (wired to real sector data)
# ─────────────────────────────────────────────────────────
def page_analysis():
    import plotly.graph_objects as go

    data_src = "live consolidated data" if not _CONSOLIDATED_DF.empty else "built-in demo data"
    st.markdown("## Analysis & Trends")
    st.caption(f"Sector aggregated across all TIP member companies · Source: {data_src}")

    if _USING_FALLBACK_DATA:
        st.warning("⚠️ No consolidated master CSV found — charts show illustrative fallback data.", icon=None)

    # ── Selectors: Company | Time range  (no Year — auto from range) ─────────────
    overlay_company = None
    overlay_year    = None
    _range_opts = {
        "Last 3 years":  3,  "Last 5 years":  5,  "Last 7 years":  7,
        "Last 8 years":  8,  "Last 10 years": 10, "Last 12 years": 12, "All": 0,
    }

    if st.session_state.get("is_dss", False) and not _CONSOLIDATED_DF.empty:
        companies_in_db = dl.get_companies(_CONSOLIDATED_DF) or COMPANIES
        co_options  = ["All Companies"] + companies_in_db
        sel_co_col, sel_rng_col = st.columns([3, 1])
        with sel_co_col:
            overlay_sel = st.selectbox("Company", co_options, key="analysis_overlay_co")
        with sel_rng_col:
            _range_label = st.selectbox("Time range", list(_range_opts.keys()),
                                        index=1, key="analysis_year_range")
        if overlay_sel != "All Companies":
            overlay_company = overlay_sel
    else:
        _range_label = st.selectbox("Time range", list(_range_opts.keys()),
                                    index=1, key="analysis_year_range")

    _n = _range_opts[_range_label]
    yrs_int = LONG_YEARS[-_n:] if _n else LONG_YEARS
    yrs     = [str(y) for y in yrs_int]
    # Auto-derive overlay_year as most recent year with data for selected company
    if overlay_company:
        _co_avail_yrs = dl.get_years(_CONSOLIDATED_DF, overlay_company) or [CURR_YEAR]
        _co_in_range  = [y for y in _co_avail_yrs if y in yrs_int]
        overlay_year  = max(_co_in_range) if _co_in_range else max(_co_avail_yrs)
    else:
        overlay_year = yrs_int[-1] if yrs_int else CURR_YEAR

    C = {
        "navy":"#0A2240","red":"#C8102E","green":"#00916E","blue":"#1D4ED8",
        "teal":"#0891B2","amber":"#D97706","purple":"#7C3AED","coral":"#EA580C",
        "gray":"#6B7280","grid":"#F3F4F6","bg":"#FFFFFF",
    }
    PALETTE_10 = ["#C8102E","#0A2240","#00916E","#1D4ED8","#D97706",
                  "#7C3AED","#0891B2","#EA580C","#059669","#DB2777"]

    def _layout(title="", height=300, legend_h=True, **kw):
        base = dict(
            title=dict(text=title, font=dict(size=13, color=C["navy"])),
            height=height, margin=dict(l=10,r=10,t=40,b=30),
            plot_bgcolor=C["bg"], paper_bgcolor=C["bg"],
            xaxis=dict(gridcolor=C["grid"], tickfont=dict(size=10)),
            yaxis=dict(gridcolor=C["grid"], tickfont=dict(size=10)),
            legend=dict(orientation="h" if legend_h else "v",
                        y=1.12 if legend_h else 1, font=dict(size=10)),
            hovermode="x unified",
        )
        base.update(kw)
        return base

    def _line(x, y, name, color, dash="solid", width=2, fill=None, fill_color=None, marker_size=4):
        kw = dict(x=x, y=y, name=name, mode="lines+markers",
                  line=dict(color=color, width=width, dash=dash),
                  marker=dict(size=marker_size, color=color),
                  hovertemplate="%{y:.2f}<extra>" + name + "</extra>")
        if fill:
            kw["fill"] = fill
            kw["fillcolor"] = fill_color or "rgba(128,128,128,.08)"
        return go.Scatter(**kw)

    df = _CONSOLIDATED_DF
    has_wide = (not df.empty and "Row_Label" not in df.columns)

    def _sector(col, divisor=1):
        if has_wide and col in df.columns:
            return (df.groupby("Year")[col].sum() / divisor).reindex(yrs_int)
        return None

    def _sector_mean(col):
        if has_wide and col in df.columns:
            return df.groupby("Year")[col].mean().reindex(yrs_int)
        return None

    def _co_series(company, col, divisor=1):
        if has_wide and col in df.columns:
            s = df[df["Company"]==company].set_index("Year")[col] / divisor
            return s.reindex(yrs_int)
        return None

    def _safe(series, fallback):
        if series is None:
            return fallback
        result = []
        for i, v in enumerate(series.values):
            fb = fallback[i] if i < len(fallback) else (fallback[-1] if fallback else 0.0)
            try:   result.append(float(v) if (v == v and v is not None) else fb)
            except: result.append(fb)
        return result

    companies = sorted(df["Company"].unique().tolist()) if has_wide else []

    energy_total  = _safe(_sector("Total energy", 1e6),            LONG_DATA["energy"])
    co2_total     = _safe(_sector("Total CO2", 1e6),               LONG_DATA["co2"])
    scope1_total  = _safe(_sector("Total CO2 - Scope 1", 1e6),     LONG_DATA["scope1"])
    scope2_total  = _safe(_sector("Total CO2 - Scope 2", 1e6),     LONG_DATA["scope2"])
    water_total   = _safe(_sector("Water intake", 1e6),            LONG_DATA["water"])
    energy_kpi    = _safe(_sector_mean("Total energy - KPI"),      LONG_DATA["energy_kpi"])
    co2_kpi       = _safe(_sector_mean("Total CO2 - KPI"),         LONG_DATA["co2_kpi"])
    water_kpi_v   = _safe(_sector_mean("Water intake - KPI"),      [None]*len(yrs_int))
    renew_pct     = _safe(_sector_mean("Renewable_Electricity_Share_%"), LONG_DATA["renew_pct"])
    waste_recov   = _safe(_sector_mean("Waste_Recovery_Rate_%"),   LONG_DATA["waste_recov"])
    iso_cert      = _safe(_sector_mean("ISO_Certification_%"),     [93.0]*len(yrs_int))
    waste_total_v = _safe(_sector("Total Waste"),                  [v*330000 for v in LONG_DATA["prod"]])
    waste_recov_a = _safe(_sector("Waste Recovered"),              [v*280000 for v in LONG_DATA["prod"]])

    # ── Headline KPI strip — dynamic: company data or sector totals ──────────────
    def _delta(cur, prv, good_if_down=True):
        if prv and prv != 0:
            pct = (cur - prv) / abs(prv) * 100
            good = (pct < 0) == good_if_down
            arrow = "▼" if pct < 0 else "▲"
            col = "#00916E" if good else "#C8102E"
            return f'<span style="color:{col};font-size:11px">{arrow} {abs(pct):.1f}%</span>'
        return '<span style="font-size:11px;color:#9CA3AF">—</span>'

    latest_yr = yrs_int[-1] if yrs_int else CURR_YEAR
    _first_yr = yrs_int[0]  if yrs_int else CURR_YEAR - 10
    _last_yr  = yrs_int[-1] if yrs_int else CURR_YEAR

    if overlay_company:
        # Company-specific KPI boxes
        ov_inp, ov_out = _load_company_year_outputs(overlay_company, overlay_year)
        ov_rt = max(ov_inp.renew_elec_purchased + ov_inp.nonrenew_elec_purchased
                    + ov_inp.self_gen_elec, 1)
        ov_renew = ov_inp.renew_elec_purchased / ov_rt * 100

        # Prior year for delta
        ov_hist   = dl.get_company_hist(_CONSOLIDATED_DF, overlay_company)
        ov_prev_out = None
        if overlay_year - 1 in dl.get_years(_CONSOLIDATED_DF, overlay_company):
            _, ov_prev_out = _load_company_year_outputs(overlay_company, overlay_year - 1)

        def _co_delta(cur, prev_val, good_if_down=True):
            return _delta(cur, prev_val, good_if_down) if prev_val else _delta(cur, None)

        kpi_items = [
            (f"Total Energy {overlay_year}", f"{ov_inp.nat_gas + ov_inp.nonrenew_elec_purchased + ov_inp.renew_elec_purchased:.0f}",
             "GJ",
             _co_delta(ov_out.total_energy, ov_prev_out.total_energy if ov_prev_out else None)),
            (f"Total CO₂ {overlay_year}", f"{ov_out.total_co2:,.0f}", "tCO₂",
             _co_delta(ov_out.total_co2, ov_prev_out.total_co2 if ov_prev_out else None)),
            ("CO₂ Intensity",  f"{ov_out.co2_kpi:.3f}", "tCO₂/t",
             _co_delta(ov_out.co2_kpi, ov_prev_out.co2_kpi if ov_prev_out else None)),
            ("Renewable Electricity", f"{ov_renew:.1f}%", "of elec",
             _co_delta(ov_renew, None, False)),
            ("Waste Recovery", f"{ov_out.waste_recovery_pct*100:.1f}%", "of waste",
             _co_delta(ov_out.waste_recovery_pct, ov_prev_out.waste_recovery_pct if ov_prev_out else None, False)),
        ]
    else:
        # Sector aggregate KPI boxes (dynamic from real data)
        kpi_items = [
            (f"Total Energy {latest_yr}", f"{energy_total[-1]:.1f}M", "GJ",
             _delta(energy_total[-1], energy_total[-2] if len(energy_total) > 1 else None, True)),
            (f"Total CO₂ {latest_yr}", f"{co2_total[-1]:.2f}M", "tCO₂",
             _delta(co2_total[-1], co2_total[-2] if len(co2_total) > 1 else None, True)),
            ("CO₂ Intensity", f"{co2_kpi[-1]:.3f}", "tCO₂/t",
             _delta(co2_kpi[-1], co2_kpi[-2] if len(co2_kpi) > 1 else None, True)),
            ("Renewable Electricity", f"{renew_pct[-1]:.1f}%", "of elec",
             _delta(renew_pct[-1], renew_pct[-2] if len(renew_pct) > 1 else None, False)),
            ("Waste Recovery", f"{waste_recov[-1]:.1f}%", "of waste",
             _delta(waste_recov[-1], waste_recov[-2] if len(waste_recov) > 1 else None, False)),
        ]
    kpi_cols = st.columns(5)
    for i, (label, val, unit, delta_html) in enumerate(kpi_items):
        kpi_cols[i].markdown(
            f'''<div style="border:0.5px solid #E5E7EB;border-radius:8px;
                padding:12px 14px;background:#fff">
            <div style="font-size:11px;color:#6B7280;margin-bottom:3px">{label}</div>
            <div style="font-size:21px;font-weight:600;color:#0A2240;line-height:1.1">{val}</div>
            <div style="font-size:11px;color:#9CA3AF">{unit}</div>
            <div style="margin-top:3px">{delta_html}</div>
            </div>''', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Tabs — client sees own-company tabs; DSS sees all tabs ───────────────
    is_dss_user = st.session_state.get("is_dss", False)

    if is_dss_user:
        tab_gen, tab_energy, tab_co2, tab_p3, tab_p4, tab_people = st.tabs([
            "Overview",
            "Energy",
            "CO₂ Emissions",
            "Water",
            "Waste & Environment",
            "People & Governance",
        ])
        tab_p12 = tab_energy   # keep backward compat alias
    else:
        tab_gen, tab_energy, tab_co2, tab_p3, tab_p4, tab_people = st.tabs([
            "Overview",
            "Energy",
            "CO₂ Emissions",
            "Water",
            "Waste & Environment",
            "People & Governance",
        ])
        tab_p12 = tab_energy

    # ══════════════════════════════════════════════════════════════════════════
    # TIP CHART DESIGN SYSTEM — matches official TIP ESG report (2021-2024)
    # ══════════════════════════════════════════════════════════════════════════
    from plotly.subplots import make_subplots

    # TIP colours (official report palette)
    TC = {
        "bar_blue":   "#B8CDD9",   # light blue bars  (energy, CO2 absolute)
        "bar_blue2":  "#2D4A5A",   # dark teal bars   (scope 2, steam, SBT-none)
        "bar_beige":  "#C8B49A",   # beige bars       (water, waste)
        "bar_green":  "#7BAF74",   # green            (renewable, SBT-validated)
        "bar_orange": "#E0935A",   # orange           (other fuels)
        "bar_sand":   "#D4C5A9",   # sand             (non-renew elec background)
        "bar_commit": "#9FB8C5",   # soft blue        (SBT-committed)
        "line_dark":  "#2D4A5A",   # primary line
        "line_light": "#8FA5B5",   # secondary line
    }
    # Text colours: white on dark bars, near-black on light bars
    TXT = {
        "bar_blue":   "#2C3E50",
        "bar_blue2":  "white",
        "bar_beige":  "#2C3E50",
        "bar_green":  "white",
        "bar_orange": "white",
        "bar_sand":   "#2C3E50",
        "bar_commit": "#2C3E50",
    }

    def _tlayout(title="", h=330, r=65, show_legend=True, leg_y=-0.24):
        return dict(
            title=dict(text=f"<b>{title}</b>",
                       font=dict(size=14, color="#1C2E3F", family="Arial, sans-serif"), x=0),
            height=h, margin=dict(l=55, r=r, t=50, b=60),
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            xaxis=dict(
                showgrid=False, linecolor="#999", linewidth=1.2,
                showline=True, mirror=False,
                tickfont=dict(size=12, color="#1C2E3F", family="Arial"),
                tickangle=0,
                type="category",
            ),
            yaxis=dict(
                showgrid=True, gridcolor="rgba(0,0,0,0.07)", zeroline=False,
                showline=True, linecolor="#999", linewidth=1.2,
                tickfont=dict(size=12, color="#1C2E3F", family="Arial"),
                showticklabels=True,
                autorange=True,           # ensures top of chart has breathing room
            ),
            legend=dict(orientation="h", x=0.5, xanchor="center", y=leg_y,
                        font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
            hovermode="x unified", showlegend=show_legend,
        )

    def _y2(label=""):
        """Right Y-axis — fully visible, properly sized."""
        return dict(
            tickfont=dict(size=12, color="#1C2E3F", family="Arial"),
            showgrid=False, zeroline=False,
            showline=True, linecolor="#999", linewidth=1.2,
            showticklabels=True,
            autorange=True,
            title=dict(text=f"<b>{label}</b>" if label else "",
                       font=dict(size=12, color="#333", family="Arial")),
        )

    def _omk(col, sz=9):
        """Open-circle marker."""
        return dict(symbol="circle", size=sz, color="white",
                    line=dict(color=col, width=2))


    def _dual(xs, bv, bl, bc, lv, ll, lc, title="", h=330,
              bfmt=".1f", lfmt=".2f", byt="", lyt=""):
        """Dual-axis bar (left) + line (right).
        TIP report style:
        - Bar value centred vertically in the MIDDLE of the bar (clearly readable)
        - Line value alternates top/bottom near marker (never inside bar)
        - Right Y-axis fully visible; wide right margin prevents label clipping
        """
        fig = make_subplots(specs=[[{"secondary_y": True}]])

        # ── Bar trace — value centred in bar middle ────────────────────────────
        max_bv = max((v for v in bv if v is not None and v == v), default=1) or 1
        bar_texts = [
            f"<b>{v:{bfmt}}</b>" if (v is not None and v == v and abs(v) / max_bv > 0.12)
            else ""
            for v in bv
        ]
        fig.add_trace(go.Bar(
            x=xs, y=bv, name=bl, marker_color=bc, marker_line_width=0, width=0.5,
            text=bar_texts,
            textposition="inside",
            insidetextanchor="middle",          # ← centred in the bar
            textfont=dict(size=14, color="#1C2E3F", family="Arial, sans-serif"),
            hovertemplate=f"{bl}: %{{y:{bfmt}}}<extra></extra>",
            cliponaxis=False,
        ), secondary_y=False)

        # ── Line trace — values alternate above/below marker ─────────────────
        n = len(lv)
        line_texts = [f"{v:{lfmt}}" if v is not None and v == v else "" for v in lv]
        text_pos   = ["top center" if i % 2 == 0 else "bottom center" for i in range(n)]
        fig.add_trace(go.Scatter(
            x=xs, y=lv, name=ll, mode="lines+markers+text",
            line=dict(color=lc, width=2.5), marker=_omk(lc, 10),
            text=line_texts,
            textposition=text_pos,
            textfont=dict(size=12, color="#1C2E3F", family="Arial, sans-serif"),
            hovertemplate=f"{ll}: %{{y:{lfmt}}}<extra></extra>",
            cliponaxis=False,
        ), secondary_y=True)

        lay = _tlayout(title, h, r=100)
        # Add 20% headroom above max bar so bar labels are never clipped
        valid_bv = [v for v in bv if v is not None and v == v]
        if valid_bv:
            lay["yaxis"]["range"] = [0, max(valid_bv) * 1.22]
            lay["yaxis"].pop("autorange", None)
        lay["yaxis"]["title"] = dict(
            text=f"<b>{byt}</b>" if byt else "",
            font=dict(size=12, color="#333", family="Arial, sans-serif"),
        )
        lay["yaxis2"] = _y2(lyt)
        lay["margin"]["t"] = 55
        lay["margin"]["r"] = 115
        fig.update_layout(**lay)
        fig.update_yaxes(showticklabels=True, showline=True, linecolor="#999")
        return fig

    def _stack100(xs, traces, title="", h=330):
        """100% stacked bar. Text colour chosen per bar colour. Larger, bold labels."""
        fig = go.Figure()
        _dark = ("#2D4A5A", "#7BAF74", "#E0935A", "#9FB8C5")
        for (vals, lbl, bc) in traces:
            tc_col = "white" if bc in _dark else "#1C2E3F"
            fig.add_trace(go.Bar(
                x=xs, y=vals, name=lbl, marker_color=bc, marker_line_width=0,
                text=[f"<b>{v:.1f}%</b>" if v and v > 6 else "" for v in vals],
                textposition="inside",
                insidetextanchor="middle",
                textfont=dict(size=12, color=tc_col, family="Arial, sans-serif"),
                hovertemplate=f"{lbl}: %{{y:.1f}}%<extra></extra>",
            ))
        lay = _tlayout(title, h)
        lay["barmode"] = "stack"
        lay["yaxis"]["ticksuffix"] = "%"
        lay["yaxis"]["range"] = [0, 100]
        fig.update_layout(**lay)
        return fig

    def _stackabs(xs, traces, title="", h=330):
        """Absolute stacked bar expressed as % (Fig 6 style). Larger, bold labels."""
        fig = go.Figure()
        for (pct_vals, lbl, bc, txt_vals) in traces:
            tc_col = "white" if bc in ("#2D4A5A","#7BAF74","#E0935A") else "#1C2E3F"
            # Make text bold
            bold_txts = [f"<b>{t}</b>" if t else "" for t in txt_vals]
            fig.add_trace(go.Bar(
                x=xs, y=pct_vals, name=lbl, marker_color=bc, marker_line_width=0,
                text=bold_txts, textposition="inside",
                insidetextanchor="middle",
                textfont=dict(size=12, color=tc_col, family="Arial, sans-serif"),
                hovertemplate=f"{lbl}: %{{y:.1f}}%<extra></extra>",
            ))
        lay = _tlayout(title, h)
        lay["barmode"] = "stack"
        lay["yaxis"]["ticksuffix"] = "%"
        lay["yaxis"]["range"] = [0, 100]
        fig.update_layout(**lay)
        return fig

    def _dline(xs, s1v, s1l, s1c, s2v, s2l, s2c, title="", h=300,
               s1f=".1f", s2f=".1f", yt="", s2yt="", right_y=False):
        """Dual-line chart with open-circle markers. Larger labels, alternating positions."""
        def _line_txt(vals, fmt):
            return [f"{v:{fmt}}" if v is not None and v == v else "" for v in vals]

        def _alt_pos(n):
            return ["top center" if i % 2 == 0 else "bottom center" for i in range(n)]

        n = len(xs)
        if right_y:
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Scatter(x=xs, y=s1v, name=s1l,
                mode="lines+markers+text",
                line=dict(color=s1c, width=2.5), marker=_omk(s1c, 10),
                text=_line_txt(s1v, s1f),
                textposition=_alt_pos(n),
                textfont=dict(size=11, color="#1C2E3F", family="Arial, sans-serif"),
            ), secondary_y=False)
            fig.add_trace(go.Scatter(x=xs, y=s2v, name=s2l,
                mode="lines+markers+text",
                line=dict(color=s2c, width=2.5), marker=_omk(s2c, 10),
                text=_line_txt(s2v, s2f),
                textposition=_alt_pos(n),
                textfont=dict(size=11, color="#1C2E3F", family="Arial, sans-serif"),
            ), secondary_y=True)
            lay = _tlayout(title, h)
            lay["yaxis"]["title"]  = dict(text=yt, font=dict(size=10, color="#666"))
            lay["yaxis"]["ticksuffix"] = "%"
            lay["yaxis2"] = _y2(s2yt)
            fig.update_layout(**lay)
        else:
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=xs, y=s1v, name=s1l,
                mode="lines+markers+text",
                line=dict(color=s1c, width=2.5), marker=_omk(s1c, 10),
                text=_line_txt(s1v, s1f),
                textposition=_alt_pos(n),
                textfont=dict(size=11, color="#1C2E3F", family="Arial, sans-serif"),
            ))
            fig.add_trace(go.Scatter(x=xs, y=s2v, name=s2l,
                mode="lines+markers+text",
                line=dict(color=s2c, width=2.5), marker=_omk(s2c, 10),
                text=_line_txt(s2v, s2f),
                textposition=_alt_pos(n),
                textfont=dict(size=11, color="#1C2E3F", family="Arial, sans-serif"),
            ))
            lay = _tlayout(title, h)
            lay["yaxis"]["title"] = dict(text=yt, font=dict(size=10, color="#666"))
            lay["yaxis"]["ticksuffix"] = "%"
            fig.update_layout(**lay)
        return fig

    def _ck(suf): return _chart_key(overlay_company or "sector", overlay_year or 0, suf)

    # ── Sector series ──────────────────────────────────────────────────────────
    prod_total   = _safe(_sector("Production", 1e6),               LONG_DATA["prod"])
    sites_total  = _safe(_sector("Total no. of sites"),            [None]*len(yrs_int))
    iso_cert     = _safe(_sector_mean("ISO_Certification_%"),      [97.0]*len(yrs_int))

    # Energy mix components (absolute GJ, then computed as %)
    _sg          = _safe(_sector("Purchased Steam"),               [1.5e7]*len(yrs_int))
    _rg          = _safe(_sector("Renewable Electricity Purchased"),[2e7]*len(yrs_int))
    _nrg         = _safe(_sector("Non-Renewable Electricity Purchased"),[1.4e8]*len(yrs_int))
    _natg        = _safe(_sector("Natural Gas"),                   [2.3e8]*len(yrs_int))
    _coal        = _safe(_sector("Coal"),                          [5e5]*len(yrs_int))
    _diesel      = _safe(_sector("Diesel"),                        [2e6]*len(yrs_int))
    _lpg         = _safe(_sector("LPG"),                           [1.5e7]*len(yrs_int))
    _bio         = _safe(_sector("Biomass"),                       [1e5]*len(yrs_int))
    _other_sum   = [c+d+l+b for c,d,l,b in zip(_coal,_diesel,_lpg,_bio)]

    # Electricity renewable % (per year)
    _te          = [max(r+n,1) for r,n in zip(_rg,_nrg)]
    _renew_pct_v = [r/t*100  for r,t in zip(_rg,_te)]
    _nonrw_pct_v = [100-v    for v in _renew_pct_v]

    # Energy mix as % of total energy
    def _fp(vals): return [v/(max(e,1)*1e6)*100 for v,e in zip(vals, energy_total)]
    _steam_pct   = _fp(_sg)
    _renew_pct2  = _fp(_rg)
    _nonrw_pct2  = _fp(_nrg)
    _natg_pct    = _fp(_natg)
    _other_pct   = _fp(_other_sum)

    # Waste
    waste_recov_pct = [r/max(t,1)*100 for r,t in zip(waste_recov_a, waste_total_v)]
    waste_elim_pct  = [100-v          for v in waste_recov_pct]
    waste_intensity = [wt/max(p*1e6,1)*1000
                       for wt,p in zip(waste_total_v, prod_total)]  # kg/t

    # Production index relative to first year
    _p0 = max(prod_total[0], 1)
    prod_idx = [v/_p0*100 for v in prod_total]

    # Social / People metrics — sourced from live data where columns exist,
    # else display an info panel (do NOT hardcode)
    _HS_EXT_COL   = "HS_External_Audit_%"   # not yet in schema
    _HS_INT_COL   = "HS_Internal_Audit_%"   # not yet in schema
    _WOMEN_BD_COL = "Women_Board_%"         # not yet in schema
    _WOMEN_TT_COL = "Women_Total_%"         # not yet in schema
    _SBT_VAL_COL  = "SBT_Validated"         # not yet in schema
    _SBT_COM_COL  = "SBT_Committed"         # not yet in schema
    _SBT_NON_COL  = "SBT_Not_Committed"     # not yet in schema
    _social_available = has_wide and all(
        c in df.columns for c in [_HS_EXT_COL, _HS_INT_COL, _WOMEN_BD_COL, _WOMEN_TT_COL])

    def _social_series(col):
        if has_wide and col in df.columns:
            return _safe(_sector_mean(col), [None]*len(yrs_int))
        return None

    _hs_ext  = _social_series(_HS_EXT_COL)
    _hs_int  = _social_series(_HS_INT_COL)
    _wb      = _social_series(_WOMEN_BD_COL)
    _wt      = _social_series(_WOMEN_TT_COL)

    def _sbt_series(col):
        if has_wide and col in df.columns:
            return _safe(_sector(""+col), [None]*len(yrs_int))
        return None

    _sbt_v = _sbt_series(_SBT_VAL_COL)
    _sbt_c = _sbt_series(_SBT_COM_COL)
    _sbt_n = _sbt_series(_SBT_NON_COL)
    _sbt_available = _sbt_v is not None and any(v is not None for v in _sbt_v)

    def _no_data_msg(metric, pathway):
        st.info(
            f"**{metric}** data is tracked in the TIP annual report under **{pathway}**. "
            "To enable this chart, add the corresponding fields to the KPI submission form "
            "and rebuild the master database.",
            icon="📊"
        )

    # ── Overview Tab ────────────────────────────────────────────────────────────
    with tab_gen:
        lbl_pfx = f"({overlay_company.split()[0]})" if overlay_company else "(Sector)"
        st.caption(f"KPI headline summary · {lbl_pfx} · {yrs_int[0]}–{yrs_int[-1]}")

        def _ov(col, divisor=1):
            if overlay_company and has_wide and col in df.columns:
                s = df[df["Company"]==overlay_company].set_index("Year")[col]/divisor
                return _safe(s.reindex(yrs_int), [None]*len(yrs_int))
            return None

        plot_energy = _ov("Total energy", 1e6) or energy_total
        plot_co2    = _ov("Total CO2",    1e6) or co2_total
        plot_water  = _ov("Water intake", 1e6) or water_total
        plot_ekpi   = _ov("Total energy - KPI") or energy_kpi
        plot_c2kpi  = _ov("Total CO2 - KPI")    or co2_kpi
        plot_wkpi   = _ov("Water intake - KPI") or water_kpi_v

        # 5 KPI cards
        def _delta(cur, prv, good_if_down=True):
            if prv and prv != 0:
                pct = (cur - prv)/abs(prv)*100
                good = (pct < 0) == good_if_down
                arrow = "▼" if pct < 0 else "▲"
                col2 = "#00916E" if good else "#C8102E"
                return f'<span style="color:{col2};font-size:11px">{arrow} {abs(pct):.1f}%</span>'
            return '<span style="font-size:11px;color:#9CA3AF">—</span>'

        latest_yr = yrs_int[-1] if yrs_int else CURR_YEAR
        _first_yr = yrs_int[0]  if yrs_int else CURR_YEAR - 10
        _last_yr  = yrs_int[-1] if yrs_int else CURR_YEAR

        def _sfmt(v, fmt, fallback="—"):
            """Safe format — returns fallback if v is None or NaN."""
            try:
                if v is None or v != v: return fallback
                return format(float(v), fmt)
            except Exception:
                return fallback

        kpi_items = [
            (f"Total Energy {latest_yr}", _sfmt(plot_energy[-1], ".1f") + "M", "GJ",
             _delta(plot_energy[-1], plot_energy[-2] if len(plot_energy)>1 else None, True)),
            (f"Total CO₂ {latest_yr}", _sfmt(plot_co2[-1], ".2f") + "M", "tCO₂",
             _delta(plot_co2[-1], plot_co2[-2] if len(plot_co2)>1 else None, True)),
            ("CO₂ Intensity", _sfmt(plot_c2kpi[-1], ".3f"), "tCO₂/t",
             _delta(plot_c2kpi[-1], plot_c2kpi[-2] if len(plot_c2kpi)>1 else None, True)),
            ("Renewable Electricity", _sfmt(renew_pct[-1], ".1f") + "%", "of elec",
             _delta(renew_pct[-1], renew_pct[-2] if len(renew_pct)>1 else None, False)),
            ("Waste Recovery", _sfmt(waste_recov[-1], ".1f") + "%", "of waste",
             _delta(waste_recov[-1], waste_recov[-2] if len(waste_recov)>1 else None, False)),
        ]
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
        st.caption("Select a pathway tab above to view TIP ESG report charts for that KPI category.")

        # Mini overview — 2 summary charts
        c1, c2 = st.columns(2)
        with c1:
            st.plotly_chart(_dual(
                yrs, plot_energy, "Energy (M GJ)", TC["bar_blue"],
                plot_ekpi, "Energy intensity (GJ/t)", TC["line_dark"],
                title="Energy consumption & intensity",
                byt="M GJ", lyt="GJ/t", bfmt=".1f", lfmt=".1f",
            ), use_container_width=True, key=_ck("ov01"))
        with c2:
            st.plotly_chart(_dual(
                yrs, plot_co2, "CO₂ (Mt)", TC["bar_blue"],
                plot_c2kpi, "CO₂ intensity (tCO₂/ton)", TC["line_dark"],
                title="CO₂ emissions & intensity",
                byt="Mt CO₂e", lyt="tCO₂/ton", bfmt=".1f", lfmt=".2f",
            ), use_container_width=True, key=_ck("ov02"))

        # Client vs sector comparison (own company only)
        if not is_dss_user:
            client_co = st.session_state.get("user_company", "")
            if has_wide and client_co and client_co in df["Company"].values:
                st.markdown("---")
                st.markdown(f"##### {client_co} — your performance vs TIP sector")
                _c1, _c2 = st.columns(2)
                with _c1:
                    _s = _co_series(client_co, "Total energy - KPI") if has_wide and "Total energy - KPI" in df.columns else None
                    if _s is not None:
                        _cv = [float(v) if not np.isnan(float(v)) else None for v in _s.values]
                        _f = go.Figure()
                        _f.add_trace(go.Scatter(x=yrs, y=energy_kpi, name="TIP sector avg",
                            mode="lines+markers", line=dict(color=TC["line_light"], width=1.8, dash="dot"),
                            marker=_omk(TC["line_light"], 7)))
                        _f.add_trace(go.Scatter(x=yrs, y=_cv, name=client_co.split()[0],
                            mode="lines+markers+text", line=dict(color=TC["line_dark"], width=2.5),
                            marker=_omk(TC["line_dark"]),
                            text=[f"{v:.2f}" if v else "" for v in _cv],
                            textposition="top center", textfont=dict(size=12, color="#1C2E3F", family="Arial")))
                        lay = _tlayout("Energy intensity vs sector (GJ/t)", 280)
                        _f.update_layout(**lay)
                        st.plotly_chart(_f, use_container_width=True, key=_ck("ov03"))
                with _c2:
                    _s2 = _co_series(client_co, "Total CO2 - KPI") if has_wide and "Total CO2 - KPI" in df.columns else None
                    if _s2 is not None:
                        _cv2 = [float(v) if not np.isnan(float(v)) else None for v in _s2.values]
                        _f2 = go.Figure()
                        _f2.add_trace(go.Scatter(x=yrs, y=co2_kpi, name="TIP sector avg",
                            mode="lines+markers", line=dict(color=TC["line_light"], width=1.8, dash="dot"),
                            marker=_omk(TC["line_light"], 7)))
                        _f2.add_trace(go.Scatter(x=yrs, y=_cv2, name=client_co.split()[0],
                            mode="lines+markers+text", line=dict(color=TC["line_dark"], width=2.5),
                            marker=_omk(TC["line_dark"]),
                            text=[f"{v:.3f}" if v else "" for v in _cv2],
                            textposition="top center", textfont=dict(size=12, color="#1C2E3F", family="Arial")))
                        lay2 = _tlayout("CO₂ intensity vs sector (tCO₂/t)", 280)
                        _f2.update_layout(**lay2)
                        st.plotly_chart(_f2, use_container_width=True, key=_ck("ov04"))

    # helper: company overlay
    def _ov_series(col, divisor=1):
        if overlay_company and has_wide and col in df.columns:
            s = df[df["Company"]==overlay_company].set_index("Year")[col]/divisor
            return _safe(s.reindex(yrs_int), [None]*len(yrs_int))
        return None

    # ── Energy Tab ──────────────────────────────────────────────────────────────
    with tab_energy:
        st.markdown("##### Energy consumption & intensity")
        st.caption("Total energy (PJNCV) · Energy intensity (GJ/t) · Energy mix · Renewable electricity share")

        c1, c2 = st.columns(2)
        with c1:
            # Dual-axis: total energy bar + energy intensity line
            plot_e = _ov_series("Total energy", 1e6) or energy_total
            _raw_ei = _ov_series("Total energy - KPI")
            plot_ei = _raw_ei if (_raw_ei and any(v is not None for v in _raw_ei)) else energy_kpi
            st.plotly_chart(_dual(
                yrs, plot_e, "Energy consumption (M GJ)", TC["bar_blue"],
                plot_ei, "Energy intensity (GJ/t)", TC["line_dark"],
                title="Total energy consumption & intensity",
                byt="M GJ (PJNCV)", lyt="GJ/t", bfmt=".1f", lfmt=".1f",
            ), use_container_width=True, key=_ck("e01"))
        with c2:
            # Production + sites (Fig 3)
            _raw_pi = _ov_series("Production", 1e6)
            plot_pi  = _raw_pi if (_raw_pi and any(v is not None for v in _raw_pi)) else prod_total
            _p0v_raw = next((v for v in plot_pi if v is not None and v == v), 1)
            _p0v = max(_p0v_raw, 1) if _p0v_raw else 1
            plot_pidx = [v/_p0v*100 if (v is not None and v == v) else None for v in plot_pi]
            st.plotly_chart(_dual(
                yrs, plot_pidx, "Production level (% rel. to first year)", TC["bar_blue"],
                sites_total, "Number of sites", TC["line_light"],
                title="Production levels & number of sites",
                byt="Production (%)", lyt="Number of sites", bfmt=".2f", lfmt=".0f",
            ), use_container_width=True, key=_ck("e02"))

        c3, c4 = st.columns(2)
        with c3:
            # Fig 5 — Energy mix 5-category stacked 100%
            traces5 = [
                (_steam_pct,   "Purchased steam",                          TC["bar_blue2"]),
                (_renew_pct2,  "Renewable electricity (purchased+self-gen)",TC["bar_green"]),
                (_nonrw_pct2,  "Non-renewable electricity purchased",       TC["bar_sand"]),
                (_natg_pct,    "Natural gas",                              "#8FA5B5"),
                (_other_pct,   "Other (LPG, fuel oil, coal, diesel, etc.)",TC["bar_orange"]),
            ]
            st.plotly_chart(_stack100(yrs, traces5, "Energy mix (%)"), use_container_width=True, key=_ck("e03"))
        with c4:
            # Fig 6 — Electricity from renewable sources
            r_texts  = [f"{v:.1f}%" if v else "" for v in _renew_pct_v]
            nr_texts = [f"{v:.1f}%" if v else "" for v in _nonrw_pct_v]
            st.plotly_chart(_stackabs(
                yrs,
                [(_renew_pct_v, "Renewable electricity (GJ)", TC["bar_blue2"], r_texts),
                 (_nonrw_pct_v, "Non-renewable electricity (GJ)", TC["bar_sand"], nr_texts)],
                "Electricity from renewable sources (%)",
            ), use_container_width=True, key=_ck("e04"))

        # When a company is selected: show that company's trend vs sector (all companies)
        if is_dss_user and overlay_company and has_wide:
            st.markdown("---")
            st.markdown(f"##### {overlay_company} — energy intensity trend vs sector")
            f_co = go.Figure()
            # Sector average line
            f_co.add_trace(go.Scatter(
                x=yrs, y=energy_kpi, name="TIP Sector Avg",
                mode="lines", line=dict(color=TC["line_light"], width=1.8, dash="dot"),
                marker=_omk(TC["line_light"], 6),
            ))
            # Selected company line
            _s_co = _co_series(overlay_company, "Total energy - KPI")
            if _s_co is not None:
                _co_vals = [float(v) if (v == v and v is not None) else None
                            for v in _s_co.reindex(yrs_int).values]
                f_co.add_trace(go.Scatter(
                    x=yrs, y=_co_vals, name=overlay_company.split()[0],
                    mode="lines+markers+text",
                    line=dict(color=TC["line_dark"], width=2.5),
                    marker=_omk(TC["line_dark"]),
                    text=[f"{v:.1f}" if v else "" for v in _co_vals],
                    textposition="top center", textfont=dict(size=12, color="#1C2E3F", family="Arial"),
                ))
            lay_co = _tlayout(
                f"Energy intensity (GJ/t) — {overlay_company.split()[0]} vs sector avg", 300)
            f_co.update_layout(**lay_co)
            st.plotly_chart(f_co, use_container_width=True, key=_ck("e05"))

    # ── CO₂ Tab ─────────────────────────────────────────────────────────────────
    with tab_co2:
        st.markdown("##### CO₂ emissions & decarbonisation")
        st.caption("Total CO₂ (Mt CO₂e) · CO₂ intensity (tCO₂/t) · Scope 1 vs Scope 2 · Science-based targets")

        c1, c2 = st.columns(2)
        with c1:
            # Fig 7 — CO₂ dual-axis
            plot_c = _ov_series("Total CO2", 1e6) or co2_total
            plot_ck = _ov_series("Total CO2 - KPI") or co2_kpi
            st.plotly_chart(_dual(
                yrs, plot_c, "CO₂ emissions (Mt CO₂e)", TC["bar_blue"],
                plot_ck, "CO₂ intensity (tCO₂/ton)", TC["line_dark"],
                title="Total CO₂ emissions & intensity",
                byt="Mt CO₂e", lyt="tCO₂/ton", bfmt=".1f", lfmt=".2f",
            ), use_container_width=True, key=_ck("c01"))
        with c2:
            # Scope 1 vs Scope 2 stacked bar
            plot_s1 = _ov_series("Total CO2 - Scope 1", 1e6) or scope1_total
            plot_s2 = _ov_series("Total CO2 - Scope 2", 1e6) or scope2_total
            f_sc = go.Figure()
            f_sc.add_trace(go.Bar(x=yrs, y=plot_s1, name="Scope 1 — direct emissions",
                marker_color=TC["bar_blue"], marker_line_width=0,
                text=[f"{v:.2f}" if v else "" for v in plot_s1],
                textposition="inside", insidetextanchor="middle", textfont=dict(size=12, color="#1C2E3F", family="Arial")))
            f_sc.add_trace(go.Bar(x=yrs, y=plot_s2, name="Scope 2 — indirect emissions",
                marker_color=TC["bar_blue2"], marker_line_width=0,
                text=[f"{v:.2f}" if v else "" for v in plot_s2],
                textposition="inside", insidetextanchor="middle", textfont=dict(size=12, color="white", family="Arial")))
            lay_sc = _tlayout("CO₂ Scope 1 vs Scope 2 (Mt CO₂e)", 330)
            lay_sc["barmode"] = "stack"
            lay_sc["yaxis"]["title"] = dict(text="Mt CO₂e", font=dict(size=9))
            f_sc.update_layout(**lay_sc)
            st.plotly_chart(f_sc, use_container_width=True, key=_ck("c02"))

        # Science-based targets (Fig 8) — only if column exists
        st.markdown("---")
        st.markdown("##### Science-based targets (SBT)")
        if _sbt_available:
            c3, c4 = st.columns(2)
            with c3:
                fig8 = go.Figure()
                for vals, lbl, bc in [
                    (_sbt_n, "Not committed",  TC["bar_blue2"]),
                    (_sbt_c, "Committed",       TC["bar_commit"]),
                    (_sbt_v, "Validated",       TC["bar_green"]),
                ]:
                    if vals and any(v is not None for v in vals):
                        fig8.add_trace(go.Bar(x=yrs, y=vals, name=lbl,
                            marker_color=bc, marker_line_width=0,
                            text=[str(int(v)) if v is not None else "" for v in vals],
                            textposition="inside", insidetextanchor="middle", textfont=dict(size=13, family="Arial",
                                color="white" if bc in (TC["bar_blue2"],TC["bar_green"]) else "#1C2E3F")))
                lay8 = _tlayout("Members with science-based targets", 310)
                lay8["barmode"] = "stack"
                lay8["yaxis"]["title"] = dict(text="Number of TIP members", font=dict(size=9))
                fig8.update_layout(**lay8)
                st.plotly_chart(fig8, use_container_width=True, key=_ck("c03"))
        else:
            _no_data_msg("Science-based targets (SBT)",
                "Operations — Manufacturing (CO₂ pathway)")

        # When a company is selected: show that company's CO₂ trend vs sector
        if is_dss_user and overlay_company and has_wide:
            st.markdown("---")
            st.markdown(f"##### {overlay_company} — CO₂ intensity trend vs sector")
            f_cr = go.Figure()
            f_cr.add_trace(go.Scatter(
                x=yrs, y=co2_kpi, name="TIP Sector Avg",
                mode="lines", line=dict(color=TC["line_light"], width=1.8, dash="dot"),
                marker=_omk(TC["line_light"], 6),
            ))
            _s_co2 = _co_series(overlay_company, "Total CO2 - KPI")
            if _s_co2 is not None:
                _co2_vals = [float(v) if (v == v and v is not None) else None
                             for v in _s_co2.reindex(yrs_int).values]
                f_cr.add_trace(go.Scatter(
                    x=yrs, y=_co2_vals, name=overlay_company.split()[0],
                    mode="lines+markers+text",
                    line=dict(color=TC["line_dark"], width=2.5),
                    marker=_omk(TC["line_dark"]),
                    text=[f"{v:.3f}" if v else "" for v in _co2_vals],
                    textposition="top center", textfont=dict(size=12, color="#1C2E3F", family="Arial"),
                ))
            lay_cr = _tlayout(
                f"CO₂ intensity (tCO₂/t) — {overlay_company.split()[0]} vs sector avg", 300)
            f_cr.update_layout(**lay_cr)
            st.plotly_chart(f_cr, use_container_width=True, key=_ck("c04"))

    # ── Water Tab ───────────────────────────────────────────────────────────────
    with tab_p3:
        st.markdown("##### Water withdrawals & intensity")
        st.caption("Total water withdrawals (M m³) · Water intensity (m³/metric t of production)")

        plot_w  = _ov_series("Water intake", 1e6) or water_total
        plot_wk = _ov_series("Water intake - KPI") or water_kpi_v

        c1, c2 = st.columns(2)
        with c1:
            st.plotly_chart(_dual(
                yrs, plot_w, "Water withdrawals (M m³)", TC["bar_beige"],
                plot_wk, "Water intensity (m³/t)", TC["line_dark"],
                title="Water withdrawals & intensity",
                byt="Million m³", lyt="m³/t", bfmt=".1f", lfmt=".1f",
            ), use_container_width=True, key=_ck("w01"))
        with c2:
            if is_dss_user and overlay_company and has_wide and companies:
                # Show selected company vs sector avg when a company is chosen
                f_wt = go.Figure()
                for i, co in enumerate(companies):
                    s = _co_series(co, "Water intake - KPI")
                    if s is not None:
                        # reindex to exactly yrs_int to avoid float x interpolation
                        vals_s = s.reindex(yrs_int)
                        vals = [float(v) if (v == v and v is not None) else None
                                for v in vals_s.values]
                        is_selected = (co == overlay_company)
                        f_wt.add_trace(go.Scatter(
                            x=yrs, y=vals, name=co.split()[0],
                            mode="lines+markers",
                            line=dict(color=PALETTE_10[i%10],
                                      width=2.5 if is_selected else 1.2,
                                      dash="solid" if is_selected else "solid"),
                            marker=dict(size=6 if is_selected else 3),
                            opacity=1.0 if is_selected else 0.45,
                        ))
                f_wt.add_trace(go.Scatter(x=yrs, y=water_kpi_v, name="Sector avg",
                    mode="lines", line=dict(color="#000", width=2, dash="dot")))
                lay_wt = _tlayout("Water intensity by company vs sector avg (m³/t)", 330, r=12)
                lay_wt["xaxis"]["type"] = "category"
                f_wt.update_layout(**lay_wt)
                st.plotly_chart(f_wt, use_container_width=True, key=_ck("w02"))
            else:
                # No company selected: show sector trend only
                f_wt2 = go.Figure()
                f_wt2.add_trace(go.Scatter(x=yrs, y=water_kpi_v, name="Sector avg water intensity",
                    mode="lines+markers+text",
                    line=dict(color=TC["line_dark"], width=2.5),
                    marker=_omk(TC["line_dark"]),
                    text=[f"{v:.1f}" if v else "" for v in water_kpi_v],
                    textposition="top center", textfont=dict(size=12, color="#1C2E3F", family="Arial"),
                ))
                lay_wt2 = _tlayout("Sector avg water intensity (m³/t)", 330, r=12)
                lay_wt2["xaxis"]["type"] = "category"
                lay_wt2["showlegend"] = False
                f_wt2.update_layout(**lay_wt2)
                st.plotly_chart(f_wt2, use_container_width=True, key=_ck("w02"))

    # ── Waste & Environment Tab ──────────────────────────────────────────────────
    with tab_p4:
        st.markdown("##### Waste management & ISO 14001 certification")
        st.caption("Total waste generated · Waste recovery vs disposal · ISO 14001 site certification")

        c1, c2 = st.columns(2)
        with c1:
            # Waste total (bar) + intensity (line)
            st.plotly_chart(_dual(
                yrs, [wt/1e6 for wt in waste_total_v], "Waste generated (Mt)", TC["bar_beige"],
                waste_intensity, "Waste intensity (kg/t)", TC["line_dark"],
                title="Total waste generated & intensity",
                byt="Mt", lyt="kg/t", bfmt=".2f", lfmt=".1f",
            ), use_container_width=True, key=_ck("wst01"))
        with c2:
            # Waste recovery vs disposal stacked 100% — percentages inside bars
            # Absolute tonnage table shown below (matches TIP report Fig 11 layout)
            st.plotly_chart(_stack100(
                yrs,
                # Order: recovery (beige, bottom) then disposal (dark, top) —
                # matches TIP report where recovery is the dominant lower section
                [(waste_recov_pct, "Sent for recovery (%)",  TC["bar_beige"]),
                 (waste_elim_pct,  "Sent for disposal (%)",  TC["bar_blue2"])],
                "Waste sent for recovery vs disposal (%)",
            ), use_container_width=True, key=_ck("wst02"))

            # ── Absolute tonnage table below chart (TIP report Fig 11 style) ──
            # Build rows: label | yr1 | yr2 | ...
            _hdr = "| Metric |" + "".join(f" {y} |" for y in yrs)
            _sep = "| --- |" + " --- |" * len(yrs)
            _rec_vals = "| Sent for recovery (t) |" + "".join(
                f" {int(v):,} |" if v else " — |" for v in waste_recov_a)
            _dis_vals = "| Sent for disposal (t) |" + "".join(
                f" {int(max(t-r,0)):,} |" if (t and r) else " — |"
                for t, r in zip(waste_total_v, waste_recov_a))
            _table_md = "\n".join([_hdr, _sep, _rec_vals, _dis_vals])
            st.markdown(_table_md, unsafe_allow_html=False)

        c3, c4 = st.columns(2)
        with c3:
            # ISO 14001 — dual line: % certified + site count
            st.plotly_chart(_dline(
                yrs, iso_cert, "ISO 14001 certified sites (%)", TC["line_dark"],
                sites_total,   "Number of sites",               TC["line_light"],
                title="ISO 14001 certification & site count",
                yt="% certified sites", s2yt="Number of sites",
                right_y=True, s1f=".0f", s2f=".0f", h=330,
            ), use_container_width=True, key=_ck("wst03"))
        with c4:
            # Waste recovery trend — multi-year progress lines per company
            f_wr = go.Figure()
            if has_wide and companies:
                for i, co in enumerate(companies):
                    s = _co_series(co, "Waste_Recovery_Rate_%")
                    if s is None: s = _co_series(co, "Recovery Rate")
                    if s is not None:
                        vals = [float(v) if (v == v and v is not None) else None
                                for v in s.reindex(yrs_int).values]
                        # Highlight selected company if one is chosen
                        is_sel = (overlay_company and co == overlay_company)
                        f_wr.add_trace(go.Scatter(
                            x=yrs, y=vals, name=co.split()[0],
                            mode="lines+markers",
                            line=dict(color=PALETTE_10[i%10],
                                      width=2.5 if is_sel else 1.2),
                            marker=dict(size=6 if is_sel else 3),
                            opacity=1.0 if (is_sel or not overlay_company) else 0.4,
                            hovertemplate=f"{co.split()[0]}: %{{y:.1f}}%<extra></extra>",
                        ))
            # Sector average
            f_wr.add_trace(go.Scatter(
                x=yrs, y=waste_recov, name="Sector avg",
                mode="lines", line=dict(color="#000", width=2, dash="dot"),
            ))
            f_wr.add_hline(y=80, line_dash="dot", line_color=TC["line_dark"],
                annotation_text="80% target", annotation_font_size=9)
            lay_wr = _tlayout("Waste recovery rate — company progress (%)", 330, r=12)
            lay_wr["yaxis"]["ticksuffix"] = "%"
            lay_wr["yaxis"]["range"] = [0, 105]
            f_wr.update_layout(**lay_wr)
            st.plotly_chart(f_wr, use_container_width=True, key=_ck("wst04"))

    # ── People & Governance Tab ──────────────────────────────────────────────────
    with tab_people:
        st.markdown("##### People — H&S audit coverage & workforce diversity")
        st.caption("Health & Safety audited sites · Women's representation in workforce and on boards")

        if _social_available:
            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(_dline(
                    yrs, _hs_ext, "Externally audited H&S (%)", TC["line_dark"],
                    _hs_int,      "Internally audited H&S (%)", TC["line_light"],
                    title="H&S system audit coverage (%)",
                    yt="Audited sites (%)", h=320, s1f=".1f", s2f=".1f",
                ), use_container_width=True, key=_ck("pp01"))
            with c2:
                st.plotly_chart(_dline(
                    yrs, _wb, "Women on Board of Directors (%)", TC["line_dark"],
                    _wt,      "Women in total employees (%)",    TC["line_light"],
                    title="Women's representation (%)",
                    yt="Women's representation (%)", h=320, s1f=".1f", s2f=".1f",
                ), use_container_width=True, key=_ck("pp02"))
        else:
            st.info(
                "H&S audit coverage and women's representation data are tracked in the "
                "TIP annual report under Impact Pathway 4 (Operations: Employees) but "
                "are not yet in the KPI submission form. "
                "Add fields HS_External_Audit_%, HS_Internal_Audit_%, "
                "Women_Board_%, Women_Total_% to the entry form to enable these charts.",
                icon="📊",
            )


# ─────────────────────────────────────────────────────────
# PAGE 3 -- BENCHMARKING
# ─────────────────────────────────────────────────────────
def _compute_industry_scores(df, year):
    """Compute sector median scores (0–100, 100=best) for the 5 TIP KPIs."""
    KPI_MAP = [
        ("Total CO2 - KPI",              True,  0.55, 0.82),
        ("Total energy - KPI",           True,  8.0,  10.5),
        ("Water intake - KPI",           True,  5.5,  9.0),
        ("Renewable_Electricity_Share_%", False, 0.0,  100.0),
        ("Waste_Recovery_Rate_%",         False, 70.0, 100.0),
    ]
    if df.empty or "Row_Label" in df.columns:
        return [50.0] * 5
    yr_df = df[df["Year"] == year]
    if yr_df.empty:
        # Try nearest year
        nearest = df["Year"].dropna().unique()
        if len(nearest):
            yr_df = df[df["Year"] == nearest[abs(nearest - year).argmin()]]
        if yr_df.empty:
            return [50.0] * 5
    scores = []
    for col, lower_better, best, worst in KPI_MAP:
        if col in yr_df.columns and yr_df[col].notna().any():
            med  = float(yr_df[col].median())
            span = abs(worst - best) or 1
            s    = ((worst - med) / span * 100 if lower_better
                    else (med - best) / span * 100)
            scores.append(round(max(0, min(100, s)), 1))
        else:
            scores.append(50.0)
    return scores


def _load_company_year_outputs(company: str, year: int):
    """
    Load inputs and compute outputs for any company+year from the consolidated DB.
    Returns (TemplateInputs, TemplateOutputs) — never falls back to session state
    (session state belongs to the logged-in client, not the selected company).
    """
    hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    if hist:
        sd = dl.get_step_data(hist, year)
        sd_clean = {k: v for k, v in sd.items() if k in _VALID_TEMPLATE_FIELDS}
        if sd_clean:
            inp = TemplateInputs(company=company, year=year, **sd_clean)
            return inp, calculate(inp)
    # Neutral fallback — do NOT use session state (that's the logged-in client's data)
    inp = TemplateInputs(company=company, year=year)
    return inp, calculate(inp)


def _compute_kpi_improvement(company: str, base_year: int, end_year: int) -> dict:
    """
    Compute improvement % for each KPI between base_year and end_year.
    Returns {kpi_attr: pct_change_string} for CO2, energy, water KPIs + raw fields.
    """
    base_inp, base_out = _load_company_year_outputs(company, base_year)
    end_inp,  end_out  = _load_company_year_outputs(company, end_year)

    def _pct(b, e):
        if b and b != 0 and e:
            return f"{(e - b) / abs(b) * 100:+.1f}%"
        return "N/A"

    renew_base = (base_inp.renew_elec_purchased + base_inp.self_gen_elec) / max(base_out.total_electricity, 1) * 100
    renew_end  = (end_inp.renew_elec_purchased  + end_inp.self_gen_elec)  / max(end_out.total_electricity,  1) * 100
    wrec_base  = base_out.waste_recovery_pct * 100
    wrec_end   = end_out.waste_recovery_pct  * 100

    return {
        "CO₂ intensity":        _pct(base_out.co2_kpi,    end_out.co2_kpi),
        "Energy intensity":     _pct(base_out.energy_kpi, end_out.energy_kpi),
        "Water intensity":      _pct(base_out.water_kpi,  end_out.water_kpi),
        "Renewable electricity":_pct(renew_base,           renew_end),
        "Waste recovery rate":  _pct(wrec_base,            wrec_end),
    }




def _chart_key(*args) -> str:
    """Unique chart key that changes with company/year selection → forces animation replay."""
    return "__".join(str(a).replace(" ","_") for a in args)


def _compute_industry_scores(df, year):
    """Compute sector median scores (0–100, 100=best) for the 5 TIP KPIs."""
    KPI_MAP = [
        ("Total CO2 - KPI",              True,  0.55, 0.82),
        ("Total energy - KPI",           True,  8.0,  10.5),
        ("Water intake - KPI",           True,  5.5,  9.0),
        ("Renewable_Electricity_Share_%", False, 0.0,  100.0),
        ("Waste_Recovery_Rate_%",         False, 70.0, 100.0),
    ]
    if df.empty or "Row_Label" in df.columns:
        return [50.0] * 5
    yr_df = df[df["Year"] == year]
    if yr_df.empty:
        # Try nearest year
        nearest = df["Year"].dropna().unique()
        if len(nearest):
            yr_df = df[df["Year"] == nearest[abs(nearest - year).argmin()]]
        if yr_df.empty:
            return [50.0] * 5
    scores = []
    for col, lower_better, best, worst in KPI_MAP:
        if col in yr_df.columns and yr_df[col].notna().any():
            med  = float(yr_df[col].median())
            span = abs(worst - best) or 1
            s    = ((worst - med) / span * 100 if lower_better
                    else (med - best) / span * 100)
            scores.append(round(max(0, min(100, s)), 1))
        else:
            scores.append(50.0)
    return scores


def _load_company_year_outputs(company: str, year: int):
    """
    Load inputs and compute outputs for any company+year from the consolidated DB.
    Returns (TemplateInputs, TemplateOutputs) — never falls back to session state
    (session state belongs to the logged-in client, not the selected company).
    """
    hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    if hist:
        sd = dl.get_step_data(hist, year)
        sd_clean = {k: v for k, v in sd.items() if k in _VALID_TEMPLATE_FIELDS}
        if sd_clean:
            inp = TemplateInputs(company=company, year=year, **sd_clean)
            return inp, calculate(inp)
    # Neutral fallback — do NOT use session state (that's the logged-in client's data)
    inp = TemplateInputs(company=company, year=year)
    return inp, calculate(inp)


def _compute_kpi_improvement(company: str, base_year: int, end_year: int) -> dict:
    """
    Compute improvement % for each KPI between base_year and end_year.
    Returns {kpi_attr: pct_change_string} for CO2, energy, water KPIs + raw fields.
    """
    base_inp, base_out = _load_company_year_outputs(company, base_year)
    end_inp,  end_out  = _load_company_year_outputs(company, end_year)

    def _pct(b, e):
        if b and b != 0 and e:
            return f"{(e - b) / abs(b) * 100:+.1f}%"
        return "N/A"

    renew_base = (base_inp.renew_elec_purchased + base_inp.self_gen_elec) / max(base_out.total_electricity, 1) * 100
    renew_end  = (end_inp.renew_elec_purchased  + end_inp.self_gen_elec)  / max(end_out.total_electricity,  1) * 100
    wrec_base  = base_out.waste_recovery_pct * 100
    wrec_end   = end_out.waste_recovery_pct  * 100

    return {
        "CO₂ intensity":        _pct(base_out.co2_kpi,    end_out.co2_kpi),
        "Energy intensity":     _pct(base_out.energy_kpi, end_out.energy_kpi),
        "Water intensity":      _pct(base_out.water_kpi,  end_out.water_kpi),
        "Renewable electricity":_pct(renew_base,           renew_end),
        "Waste recovery rate":  _pct(wrec_base,            wrec_end),
    }




def _chart_key(*args) -> str:
    """Unique chart key that changes with company/year selection → forces animation replay."""
    return "__".join(str(a).replace(" ","_") for a in args)


def _compute_industry_scores(df, year):
    """Compute sector median scores (0–100, 100=best) for the 5 TIP KPIs."""
    KPI_MAP = [
        ("Total CO2 - KPI",              True,  0.55, 0.82),
        ("Total energy - KPI",           True,  8.0,  10.5),
        ("Water intake - KPI",           True,  5.5,  9.0),
        ("Renewable_Electricity_Share_%", False, 0.0,  100.0),
        ("Waste_Recovery_Rate_%",         False, 70.0, 100.0),
    ]
    if df.empty or "Row_Label" in df.columns:
        return [50.0] * 5
    yr_df = df[df["Year"] == year]
    if yr_df.empty:
        # Try nearest year
        nearest = df["Year"].dropna().unique()
        if len(nearest):
            yr_df = df[df["Year"] == nearest[abs(nearest - year).argmin()]]
        if yr_df.empty:
            return [50.0] * 5
    scores = []
    for col, lower_better, best, worst in KPI_MAP:
        if col in yr_df.columns and yr_df[col].notna().any():
            med  = float(yr_df[col].median())
            span = abs(worst - best) or 1
            s    = ((worst - med) / span * 100 if lower_better
                    else (med - best) / span * 100)
            scores.append(round(max(0, min(100, s)), 1))
        else:
            scores.append(50.0)
    return scores


def _load_company_year_outputs(company: str, year: int):
    """
    Load inputs and compute outputs for any company+year from the consolidated DB.
    Returns (TemplateInputs, TemplateOutputs) — never falls back to session state
    (session state belongs to the logged-in client, not the selected company).
    """
    hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    if hist:
        sd = dl.get_step_data(hist, year)
        sd_clean = {k: v for k, v in sd.items() if k in _VALID_TEMPLATE_FIELDS}
        if sd_clean:
            inp = TemplateInputs(company=company, year=year, **sd_clean)
            return inp, calculate(inp)
    # Neutral fallback — do NOT use session state (that's the logged-in client's data)
    inp = TemplateInputs(company=company, year=year)
    return inp, calculate(inp)


def _compute_kpi_improvement(company: str, base_year: int, end_year: int) -> dict:
    """
    Compute improvement % for each KPI between base_year and end_year.
    Returns {kpi_attr: pct_change_string} for CO2, energy, water KPIs + raw fields.
    """
    base_inp, base_out = _load_company_year_outputs(company, base_year)
    end_inp,  end_out  = _load_company_year_outputs(company, end_year)

    def _pct(b, e):
        if b and b != 0 and e:
            return f"{(e - b) / abs(b) * 100:+.1f}%"
        return "N/A"

    renew_base = (base_inp.renew_elec_purchased + base_inp.self_gen_elec) / max(base_out.total_electricity, 1) * 100
    renew_end  = (end_inp.renew_elec_purchased  + end_inp.self_gen_elec)  / max(end_out.total_electricity,  1) * 100
    wrec_base  = base_out.waste_recovery_pct * 100
    wrec_end   = end_out.waste_recovery_pct  * 100

    return {
        "CO₂ intensity":        _pct(base_out.co2_kpi,    end_out.co2_kpi),
        "Energy intensity":     _pct(base_out.energy_kpi, end_out.energy_kpi),
        "Water intensity":      _pct(base_out.water_kpi,  end_out.water_kpi),
        "Renewable electricity":_pct(renew_base,           renew_end),
        "Waste recovery rate":  _pct(wrec_base,            wrec_end),
    }




def _chart_key(*args) -> str:
    """Unique chart key that changes with company/year selection → forces animation replay."""
    return "__".join(str(a).replace(" ","_") for a in args)


def _compute_industry_scores(df, year):
    """Compute sector median scores (0–100, 100=best) for the 5 TIP KPIs."""
    KPI_MAP = [
        ("Total CO2 - KPI",              True,  0.55, 0.82),
        ("Total energy - KPI",           True,  8.0,  10.5),
        ("Water intake - KPI",           True,  5.5,  9.0),
        ("Renewable_Electricity_Share_%", False, 0.0,  100.0),
        ("Waste_Recovery_Rate_%",         False, 70.0, 100.0),
    ]
    if df.empty or "Row_Label" in df.columns:
        return [50.0] * 5
    yr_df = df[df["Year"] == year]
    if yr_df.empty:
        # Try nearest year
        nearest = df["Year"].dropna().unique()
        if len(nearest):
            yr_df = df[df["Year"] == nearest[abs(nearest - year).argmin()]]
        if yr_df.empty:
            return [50.0] * 5
    scores = []
    for col, lower_better, best, worst in KPI_MAP:
        if col in yr_df.columns and yr_df[col].notna().any():
            med  = float(yr_df[col].median())
            span = abs(worst - best) or 1
            s    = ((worst - med) / span * 100 if lower_better
                    else (med - best) / span * 100)
            scores.append(round(max(0, min(100, s)), 1))
        else:
            scores.append(50.0)
    return scores


def _load_company_year_outputs(company: str, year: int):
    """
    Load inputs and compute outputs for any company+year from the consolidated DB.
    Returns (TemplateInputs, TemplateOutputs) — never falls back to session state
    (session state belongs to the logged-in client, not the selected company).
    """
    hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    if hist:
        sd = dl.get_step_data(hist, year)
        sd_clean = {k: v for k, v in sd.items() if k in _VALID_TEMPLATE_FIELDS}
        if sd_clean:
            inp = TemplateInputs(company=company, year=year, **sd_clean)
            return inp, calculate(inp)
    # Neutral fallback — do NOT use session state (that's the logged-in client's data)
    inp = TemplateInputs(company=company, year=year)
    return inp, calculate(inp)


def _compute_kpi_improvement(company: str, base_year: int, end_year: int) -> dict:
    """
    Compute improvement % for each KPI between base_year and end_year.
    Returns {kpi_attr: pct_change_string} for CO2, energy, water KPIs + raw fields.
    """
    base_inp, base_out = _load_company_year_outputs(company, base_year)
    end_inp,  end_out  = _load_company_year_outputs(company, end_year)

    def _pct(b, e):
        if b and b != 0 and e:
            return f"{(e - b) / abs(b) * 100:+.1f}%"
        return "N/A"

    renew_base = (base_inp.renew_elec_purchased + base_inp.self_gen_elec) / max(base_out.total_electricity, 1) * 100
    renew_end  = (end_inp.renew_elec_purchased  + end_inp.self_gen_elec)  / max(end_out.total_electricity,  1) * 100
    wrec_base  = base_out.waste_recovery_pct * 100
    wrec_end   = end_out.waste_recovery_pct  * 100

    return {
        "CO₂ intensity":        _pct(base_out.co2_kpi,    end_out.co2_kpi),
        "Energy intensity":     _pct(base_out.energy_kpi, end_out.energy_kpi),
        "Water intensity":      _pct(base_out.water_kpi,  end_out.water_kpi),
        "Renewable electricity":_pct(renew_base,           renew_end),
        "Waste recovery rate":  _pct(wrec_base,            wrec_end),
    }




def _chart_key(*args) -> str:
    """Unique chart key that changes with company/year selection → forces animation replay."""
    return "__".join(str(a).replace(" ","_") for a in args)

def page_benchmarking():
    """
    Benchmarking — KPI-topic tabs (General, CO₂, Energy, Electricity, Water, Waste).
    Client: no company selector (own company only).
    DSS+: company dropdown.
    No peer company names exposed in any chart.
    PDF download available per tab.
    """
    from pdf_report import generate_executive_pdf, build_kpi_dict_from_outputs, REPORTLAB_OK
    import io

    st.markdown(section_header_html("Benchmarking",
        "Industry peer comparison · TIP sector quartiles"), unsafe_allow_html=True)

    companies_in_db = dl.get_companies(_CONSOLIDATED_DF) or COMPANIES
    is_dss = st.session_state.get("is_dss", False)

    # ── Selectors: Company | Time range  (no Year — auto from most recent data) ──
    _b_range_opts = {
        "Last 3 years":  3,  "Last 5 years":  5,  "Last 7 years":  7,
        "Last 8 years":  8,  "Last 10 years": 10, "Last 12 years": 12, "All": 0,
    }
    if is_dss:
        default_co = (st.session_state.get("reporting_company") or
                      st.session_state.get("user_company") or companies_in_db[0])
        if default_co not in companies_in_db: default_co = companies_in_db[0]
        bc1, bc2 = st.columns([3, 1])
        with bc1:
            company = st.selectbox("Company", companies_in_db,
                                   index=companies_in_db.index(default_co),
                                   key="bench_company_dss")
        with bc2:
            _b_range_lbl = st.selectbox("Time range", list(_b_range_opts.keys()),
                                        index=1, key="bench_year_range")
    else:
        company  = st.session_state.user_company
        bc2, _   = st.columns([1, 3])
        with bc2:
            _b_range_lbl = st.selectbox("Time range", list(_b_range_opts.keys()),
                                        index=1, key="bench_year_range")

    # ── Derive rep_year automatically (most recent year in range with data) ──────
    _bn      = _b_range_opts[_b_range_lbl]
    _all_db_yrs = sorted(dl.get_years(_CONSOLIDATED_DF, company) or [CURR_YEAR])
    _avail_in_range = ([y for y in _all_db_yrs if not _bn or y >= (_all_db_yrs[-1] - _bn + 1)]
                       or _all_db_yrs)
    rep_year = max(_avail_in_range)

    # ── Load data ─────────────────────────────────────────────────────────────
    inp, out = _load_company_year_outputs(company, rep_year)
    bench_kpis = dl.get_benchmark_kpis(_CONSOLIDATED_DF, rep_year)
    renew_val  = (inp.renew_elec_purchased + inp.self_gen_elec) / max(out.total_electricity, 1) * 100
    waste_pct  = out.waste_recovery_pct * 100

    def live_bench(col, val, unit, lb):
        vals = bench_kpis[col].dropna().values if (not bench_kpis.empty and col in bench_kpis.columns) else []
        if len(vals) >= 3:
            q25, med, q75 = (float(np.percentile(vals, p)) for p in [25, 50, 75])
            lo,  hi       = float(np.percentile(vals, 10)), float(np.percentile(vals, 90))
        else:
            q25, med, q75, lo, hi = val*.85, val, val*1.15, val*.7, val*1.3
        b = BenchmarkResult(col, val, q25, med, q75, unit, lb)
        b._lo, b._hi, b._vals = lo, hi, vals
        return b

    BM = [
        live_bench("Total CO2 - KPI",              out.co2_kpi,  "tCO₂/t", True),
        live_bench("Total energy - KPI",            out.energy_kpi,"GJ/t",   True),
        live_bench("Water intake - KPI",            out.water_kpi, "m³/t",   True),
        live_bench("Renewable_Electricity_Share_%", renew_val,     "%",      False),
        live_bench("Waste_Recovery_Rate_%",         waste_pct,     "%",      False),
    ]
    KPI_NAMES = ["CO₂ Intensity","Energy Intensity","Water Intensity","Renewable Elec.","Waste Recovery"]
    KPI_COLORS = [CAT_CO2, CAT_ENERGY, CAT_WATER, CAT_RENEW, CAT_WASTE]
    for b, n in zip(BM, KPI_NAMES): b.kpi_name = n

    # ── Enhanced KPI summary boxes with position bar ──────────────────────────
    chip_cols = st.columns(5)
    for i, (b, color) in enumerate(zip(BM, KPI_COLORS)):
        # Position within sector range as 0–100
        rng = max(b._hi - b._lo, 0.001)
        pos = (b.company_value - b._lo) / rng   # 0=best for lb, 1=worst
        pos_pct = (1 - pos) * 100 if b.lower_is_better else pos * 100  # 100=best always
        pos_pct = max(0, min(100, pos_pct))
        rank_col = GREEN if pos_pct >= 70 else (AMBER if pos_pct >= 40 else RED)
        rank_lbl = ("Top quartile" if pos_pct >= 75 else
                    "Above median" if pos_pct >= 50 else
                    "Below median" if pos_pct >= 25 else "Bottom quartile")
        with chip_cols[i]:
            st.markdown(f"""
            <div style="background:#fff;border:1px solid {BORDER};border-radius:8px;
                padding:12px;animation:tipFadeIn 400ms ease-out {i*60}ms both">
              <div style="font-size:9.5px;color:{MUTED};font-weight:600;text-transform:uppercase;
                  letter-spacing:.5px;margin-bottom:6px">{b.kpi_name}</div>
              <div style="font-size:22px;font-weight:700;color:{color};
                  font-variant-numeric:tabular-nums;line-height:1">{b.company_value:.2f}</div>
              <div style="font-size:9px;color:{MUTED};margin-bottom:8px">{b.unit}</div>
              <div style="background:#F1F5F9;border-radius:4px;height:5px;overflow:hidden;margin-bottom:5px">
                <div style="background:{rank_col};width:{pos_pct:.0f}%;height:100%;border-radius:4px;
                    transition:width 1s ease"></div>
              </div>
              <div style="display:flex;justify-content:space-between;font-size:9px;color:{MUTED}">
                <span>{"Worst" if b.lower_is_better else "Low"}</span>
                <span style="color:{rank_col};font-weight:600">{rank_lbl}</span>
                <span>{"Best" if b.lower_is_better else "High"}</span>
              </div>
            </div>""", unsafe_allow_html=True)

    st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)

    # ── Helpers ────────────────────────────────────────────────────────────────
    comp_hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    all_years = sorted(dl.get_years(_CONSOLIDATED_DF, company) or [rep_year])
    from formula_engine import TemplateInputs as TI, calculate as calc
    valid_flds = {f.name for f in TI.__dataclass_fields__.values()}

    def _co_trend(field_key):
        """Return dict year→value for a computed KPI across all years."""
        result = {}
        for y in all_years:
            sd = dl.get_step_data(comp_hist, y)
            sc = {k: v for k, v in sd.items() if k in valid_flds}
            if not sc: continue
            o  = calc(TI(company=company, year=y, **sc))
            ii = TI(company=company, year=y, **sc)
            rt = max(ii.renew_elec_purchased + ii.nonrenew_elec_purchased + ii.self_gen_elec, 1)
            result[y] = {
                "co2_kpi": o.co2_kpi, "energy_kpi": o.energy_kpi,
                "water_kpi": o.water_kpi, "waste_pct": o.waste_recovery_pct*100,
                "renew_pct": ii.renew_elec_purchased / rt * 100,
                "nonrenew_pct": ii.nonrenew_elec_purchased / rt * 100,
                "renew_gj": ii.renew_elec_purchased,
                "nonrenew_gj": ii.nonrenew_elec_purchased,
                "nat_gas": ii.nat_gas, "coal": ii.coal_sub,
                "diesel": ii.diesel, "biomass": ii.biomass,
                "water_m3": ii.water_withdrawals,
                "waste_total": ii.waste_total, "waste_rec": ii.waste_recovery,
                "scope1": o.total_co2_scope1, "scope2": o.total_co2_scope2,
            }
        return result

    trend   = _co_trend(None)
    _all_ys = sorted(trend.keys())
    ys = _all_ys[-_bn:] if _bn else _all_ys

    def _sector_series(col):
        """Sector mean, p25, p75 by year."""
        if _CONSOLIDATED_DF.empty or col not in _CONSOLIDATED_DF.columns:
            return {}, {}, {}
        grp = _CONSOLIDATED_DF.groupby("Year")[col]
        return (grp.mean().to_dict(), grp.quantile(.25).to_dict(),
                grp.quantile(.75).to_dict())

    def _anon_scatter(col, your_val, color, title, xlab, ylab, x_col=None):
        """Scatter plot of all peer values — anonymous dots + your company highlighted."""
        fig = go.Figure()
        yr_df = _CONSOLIDATED_DF[_CONSOLIDATED_DF["Year"] == rep_year]
        if not yr_df.empty and col in yr_df.columns:
            peer_vals = yr_df[yr_df["Company"] != company][col].dropna()
            x_vals    = (yr_df[yr_df["Company"] != company][x_col].dropna()
                         if x_col else pd.Series([None]*len(peer_vals)))
            # Anonymous peers
            for j, (idx, pv) in enumerate(peer_vals.items()):
                xv = float(yr_df.loc[idx, x_col]) if x_col and idx in yr_df.index else j+1
                fig.add_trace(go.Scatter(
                    x=[xv], y=[pv], mode="markers",
                    marker=dict(size=9, color="#CBD5E1",
                                line=dict(color="white", width=1)),
                    name=f"Peer {j+1}", showlegend=False,
                    hovertemplate=f"Peer: {pv:.3f}<extra></extra>",
                ))
        # Your company
        fig.add_trace(go.Scatter(
            x=[your_val if not x_col else float(
                _CONSOLIDATED_DF[(_CONSOLIDATED_DF["Company"]==company) &
                (_CONSOLIDATED_DF["Year"]==rep_year)].get(x_col, pd.Series([your_val])).iloc[0])],
            y=[your_val], mode="markers+text",
            marker=dict(size=14, color=color, symbol="diamond",
                        line=dict(color="white", width=2)),
            text=[company.split()[0]], textposition="top center",
            textfont=dict(size=10, color=color, family="Inter"),
            name="You", showlegend=False,
            hovertemplate=f"<b>You</b>: {your_val:.3f}<extra></extra>",
        ))
        fig.update_layout(**chart_layout_defaults(title, height=250, showlegend=False),
                          xaxis=dict(title=dict(text=xlab), gridcolor="#F1F5F9"),
                          yaxis=dict(title=dict(text=ylab), gridcolor="#F1F5F9"))
        apply_chart_animation(fig)
        return fig

    def _trend_vs_sector(kpi_key, sec_col, label, color, show_quartiles=True):
        """Company trend line vs sector IQR band with Q1/Median/Q3 reference lines.
        Uses string x-axis (categorical) to prevent float interpolation."""
        sec_mean, sec_q25, sec_q75 = _sector_series(sec_col)
        # Only show years that are IN the selected time range (ys)
        # This ensures sector bands and company lines cover exactly the same window
        _ys_int = ys   # integers, already range-filtered
        # For sector: only include years from ys that have sector data
        yr_list  = [y for y in sorted(_ys_int) if y in sec_mean]
        # For company: only include years from ys that have company trend data
        yr_str   = [str(y) for y in yr_list]
        ys_str   = [str(y) for y in _ys_int if y in trend]

        fig = go.Figure()
        # IQR band (Q1–Q3 shaded)
        fig.add_trace(go.Scatter(
            x=yr_str, y=[sec_q75.get(y) for y in yr_list],
            fill=None, mode="lines", line=dict(width=0), showlegend=False,
            name="Q3"))
        r, g, b = int(color[1:3],16), int(color[3:5],16), int(color[5:7],16)
        fig.add_trace(go.Scatter(
            x=yr_str, y=[sec_q25.get(y) for y in yr_list],
            fill="tonexty", mode="lines", line=dict(width=0),
            fillcolor=f"rgba({r},{g},{b},0.12)",
            name="Sector IQR (Q1–Q3)",
        ))
        if show_quartiles:
            fig.add_trace(go.Scatter(
                x=yr_str, y=[sec_q25.get(y) for y in yr_list],
                mode="lines", name="Q1 (25th pct)",
                line=dict(color="#94A3B8", width=1, dash="dot"),
                hovertemplate="Q1: %{y:.3f}<extra></extra>",
            ))
            fig.add_trace(go.Scatter(
                x=yr_str, y=[sec_mean.get(y) for y in yr_list],
                mode="lines", name="Sector Median",
                line=dict(color="#64748B", width=1.5, dash="dashdot"),
                hovertemplate="Median: %{y:.3f}<extra></extra>",
            ))
            fig.add_trace(go.Scatter(
                x=yr_str, y=[sec_q75.get(y) for y in yr_list],
                mode="lines", name="Q3 (75th pct)",
                line=dict(color="#94A3B8", width=1, dash="dot"),
                hovertemplate="Q3: %{y:.3f}<extra></extra>",
            ))
        # Company line — only years that exist in trend data
        co_y = [trend[y][kpi_key] for y in _ys_int if y in trend]
        fig.add_trace(go.Scatter(
            x=ys_str, y=co_y, mode="lines+markers",
            name=company.split()[0],
            line=dict(color=color, width=2.5),
            marker=dict(size=6, color=color),
            hovertemplate="%{y:.3f}<extra>" + company.split()[0] + "</extra>",
        ))
        fig.update_layout(**chart_layout_defaults(label, height=270),
                          hovermode="x unified",
                          yaxis=dict(
                              gridcolor="#F1F5F9",
                              tickfont=dict(size=12, color="#1C2E3F"),
                              showline=True, linecolor="#999",
                              showticklabels=True,
                          ),
                          yaxis2=dict(
                              tickfont=dict(size=12, color="#1C2E3F"),
                              showgrid=False, showline=True, linecolor="#999",
                              showticklabels=True,
                          ),
                          xaxis=dict(
                              gridcolor="#F1F5F9", type="category",
                              tickfont=dict(size=12, color="#1C2E3F"),
                              showline=True, linecolor="#999",
                              showticklabels=True,
                          ))
        apply_chart_animation(fig)
        return fig

    def _pdf_download_btn(label: str, key: str, figs_data: list = None):
        """
        Generate a multi-section benchmarking PDF using matplotlib (no kaleido).
        figs_data: list of (section_title, chart_type, *data_args) tuples.
        """
        import sys, os
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        try:
            import pdf_charts_v2 as pc
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units    import mm
            from reportlab.pdfgen       import canvas as rl_canvas
            from reportlab.lib.utils    import ImageReader
            import io as _io

            W, H = A4; MARGIN = 14*mm; CW = W - 2*MARGIN
            buf = _io.BytesIO()
            c   = rl_canvas.Canvas(buf, pagesize=A4)

            # Cover header
            c.setFillColor((10/255, 34/255, 64/255))
            c.rect(0, H - 26*mm, W, 26*mm, fill=1, stroke=0)
            c.setFont("Helvetica-Bold", 13)
            c.setFillColor((1,1,1))
            c.drawString(MARGIN, H - 12*mm, f"{company}  ·  {rep_year}  —  {label} Benchmarking")
            c.setFont("Helvetica", 8)
            c.setFillColor((0.55, 0.65, 0.75))
            c.drawString(MARGIN, H - 20*mm, "TIP ESG Platform  ·  dss+ consulting  ·  WBCSD Tire Industry Project")
            cursor = H - 30*mm

            def _embed(img_bytes, title="", h=62*mm):
                nonlocal cursor
                if cursor - h < MARGIN + 10*mm:
                    c.showPage(); cursor = H - MARGIN
                if title:
                    c.setFont("Helvetica-Bold", 9); c.setFillColor((10/255,34/255,64/255))
                    c.drawString(MARGIN, cursor - 5*mm, title); cursor -= 7*mm
                reader = ImageReader(_io.BytesIO(img_bytes))
                c.drawImage(reader, MARGIN, cursor - h, width=CW, height=h,
                            preserveAspectRatio=True)
                cursor -= h + 5*mm

            if figs_data:
                # Compute sector series once for trend charts
                def _ss(col):
                    if _CONSOLIDATED_DF.empty or col not in _CONSOLIDATED_DF.columns:
                        return {}, {}, {}
                    grp = _CONSOLIDATED_DF.groupby("Year")[col]
                    return grp.mean().to_dict(), grp.quantile(.25).to_dict(), grp.quantile(.75).to_dict()

                for item in figs_data:
                    kind = item[0]
                    if kind == "radar":
                        _, dims, co_sc, sec_sc, co_name = item
                        _embed(pc.radar_chart(dims, co_sc, sec_sc, co_name), "ESG Performance Radar")
                    elif kind == "position_bar":
                        _, names, positions, colors_list = item
                        _embed(pc.position_bar(names, positions, colors_list), "Sector Percentile Position")
                    elif kind == "improvement_table":
                        _, rows = item
                        # Simple text table
                        if rows:
                            c.setFont("Helvetica-Bold", 9)
                            c.setFillColor((10/255,34/255,64/255))
                            c.drawString(MARGIN, cursor - 4*mm, "Improvement Summary")
                            cursor -= 7*mm
                            c.setFont("Helvetica", 8)
                            for row in rows:
                                c.setFillColor((10/255,34/255,64/255))
                                kpi_txt = str(row.get("KPI",""))
                                val_txt = str(list(row.values())[-1])
                                c.drawString(MARGIN, cursor - 4*mm, f"• {kpi_txt}: {val_txt}")
                                cursor -= 5*mm
                    elif kind == "line_vs_sector":
                        _, sec_col, kpi_key, title_str, color = item
                        sm, sq25, sq75 = _ss(sec_col)
                        co_y = [trend.get(y, {}).get(kpi_key) for y in ys]
                        _embed(pc.line_vs_sector(ys, co_y, sm, sq25, sq75,
                               company.split()[0], title_str, color=color), title_str)
                    elif kind == "stacked_area_scope":
                        _, title_str = item
                        _embed(pc.stacked_area(ys,
                            {"Scope 1": [trend.get(y,{}).get("scope1",0) for y in ys],
                             "Scope 2": [trend.get(y,{}).get("scope2",0) for y in ys]},
                            title_str, color_dict={"Scope 1": pc.C["co2"], "Scope 2": "#94A3B8"}),
                            title_str)
                    elif kind == "energy_mix_bar":
                        _, title_str = item
                        fuel_map = {
                            "Nat. Gas": [trend.get(y,{}).get("nat_gas",0) for y in ys],
                            "Renew. Elec": [trend.get(y,{}).get("renew_gj",0) for y in ys],
                            "Non-Renew.": [trend.get(y,{}).get("nonrenew_gj",0) for y in ys],
                            "Coal":     [trend.get(y,{}).get("coal",0) for y in ys],
                            "Diesel":   [trend.get(y,{}).get("diesel",0) for y in ys],
                        }
                        cmap = {"Nat. Gas":pc.C["energy"],"Renew. Elec":pc.C["green"],
                                "Non-Renew.":"#94A3B8","Coal":"#475569","Diesel":"#78716C"}
                        _embed(pc.stacked_bar(ys, fuel_map, title_str, color_dict=cmap), title_str)
                    elif kind == "elec_mix_bar":
                        _, title_str = item
                        total_e = [max(trend.get(y,{}).get("renew_gj",0)+trend.get(y,{}).get("nonrenew_gj",0),1) for y in ys]
                        _embed(pc.stacked_bar(ys, {
                            "Renewable":     [trend.get(y,{}).get("renew_gj",0)/t*100 for y,t in zip(ys,total_e)],
                            "Non-Renewable": [trend.get(y,{}).get("nonrenew_gj",0)/t*100 for y,t in zip(ys,total_e)],
                        }, title_str, color_dict={"Renewable":pc.C["green"],"Non-Renewable":"#94A3B8"},
                        pct_mode=True), title_str)
                    elif kind == "water_bar":
                        _, title_str = item
                        _embed(pc.bar_chart(ys, [trend.get(y,{}).get("water_m3",0)/1e6 for y in ys],
                               title_str, "M m³", color=pc.C["water"]), title_str)
                    elif kind == "waste_area":
                        _, title_str = item
                        _embed(pc.area_with_target(ys, [trend.get(y,{}).get("waste_pct",0) for y in ys],
                               title_str, "%", color=pc.C["waste"]), title_str)
                    elif kind == "waste_bar":
                        _, title_str = item
                        _embed(pc.stacked_bar(ys, {
                            "Total Waste": [trend.get(y,{}).get("waste_total",0) for y in ys],
                            "Recovered":   [trend.get(y,{}).get("waste_rec",0)   for y in ys],
                        }, title_str, color_dict={"Total Waste":"#E2E8F0","Recovered":pc.C["waste"]}),
                        title_str)

            c.save(); buf.seek(0); pdf_bytes = buf.read()
            st.download_button(
                f"⬇  Download {label} Report (PDF)",
                data=pdf_bytes,
                file_name=f"{company.replace(' ','_')}_{label.replace(' ','_')}_{rep_year}_Benchmark.pdf",
                mime="application/pdf", key=key, use_container_width=True,
            )
        except Exception as e:
            st.error(f"PDF generation failed: {e}. Make sure pdf_charts_v2.py and reportlab are installed.")

    # ── Full combined PDF — all sections in one document ─────────────────────
    def _full_bench_pdf():
        """Generate one PDF with all 6 benchmark sections: General → Waste."""
        import sys as _sys, os as _os
        _sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
        try:
            import pdf_charts_v2 as pc
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units    import mm
            from reportlab.pdfgen       import canvas as rl_canvas
            from reportlab.lib.utils    import ImageReader
            import io as _io

            W, H = A4; MARGIN = 14*mm; CW = W - 2*MARGIN
            buf = _io.BytesIO()
            cv  = rl_canvas.Canvas(buf, pagesize=A4)

            def _cover():
                cv.setFillColor((10/255, 34/255, 64/255))
                cv.rect(0, H - 30*mm, W, 30*mm, fill=1, stroke=0)
                cv.setFont("Helvetica-Bold", 14); cv.setFillColor((1,1,1))
                cv.drawString(MARGIN, H - 13*mm, f"{company}  ·  ESG Benchmarking Report  ·  {rep_year}")
                cv.setFont("Helvetica", 8); cv.setFillColor((.55,.65,.75))
                cv.drawString(MARGIN, H-21*mm, "TIP ESG Platform  ·  dss+ consulting  ·  WBCSD Tire Industry Project")

            def _section_title(cv, title, cursor):
                cv.setFillColor((.94,.95,.98)); cv.rect(MARGIN, cursor-9*mm, CW, 9*mm, fill=1, stroke=0)
                cv.setFillColor((10/255,34/255,64/255)); cv.rect(MARGIN, cursor-9*mm, 2.5, 9*mm, fill=1, stroke=0)
                cv.setFont("Helvetica-Bold", 11); cv.setFillColor((10/255,34/255,64/255))
                cv.drawString(MARGIN+5, cursor-6*mm, title)
                return cursor - 11*mm

            def _embed(cv, img_bytes, cursor, caption="", h=60*mm):
                if cursor - h < MARGIN + 15*mm:
                    cv.showPage(); _cover(); cursor = H - 34*mm
                if caption:
                    cv.setFont("Helvetica", 8); cv.setFillColor((.4,.4,.4))
                    cv.drawString(MARGIN, cursor-4*mm, caption); cursor -= 6*mm
                reader = ImageReader(_io.BytesIO(img_bytes))
                cv.drawImage(reader, MARGIN, cursor-h, width=CW, height=h, preserveAspectRatio=True)
                return cursor - h - 4*mm

            def _ss(col):
                if _CONSOLIDATED_DF.empty or col not in _CONSOLIDATED_DF.columns:
                    return {},{},{}
                grp = _CONSOLIDATED_DF.groupby("Year")[col]
                return grp.mean().to_dict(), grp.quantile(.25).to_dict(), grp.quantile(.75).to_dict()

            _cover()
            cursor = H - 34*mm

            # ── General ──────────────────────────────────────────────────────
            cursor = _section_title(cv, "General — ESG Performance Overview", cursor)
            dims = ["CO₂ Intensity","Energy Intensity","Water Intensity","Renewable Elec.","Waste Recovery"]
            co_scores = []
            for b in BM:
                rng = max(b._hi - b._lo, 0.001)
                raw = (b.company_value - b._lo) / rng
                co_scores.append(max(0, min(100, (1-raw)*100 if b.lower_is_better else raw*100)))
            sec_sc = _compute_industry_scores(_CONSOLIDATED_DF, rep_year)
            cursor = _embed(cv, pc.radar_chart(dims, co_scores, sec_sc, company.split()[0]),
                            cursor, "ESG Radar Profile — company vs sector median")
            positions = [max(0,min(100,(1-(b.company_value-b._lo)/max(b._hi-b._lo,0.001))*100)) if b.lower_is_better
                         else max(0,min(100,((b.company_value-b._lo)/max(b._hi-b._lo,0.001))*100)) for b in BM]
            cursor = _embed(cv, pc.position_bar(
                ["CO₂ Intensity","Energy Intensity","Water Intensity","Renewable Elec.","Waste Recovery"],
                positions, [CAT_CO2, CAT_ENERGY, CAT_WATER, CAT_RENEW, CAT_WASTE]),
                cursor, "Sector Percentile Position (100 = best)", h=45*mm)

            # ── CO₂ ──────────────────────────────────────────────────────────
            cursor = _section_title(cv, "CO₂ — Carbon Emissions & Intensity", cursor)
            sm,sq25,sq75 = _ss("Total CO2 - KPI")
            cursor = _embed(cv, pc.line_vs_sector(ys, [trend.get(y,{}).get("co2_kpi") for y in ys],
                sm,sq25,sq75,company.split()[0],"CO₂ Intensity vs Sector (tCO₂/t)",color=pc.C["co2"]),
                cursor, "Company line vs sector IQR band · Q1/Median/Q3 shown")
            cursor = _embed(cv, pc.stacked_area(ys,
                {"Scope 1":[trend.get(y,{}).get("scope1",0) for y in ys],
                 "Scope 2":[trend.get(y,{}).get("scope2",0) for y in ys]},
                "Scope 1 vs Scope 2 (tCO₂)",
                color_dict={"Scope 1":pc.C["co2"],"Scope 2":"#94A3B8"}),
                cursor, "Scope 1 = fuel combustion · Scope 2 = purchased energy")

            # ── Energy ────────────────────────────────────────────────────────
            cursor = _section_title(cv, "Energy — Intensity & Fuel Mix", cursor)
            sm,sq25,sq75 = _ss("Total energy - KPI")
            cursor = _embed(cv, pc.line_vs_sector(ys, [trend.get(y,{}).get("energy_kpi") for y in ys],
                sm,sq25,sq75,company.split()[0],"Energy Intensity vs Sector (GJ/t)",color=pc.C["energy"]),
                cursor)
            cursor = _embed(cv, pc.stacked_bar(ys,
                {"Nat. Gas":[trend.get(y,{}).get("nat_gas",0) for y in ys],
                 "Renew. Elec":[trend.get(y,{}).get("renew_gj",0) for y in ys],
                 "Diesel":[trend.get(y,{}).get("diesel",0) for y in ys],
                 "Coal":[trend.get(y,{}).get("coal",0) for y in ys]},
                "Energy Mix by Source (GJ)",
                color_dict={"Nat. Gas":pc.C["energy"],"Renew. Elec":pc.C["green"],
                            "Diesel":"#78716C","Coal":"#475569"}),
                cursor, "Fuel mix evolution over all available years")

            # ── Electricity ───────────────────────────────────────────────────
            cursor = _section_title(cv, "Electricity — Renewable vs Non-Renewable", cursor)
            total_e = [max(trend.get(y,{}).get("renew_gj",0)+trend.get(y,{}).get("nonrenew_gj",0),1) for y in ys]
            cursor = _embed(cv, pc.stacked_bar(ys,
                {"Renewable":[trend.get(y,{}).get("renew_gj",0)/t*100 for y,t in zip(ys,total_e)],
                 "Non-Renewable":[trend.get(y,{}).get("nonrenew_gj",0)/t*100 for y,t in zip(ys,total_e)]},
                "Electricity Mix (%)", pct_mode=True,
                color_dict={"Renewable":pc.C["green"],"Non-Renewable":"#94A3B8"}), cursor)
            sm,sq25,sq75 = _ss("Renewable_Electricity_Share_%")
            cursor = _embed(cv, pc.line_vs_sector(ys,
                [trend.get(y,{}).get("renew_pct") for y in ys],
                sm,sq25,sq75,company.split()[0],"Renewable Electricity Share vs Sector (%)",color=pc.C["green"]),
                cursor, "Share of electricity from renewable sources vs TIP sector quartiles")

            # ── Water ─────────────────────────────────────────────────────────
            cursor = _section_title(cv, "Water — Intensity & Withdrawals", cursor)
            sm,sq25,sq75 = _ss("Water intake - KPI")
            cursor = _embed(cv, pc.line_vs_sector(ys,
                [trend.get(y,{}).get("water_kpi") for y in ys],
                sm,sq25,sq75,company.split()[0],"Water Intensity vs Sector (m³/t)",color=pc.C["water"]),
                cursor)
            cursor = _embed(cv, pc.bar_chart(ys,
                [trend.get(y,{}).get("water_m3",0)/1e6 for y in ys],
                "Water Withdrawals (M m³)","M m³",color=pc.C["water"]), cursor)

            # ── Waste ─────────────────────────────────────────────────────────
            cursor = _section_title(cv, "Waste — Recovery Rate & Volumes", cursor)
            sm,sq25,sq75 = _ss("Waste_Recovery_Rate_%")
            cursor = _embed(cv, pc.area_with_target(ys,
                [trend.get(y,{}).get("waste_pct",0) for y in ys],
                "Waste Recovery Rate vs Sector (%)","% recovered",color=pc.C["waste"]),
                cursor, "Target 90% shown · TIP sector IQR band")
            cursor = _embed(cv, pc.stacked_bar(ys,
                {"Total Waste":[trend.get(y,{}).get("waste_total",0) for y in ys],
                 "Recovered":[trend.get(y,{}).get("waste_rec",0) for y in ys]},
                "Total Waste vs Recovered (T)",
                color_dict={"Total Waste":"#E2E8F0","Recovered":pc.C["waste"]}), cursor)

            # Footer on last page
            cv.setFillColor((.95,.96,.98)); cv.rect(0,0,W,11*mm,fill=1,stroke=0)
            cv.setFont("Helvetica",6); cv.setFillColor((.4,.4,.4))
            cv.drawString(MARGIN, 5*mm, "TIP ESG Platform · dss+ consulting · WBCSD Tire Industry Project · Methodology: GHG Protocol")
            from datetime import date as _ddate
            cv.drawRightString(W-MARGIN, 5*mm, f"Generated {_ddate.today():%d %b %Y} · {company} · {rep_year}")
            cv.save(); buf.seek(0)
            return buf.read()
        except Exception as ex:
            try:
                import traceback as _tb
                _err = _tb.format_exc().strip().split("\n")[-1]
                st.session_state["_pdf_bench_error"] = _err
            except Exception:
                pass
            return None

    _pdf_col, _ = st.columns([2, 4])
    with _pdf_col:
        _pdf_all = _full_bench_pdf()
        if _pdf_all:
            # Clear any previous error on success
            st.session_state.pop("_pdf_bench_error", None)
            st.download_button(
                "⬇  Download Full Benchmarking Report (PDF)",
                data=_pdf_all,
                file_name=f"{company.replace(' ','_')}_Benchmarking_{rep_year}.pdf",
                mime="application/pdf", key="dl_full_bench",
                use_container_width=True, type="primary",
            )
        else:
            _pdf_err = st.session_state.pop("_pdf_bench_error", "")
            if _pdf_err:
                st.error(
                    f"⚠ PDF generation failed: {_pdf_err}\n\n"
                    f"Install with the **same Python that runs Streamlit**: "
                    f"`pip install reportlab matplotlib`",
                    icon=None,
                )
            else:
                st.warning(
                    "⚠ PDF requires `reportlab` and `matplotlib`. "
                    "Install with the **same Python** that runs Streamlit: "
                    "`pip install reportlab matplotlib`",
                    icon=None,
                )

    # ── KPI Tabs ──────────────────────────────────────────────────────────────
    tab_co2, tab_energy, tab_elec, tab_water, tab_waste = st.tabs([
        "CO₂ Emissions", "Energy", "Electricity", "Water", "Waste"
    ])

    # ── TIP Chart helpers shared with benchmarking ───────────────────────────
    from plotly.subplots import make_subplots as _msp

    _TC = {
        "bar_blue":   "#B8CDD9", "bar_blue2":  "#2D4A5A",
        "bar_beige":  "#C8B49A", "bar_beige2": "#8A7B68",
        "bar_green":  "#7BAF74", "bar_orange": "#E0935A",
        "bar_sand":   "#D4C5A9","bar_committed":"#9FB8C5",
        "line_dark":  "#2D4A5A", "line_light": "#8FA5B5",
    }

    def _blt(title="", h=330):
        return dict(
            title=dict(text=f"<b>{title}</b>",
                       font=dict(size=14, color="#1C2E3F", family="Arial, sans-serif"), x=0),
            height=h, margin=dict(l=55, r=110, t=50, b=60),
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            xaxis=dict(
                showgrid=False, linecolor="#999", linewidth=1.2,
                showline=True, mirror=False,
                tickfont=dict(size=12, color="#1C2E3F", family="Arial"),
                tickangle=0, type="category",
            ),
            yaxis=dict(
                showgrid=True, gridcolor="rgba(0,0,0,0.07)", zeroline=False,
                showline=True, linecolor="#999", linewidth=1.2,
                tickfont=dict(size=12, color="#1C2E3F", family="Arial"),
                showticklabels=True,
                autorange=True,           # ensures top of chart has breathing room
            ),
            legend=dict(orientation="h", x=0.5, xanchor="center", y=-0.24,
                        font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
            hovermode="x unified", showlegend=True,
        )

    def _open_mk(col, sz=9):
        return dict(symbol="circle", size=sz, color="white", line=dict(color=col, width=2))

    def _b_dbline(xs, bv, bl, bc, lv, ll, lc, title="", h=330,
                  bfmt=".1f", lfmt=".2f", byt="", lyt=""):
        fig = _msp(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Bar(x=xs, y=bv, name=bl, marker_color=bc, marker_line_width=0,
            text=[f"{v:{bfmt}}" if v is not None and v==v else "" for v in bv],
            textposition="outside", textfont=dict(size=12, color="#1C2E3F", family="Arial"), width=0.5,
        ), secondary_y=False)
        fig.add_trace(go.Scatter(x=xs, y=lv, name=ll,
            mode="lines+markers+text", line=dict(color=lc, width=2.5),
            marker=_open_mk(lc),
            text=[f"{v:{lfmt}}" if v is not None and v==v else "" for v in lv],
            textposition="top center", textfont=dict(size=12, color="#1C2E3F", family="Arial"),
        ), secondary_y=True)
        lay = _blt(title, h)
        lay["yaxis"]["title"] = dict(text=byt, font=dict(size=9, color="#666"))
        lay["yaxis2"] = dict(
            tickfont=dict(size=12, color="#1C2E3F", family="Arial"),
            showgrid=False, zeroline=False,
            showline=True, linecolor="#999", linewidth=1.2,
            showticklabels=True,
            title=dict(text=lyt, font=dict(size=11, color="#444")),
        )
        fig.update_layout(**lay)
        return fig

    def _b_stack100(xs, traces, title="", h=330):
        fig = go.Figure()
        for (vals, lbl, col) in traces:
            fig.add_trace(go.Bar(x=xs, y=vals, name=lbl, marker_color=col, marker_line_width=0,
                text=[f"{v:.1f}%" if v and v>5 else "" for v in vals],
                textposition="inside", textfont=dict(size=12, color="white", family="Arial"),
                hovertemplate=f"{lbl}: %{{y:.1f}}%<extra></extra>"))
        lay = _blt(title, h)
        lay["barmode"] = "stack"
        lay["yaxis"]["ticksuffix"] = "%"
        lay["yaxis"]["range"] = [0, 100]
        fig.update_layout(**lay)
        return fig

    def _b_dline(xs, s1v, s1l, s1c, s2v, s2l, s2c, title="", h=300,
                 s1f=".1f", s2f=".1f", yt="", s2yt="", right_y=False):
        fig = _msp(specs=[[{"secondary_y": True}]]) if right_y else go.Figure()
        kw1 = dict(x=xs, y=s1v, name=s1l, mode="lines+markers+text",
                   line=dict(color=s1c, width=2.5), marker=_open_mk(s1c),
                   text=[f"{v:{s1f}}" if v is not None else "" for v in s1v],
                   textposition="top center", textfont=dict(size=12, color="#1C2E3F", family="Arial"))
        kw2 = dict(x=xs, y=s2v, name=s2l, mode="lines+markers+text",
                   line=dict(color=s2c, width=2.5), marker=_open_mk(s2c),
                   text=[f"{v:{s2f}}" if v is not None else "" for v in s2v],
                   textposition="top center", textfont=dict(size=12, color="#1C2E3F", family="Arial"))
        if right_y:
            fig.add_trace(go.Scatter(**kw1), secondary_y=False)
            fig.add_trace(go.Scatter(**kw2), secondary_y=True)
            lay = _blt(title, h)
            lay["yaxis"]["title"] = dict(text=yt, font=dict(size=9, color="#666"))
            lay["yaxis"]["ticksuffix"] = "%"
            lay["yaxis2"] = dict(
            tickfont=dict(size=12, color="#1C2E3F", family="Arial"),
            showgrid=False, zeroline=False,
            showline=True, linecolor="#999", linewidth=1.2,
            showticklabels=True,
            title=dict(text=s2yt, font=dict(size=11, color="#444")),
        )
        else:
            fig.add_trace(go.Scatter(**kw1))
            fig.add_trace(go.Scatter(**kw2))
            lay = _blt(title, h)
            lay["yaxis"]["title"] = dict(text=yt, font=dict(size=9, color="#666"))
            lay["yaxis"]["ticksuffix"] = "%"
        fig.update_layout(**lay)
        return fig

    _ys_str = [str(y) for y in ys]

    with tab_co2:
        st.caption("CO₂ intensity vs TIP sector peers — with Q1/Median/Q3 reference bands")
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            st.plotly_chart(
                _trend_vs_sector("co2_kpi","Total CO2 - KPI",
                    "CO₂ Intensity vs Sector (tCO₂/t)", _TC["bar_blue2"]),
                use_container_width=True, key=_chart_key(company, rep_year, "co2t"))
        with c2:
            # Scope 1 vs Scope 2 stacked bar (TIP Fig 7 style)
            s1v = [trend[y]["scope1"] for y in ys]
            s2v = [trend[y]["scope2"] for y in ys]
            tot = [s1+s2 for s1,s2 in zip(s1v,s2v)]
            fig_sc = go.Figure()
            fig_sc.add_trace(go.Bar(x=_ys_str, y=s1v, name="Scope 1 (direct)",
                marker_color=_TC["bar_blue"], marker_line_width=0,
                text=[f"{v:.2f}" if v else "" for v in s1v],
                textposition="inside", textfont=dict(size=12, color="white", family="Arial")))
            fig_sc.add_trace(go.Bar(x=_ys_str, y=s2v, name="Scope 2 (indirect)",
                marker_color=_TC["bar_blue2"], marker_line_width=0,
                text=[f"{v:.2f}" if v else "" for v in s2v],
                textposition="inside", textfont=dict(size=12, color="white", family="Arial")))
            lay_sc = _blt("CO₂ Scope 1 vs Scope 2 trend (tCO₂)", 310)
            lay_sc["barmode"] = "stack"
            lay_sc["yaxis"]["title"] = dict(text="tCO₂", font=dict(size=9))
            fig_sc.update_layout(**lay_sc)
            st.plotly_chart(fig_sc, use_container_width=True, key=_chart_key(company, rep_year, "4"))
        _pdf_download_btn("CO2", "dl_bench_co2", [
            ("line_vs_sector", "Total CO2 - KPI", "co2_kpi",
             "CO₂ Intensity Trend vs Sector", _TC["bar_blue2"]),
        ])

    with tab_energy:
        st.caption("Energy intensity & consumption mix — Q1/Median/Q3 reference bands shown")
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            st.plotly_chart(
                _trend_vs_sector("energy_kpi","Total energy - KPI",
                    "Energy Intensity vs Sector (GJ/t)", _TC["bar_blue2"]),
                use_container_width=True, key=_chart_key(company, rep_year, "5"))
        with c2:
            # Fig 5 style energy mix 100% stacked
            fuel_keys = [
                ("nat_gas",    "Natural Gas",                              "#8FA5B5"),
                ("renew_gj",   "Renewable Electricity (purchased+self-gen)", _TC["bar_green"]),
                ("nonrenew_gj","Non-renewable Electricity",                 _TC["bar_sand"]),
                ("coal",       "Coal",                                      "#666"),
                ("diesel",     "Diesel/LPG/Other",                         _TC["bar_orange"]),
            ]
            totals_e = [max(sum(trend[y].get(k,0) for k,_,_ in fuel_keys), 1) for y in ys]
            traces_e = []
            for fkey, flbl, fcol in fuel_keys:
                pcts = [trend[y].get(fkey,0)/tot*100 for y,tot in zip(ys,totals_e)]
                if any(p>0 for p in pcts):
                    traces_e.append((pcts, flbl, fcol))
            st.plotly_chart(_b_stack100(_ys_str, traces_e,
                "Energy mix (%)", 310),
                use_container_width=True, key=_chart_key(company, rep_year, "6"))
        _pdf_download_btn("Energy", "dl_bench_energy", [
            ("line_vs_sector", "Total energy - KPI", "energy_kpi",
             "Energy Intensity vs Sector (GJ/t)", _TC["bar_blue2"]),
        ])

    with tab_elec:
        st.caption("Electricity from renewable sources — company trend vs sector")
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            # Fig 6 style — renewable vs non-renewable stacked %
            total_e2 = [max(trend[y]["renew_gj"]+trend[y]["nonrenew_gj"], 1) for y in ys]
            renew_pct_co  = [trend[y]["renew_gj"]/t*100 for y,t in zip(ys,total_e2)]
            nonren_pct_co = [100-v for v in renew_pct_co]
            fig_e6 = go.Figure()
            fig_e6.add_trace(go.Bar(x=_ys_str, y=renew_pct_co,
                name="Renewable electricity (GJ)", marker_color=_TC["bar_blue2"],
                marker_line_width=0,
                text=[f"{v:.1f}%" if v else "" for v in renew_pct_co],
                textposition="inside", textfont=dict(size=12, color="white", family="Arial")))
            fig_e6.add_trace(go.Bar(x=_ys_str, y=nonren_pct_co,
                name="Non-renewable electricity (GJ)", marker_color=_TC["bar_sand"],
                marker_line_width=0,
                text=[f"{v:.1f}%" if v else "" for v in nonren_pct_co],
                textposition="inside", textfont=dict(size=12, color="#1C2E3F", family="Arial")))
            lay_e6 = _blt("Electricity from renewable sources (%)", 310)
            lay_e6["barmode"] = "stack"
            lay_e6["yaxis"]["ticksuffix"] = "%"
            lay_e6["yaxis"]["range"] = [0,100]
            fig_e6.update_layout(**lay_e6)
            st.plotly_chart(fig_e6, use_container_width=True, key=_chart_key(company, rep_year, "7"))
        with c2:
            st.plotly_chart(
                _trend_vs_sector("renew_pct","Renewable_Electricity_Share_%",
                    "Renewable Electricity Share vs Sector (%)", _TC["bar_green"]),
                use_container_width=True, key=_chart_key(company, rep_year, "8"))
        _pdf_download_btn("Electricity", "dl_bench_elec", [
            ("line_vs_sector", "Renewable_Electricity_Share_%", "renew_pct",
             "Renewable Electricity Share vs Sector (%)", _TC["bar_green"]),
        ])

    with tab_water:
        st.caption("Water withdrawals & intensity — company trend vs sector Q1/Median/Q3")
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            st.plotly_chart(
                _trend_vs_sector("water_kpi","Water intake - KPI",
                    "Water Intensity vs Sector (m³/t)", _TC["bar_blue2"]),
                use_container_width=True, key=_chart_key(company, rep_year, "9"))
        with c2:
            # Fig 9 style — water withdrawals bar + intensity line
            st.plotly_chart(_b_dbline(
                _ys_str,
                [trend[y]["water_m3"]/1e6 for y in ys],
                "Water withdrawals (M m³)", _TC["bar_beige"],
                [trend[y]["water_kpi"] for y in ys],
                "Water intensity (m³/t)", _TC["line_dark"],
                title="Water withdrawals & intensity",
                byt="Million m³", lyt="m³/t", bfmt=".2f", lfmt=".2f",
            ), use_container_width=True, key=_chart_key(company, rep_year, "10"))
        _pdf_download_btn("Water", "dl_bench_water", [
            ("line_vs_sector", "Water intake - KPI", "water_kpi",
             "Water Intensity Trend vs Sector", _TC["bar_blue2"]),
        ])

    with tab_waste:
        st.caption("Waste recovery rate & volumes — company trend vs sector Q1/Median/Q3")
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            st.plotly_chart(
                _trend_vs_sector("waste_pct","Waste_Recovery_Rate_%",
                    "Waste Recovery Rate vs Sector (%)", _TC["bar_blue2"]),
                use_container_width=True, key=_chart_key(company, rep_year, "11"))
        with c2:
            # Fig 10/11 style — waste total bar + recovery % stacked
            wt_vals = [trend[y]["waste_total"] for y in ys]
            wr_vals = [trend[y]["waste_rec"]   for y in ys]
            we_vals = [max(t-r,0) for t,r in zip(wt_vals,wr_vals)]
            wt_safe = [max(t,1) for t in wt_vals]
            wr_pct  = [r/t*100 for r,t in zip(wr_vals,wt_safe)]
            we_pct  = [100-v for v in wr_pct]
            st.plotly_chart(_b_stack100(
                _ys_str,
                [(wr_pct, "Sent for recovery (%)",  _TC["bar_beige"]),
                 (we_pct, "Sent for disposal (%)",  _TC["bar_blue2"])],
                title="Waste recovery vs disposal (%)", h=310,
            ), use_container_width=True, key=_chart_key(company, rep_year, "12"))
        _pdf_download_btn("Waste", "dl_bench_waste", [
            ("waste_area", "Waste Recovery Rate vs Sector (%)"),
        ])


# ─────────────────────────────────────────────────────────
def _dss_company_selector(page_key: str):
    """
    Lets a dss+ analyst pick any company + year from the consolidated DB.
    Stores selection in session_state under dss_{page_key}_company / _year.
    Returns (company, year, inp, out, prev_inp, prev_out, company_hist).
    """
    co_key = f"dss_{page_key}_company"
    yr_key = f"dss_{page_key}_year"

    companies_in_db = dl.get_companies(_CONSOLIDATED_DF) or COMPANIES
    default_co = (st.session_state.get("reporting_company") or
                  st.session_state.get("user_company") or companies_in_db[0])
    if default_co not in companies_in_db:
        default_co = companies_in_db[0]

    col1, col2, _ = st.columns([2, 1, 3])
    with col1:
        sel_co = st.selectbox(
            "Company to review", options=companies_in_db,
            index=companies_in_db.index(st.session_state.get(co_key, default_co)),
            key=f"sel_{page_key}_co"
        )
    with col2:
        avail_years = dl.get_years(_CONSOLIDATED_DF, sel_co) or [CURR_YEAR]
        sel_yr = st.selectbox(
            "Year", options=sorted(avail_years, reverse=True),
            key=f"sel_{page_key}_yr"
        )

    st.session_state[co_key] = sel_co
    st.session_state[yr_key] = sel_yr

    # Load data for selected and previous year
    hist = dl.get_company_hist(_CONSOLIDATED_DF, sel_co)

    def _make_inp_out(year):
        sd = dl.get_step_data(hist, year)
        sd_clean = {k: v for k, v in sd.items() if k in _VALID_TEMPLATE_FIELDS}
        inp = TemplateInputs(company=sel_co, year=year, **sd_clean)
        return inp, calculate(inp)

    inp,  out  = _make_inp_out(sel_yr)
    try:
        prev_inp, prev_out = _make_inp_out(sel_yr - 1)
    except Exception:
        prev_inp, prev_out = TemplateInputs(company=sel_co, year=sel_yr-1), None

    return sel_co, sel_yr, inp, out, prev_inp, prev_out, hist


def _compute_completeness(inp: "TemplateInputs", out: "TemplateOutputs") -> dict:
    """
    Returns {section_label: pct_complete} based on which fields have non-zero data.
    """
    def pct(*vals):
        filled = sum(1 for v in vals if v is not None and float(v) != 0)
        return int(filled / len(vals) * 100)

    fuel_vals = [inp.nat_gas, inp.coal_sub, inp.propane, inp.fuel_oil_heavy_a,
                 inp.diesel, inp.petrol, inp.biomass, inp.lpg, inp.other_fuels]
    fuel_filled = sum(1 for v in fuel_vals if v > 0)

    return {
        "ISO 14001":           pct(inp.total_sites, inp.iso_sites),
        "Production":          pct(inp.production),
        "Water":               pct(inp.water_withdrawals),
        "Energy — Electricity":pct(inp.renew_elec_purchased + inp.nonrenew_elec_purchased, inp.self_gen_elec),
        "Energy — Fuels":      min(100, int(fuel_filled / max(len(fuel_vals), 1) * 100)) if inp.nat_gas or inp.coal_sub else 0,
        "CO₂ Scope 1":         pct(out.total_co2_scope1),
        "CO₂ Scope 2":         pct(inp.co2_scope2_steam),
        "Waste":               pct(inp.waste_total, inp.waste_recovery),
        "Pathway 3 (SBTi)":    0,   # not captured in current template
        "Pathway 4 (H&S)":     0,   # not captured in current template
        "Pathway 4 (D&I)":     0,   # not captured in current template
    }


def _compute_readiness_score(completeness: dict, flags) -> tuple:
    """
    Score = weighted completeness average, minus penalties for flags.
    Returns (score: int, label: str)
    """
    weights = {
        "ISO 14001":1,"Production":2,"Water":2,
        "Energy — Electricity":3,"Energy — Fuels":3,
        "CO₂ Scope 1":3,"CO₂ Scope 2":2,"Waste":2,
        "Pathway 3 (SBTi)":1,"Pathway 4 (H&S)":1,"Pathway 4 (D&I)":1,
    }
    total_w  = sum(weights.values())
    raw      = sum(completeness.get(k,0) * w for k,w in weights.items()) / total_w
    n_errors   = sum(1 for f in flags if f.severity == "error")
    n_warnings = sum(1 for f in flags if f.severity == "warning")
    score = max(0, min(100, int(raw - n_errors * 10 - n_warnings * 3)))
    label = "Ready" if score >= 90 else "Review required" if score >= 70 else "Not ready"
    return score, label


# ─────────────────────────────────────────────────────────
# PAGE 4 -- VERIFICATION (dss+ only)  — fully live
# ─────────────────────────────────────────────────────────

def page_verification():
    if not st.session_state.is_dss:
        st.error("This section is restricted to dss+ analysts and managers."); return

    st.markdown("## Data Verification")
    st.caption("Select any company and reporting year from the consolidated dataset to review.")

    sel_co, sel_yr, inp, out, prev_inp, prev_out, hist = _dss_company_selector("verif")
    st.divider()

    # -- Real flags from formula_engine ----------------------------------------
    flags = validate_submission(inp, out, prev_out, threshold=20.0)

    # -- Also compute YoY field-level flags from the data ----------------------
    extra_flags = []
    yoy_fields = [
        ("nat_gas",              "Natural Gas"),
        ("renew_elec_purchased", "Renewable Electricity"),
        ("nonrenew_elec_purchased","Non-Renewable Electricity"),
        ("fuel_oil_heavy_a",     "Fuel Oil"),
        ("coal_sub",             "Coal"),
        ("water_withdrawals",    "Water Withdrawals"),
        ("production",           "Production"),
    ]
    from formula_engine import ValidationFlag
    for field, label in yoy_fields:
        cur  = getattr(inp,      field, 0) or 0
        prev = getattr(prev_inp, field, 0) or 0
        if prev > 0 and cur > 0:
            pct = (cur - prev) / abs(prev) * 100
            if abs(pct) > 20:
                direction = "increase" if pct > 0 else "decrease"
                extra_flags.append(ValidationFlag(
                    severity="warning",
                    message=f"{label} — >{20}% YoY {direction} ({pct:+.1f}%)",
                    detail=(f"{label}: {prev:,.0f} → {cur:,.0f} "
                            f"({'↑' if pct>0 else '↓'}{abs(pct):.1f}%). "
                            f"Verify with company documentation.")
                ))

    all_flags = flags + extra_flags

    # -- Completeness + summary metrics ----------------------------------------
    completeness = _compute_completeness(inp, out)
    avg_complete  = int(sum(completeness.values()) / len(completeness))
    n_err  = sum(1 for f in all_flags if f.severity == "error")
    n_warn = sum(1 for f in all_flags if f.severity == "warning")
    score, label = _compute_readiness_score(completeness, all_flags)

    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Company",      sel_co)
    c2.metric("Year",         str(sel_yr))
    c3.metric("Status",       "Ready" if score >= 90 else "Pending review")
    c4.metric("Completeness", f"{avg_complete}%")
    c5.metric("Open flags",   f"{n_warn} warning{'s' if n_warn!=1 else ''} · {n_err} error{'s' if n_err!=1 else ''}")
    st.divider()

    # -- Flag cards ------------------------------------------------------------
    if not st.session_state.get("flags_resolved_real"):
        st.session_state["flags_resolved_real"] = set()

    resolved_set = st.session_state["flags_resolved_real"]

    for i, flag in enumerate(all_flags):
        flag_id  = f"flag_{sel_co}_{sel_yr}_{i}"
        resolved = flag_id in resolved_set

        sev = "ok" if resolved else flag.severity
        icon_map  = {"ok":"✓", "warning":"!", "error":"✕", "warn":"!"}
        color_map = {"ok":"fc-ok fi-ok", "warning":"fc-warn fi-warn",
                     "error":"fc-error fi-error", "warn":"fc-warn fi-warn"}
        fc, fi = color_map.get(sev, "fc-ok fi-ok").split()
        icon   = icon_map.get(sev, "OK")
        title  = flag.message + (" — Approved" if resolved else "")
        detail = flag.detail

        # H3 FIX: escape flag content before injecting into HTML
        st.markdown(f"""<div class="flag-card {fc}">
          <div class="fc-icon {fi}">{_html.escape(icon)}</div>
          <div><div class="fc-title">{_html.escape(title)}</div>
               <div class="fc-detail">{_html.escape(detail)}</div></div>
        </div>""", unsafe_allow_html=True)

        if not resolved and flag.severity in ("warning","error"):
            cols = st.columns([6,1,1])
            with cols[1]:
                if st.button("Query", key=f"q_{flag_id}"):
                    st.toast(f"Query logged: {flag.message[:50]}...")
            with cols[2]:
                if flag.severity == "warning":
                    if st.button("Accept", key=f"a_{flag_id}", type="primary"):
                        resolved_set.add(flag_id)
                        st.session_state["flags_resolved_real"] = resolved_set
                        st.rerun()
                else:
                    if st.button("Send Back", key=f"sb_{flag_id}"):
                        st.toast(f"Submission returned to {sel_co} with error details.")

    st.divider()
    col_approve, col_flag, col_export, _ = st.columns([1.5, 1.5, 1.5, 1])
    with col_approve:
        warn_ids = [f"flag_{sel_co}_{sel_yr}_{i}"
                    for i, f in enumerate(all_flags) if f.severity == "warning"]
        if st.button("Verify & Approve", type="primary"):
            resolved_set.update(warn_ids)
            st.session_state["flags_resolved_real"] = resolved_set
            # Persist verification status so client's submission bar reflects it
            _write_verification_status(sel_co, sel_yr, "Verified")
            st.success(f"✅ {sel_co} {sel_yr} marked as Verified")
            st.rerun()
    with col_flag:
        if st.button("Mark as Pending", key="mark_pending_btn"):
            _write_verification_status(sel_co, sel_yr, "Pending")
            st.info(f"Marked {sel_co} {sel_yr} as Pending Review")
            st.rerun()
    with col_export:
        if st.button("Export Flag Report"):
            rows = [{"Flag": f.message, "Severity": f.severity, "Detail": f.detail,
                     "Status": "Resolved" if f"flag_{sel_co}_{sel_yr}_{i}" in resolved_set else "Open"}
                    for i, f in enumerate(all_flags)]
            export_df = pd.DataFrame(rows)
            st.download_button(
                "Download CSV", data=export_df.to_csv(index=False).encode(),
                file_name=f"flags_{sel_co.replace(' ','_')}_{sel_yr}.csv",
                mime="text/csv", key="dl_flags"
            )


# ─────────────────────────────────────────────────────────
# PAGE 5 -- AI READINESS (dss+ only) — fully live
# ─────────────────────────────────────────────────────────

def page_readiness():
    if not st.session_state.is_dss:
        st.error("This section is restricted to dss+ analysts and managers."); return

    st.markdown("## AI Readiness Check")
    st.caption("Select any company and reporting year to compute a live readiness score.")

    sel_co, sel_yr, inp, out, prev_inp, prev_out, hist = _dss_company_selector("ready")
    st.divider()

    # -- Compute live values ---------------------------------------------------
    flags        = validate_submission(inp, out, prev_out, threshold=20.0)
    completeness = _compute_completeness(inp, out)
    score, label = _compute_readiness_score(completeness, flags)
    n_errors     = sum(1 for f in flags if f.severity == "error")
    n_warnings   = sum(1 for f in flags if f.severity == "warning")

    renew_pct = (inp.renew_elec_purchased + inp.self_gen_elec) / max(out.total_electricity, 1) * 100
    prev_co2  = out.total_co2  # placeholder for YoY
    prev_e    = out.total_energy
    if prev_out:
        yoy_co2 = yoy_change(out.total_co2, prev_out.total_co2) or 0
        yoy_e   = yoy_change(out.total_energy, prev_out.total_energy) or 0
    else:
        yoy_co2 = yoy_e = 0

    # -- Gauge + summary -------------------------------------------------------
    col_score, col_info = st.columns([1, 3])
    with col_score:
        score_color = "#00916E" if score >= 80 else "#D97706" if score >= 60 else "#DC2626"
        fig = go.Figure(go.Indicator(
            mode="gauge+number", value=score,
            number=dict(suffix="/100", font=dict(size=32, color=score_color)),
            gauge=dict(
                axis=dict(range=[0,100]),
                bar=dict(color=score_color, thickness=.25),
                steps=[
                    dict(range=[0,60],  color="#FEE2E2"),
                    dict(range=[60,80], color="#FEF3C7"),
                    dict(range=[80,100],color="#D1FAE5"),
                ],
                threshold=dict(line=dict(color="#065F46",width=3), thickness=.75, value=score)
            )
        ))
        fig.update_layout(height=200, margin=dict(l=10,r=10,t=10,b=10), paper_bgcolor="rgba(0,0,0,0)")
        apply_chart_animation(fig)
        st.plotly_chart(fig, use_container_width=True)

    with col_info:
        st.markdown(f"### Report Readiness Score: **{score} / 100**")
        st.caption(f"{sel_co} · {sel_yr} Reporting Year · "
                   f"{n_errors} error{'s' if n_errors!=1 else ''}, "
                   f"{n_warnings} warning{'s' if n_warnings!=1 else ''}")
        if score >= 90:
            st.success(f"✅ {label} — submission can be included in consolidated report.")
        elif score >= 70:
            st.warning(f"⚠️ {label} — resolve open items before submission.")
        else:
            st.error(f"❌ {label} — significant data gaps must be addressed.")

        # Key live KPIs at a glance
        k1,k2,k3,k4 = st.columns(4)
        k1.metric("Energy KPI",   f"{out.energy_kpi:.2f} GJ/t")
        k2.metric("CO₂ KPI",      f"{out.co2_kpi:.3f} t/t")
        k3.metric("Water KPI",    f"{out.water_kpi:.2f} m³/t")
        k4.metric("Renewable %",  f"{renew_pct:.1f}%")

    st.divider()

    # -- Completeness by section -----------------------------------------------
    st.markdown("#### Data completeness by section")
    cols = st.columns(3)
    for i, (label_s, pct) in enumerate(completeness.items()):
        color = "#00916E" if pct == 100 else "#D97706" if pct >= 60 else "#DC2626"
        with cols[i % 3]:
            with st.container(border=True):
                st.caption(label_s)
                st.progress(pct / 100, text=f"{pct}%")

    st.divider()

    # ── ESG Analyst Chat (replaces old LLM insights) ──────────────────────────
    st.markdown("#### dss+ ESG Analyst")
    st.caption(f"Ask about {sel_co} {sel_yr} or any TIP company — powered by local AI (Ollama).")

    try:
        from chatbot.chatbot_engine import ESGChatbot
        _bot_key = f"_readiness_bot_{st.session_state.get('user_name','dss')}"
        if _bot_key not in st.session_state:
            st.session_state[_bot_key] = ESGChatbot(
                st.session_state.get("user_name", "dss_user"))
        _bot = st.session_state[_bot_key]

        _ok, _status = _bot.copilot.is_available()

        # Status bar
        _dot   = "🟢" if _ok else "🔴"
        _slabel = _bot.copilot.provider_label() if _ok else "Ollama not running — open from system tray"
        st.markdown(
            f'<div style="font-size:12px;color:#6B7280;margin-bottom:8px">'
            f'{_dot} {_slabel}</div>',
            unsafe_allow_html=True,
        )

        if not _ok:
            st.info("Start Ollama from your system tray, then reload this page.")
        else:
            # Quick-context chips for this company
            _chat_key = f"ai_msgs_{sel_co}_{sel_yr}"
            if _chat_key not in st.session_state:
                st.session_state[_chat_key] = []

            _msgs = st.session_state[_chat_key]

            # Render message history
            for _i, _m in enumerate(_msgs):
                _av = "👤" if _m["role"] == "user" else "🤖"
                with st.chat_message(_m["role"], avatar=_av):
                    st.markdown(_m["content"])
                    if _m.get("figure"):
                        st.plotly_chart(_m["figure"], use_container_width=True,
                                        key=f"ai_fig_{_i}")

            # Suggestion chips on empty state
            if not _msgs:
                _sugs = [
                    f"Summarise {sel_co} ESG performance in {sel_yr}",
                    f"Why did CO₂ intensity change for {sel_co.split()[0]}?",
                    f"Chart water intake for {sel_co.split()[0]} 2016–{sel_yr}",
                    f"Compare {sel_co.split()[0]} vs sector average in {sel_yr}",
                ]
                _sc = st.columns(2)
                for _si, _s in enumerate(_sugs):
                    with _sc[_si % 2]:
                        if st.button(_s, key=f"ai_chip_{_si}", use_container_width=True):
                            st.session_state[_chat_key].append(
                                {"role": "user", "content": _s, "figure": None})
                            with st.spinner("Thinking…"):
                                _resp = _bot.chat(_s)
                            st.session_state[_chat_key].append({
                                "role": "assistant",
                                "content": _resp.text,
                                "figure": _resp.figure,
                            })
                            st.rerun()

            # Chat input
            _q = st.chat_input(
                f"Ask about {sel_co} {sel_yr} or any ESG metric…",
                key=f"ai_input_{sel_co}_{sel_yr}",
            )
            if _q:
                st.session_state[_chat_key].append(
                    {"role": "user", "content": _q, "figure": None})
                with st.chat_message("user", avatar="👤"):
                    st.markdown(_q)

                with st.chat_message("assistant", avatar="🤖"):
                    _placeholder = st.empty()
                    _acc = ""
                    for _chunk in _bot.copilot.call_stream(
                        user_message  = _q,
                        data_context  = _bot.context.build_context_str(_q),
                        history       = _bot.history,
                        system_prompt = _bot.system_prompt,
                    ):
                        _acc += _chunk
                        _placeholder.markdown(_acc + "▌")

                    _placeholder.markdown(_acc)

                    _spec = _bot.graph.extract_spec(_acc)
                    _fig  = None
                    if _spec and not _bot.context.df.empty:
                        _fig = _bot.graph.build(_spec, _bot.context.df)
                        if _fig:
                            st.plotly_chart(_fig, use_container_width=True,
                                            key=f"ai_resp_fig_{len(_msgs)}")

                    _clean = _bot.graph.strip_spec(_acc)

                _bot.history.append({"role": "user",      "content": _q})
                _bot.history.append({"role": "assistant",  "content": _clean})
                if len(_bot.history) > 12:
                    _bot.history = _bot.history[-12:]
                _bot.logger.log(_q, _clean, _bot.classifier.classify(_q),
                                had_chart=(_fig is not None))

                st.session_state[_chat_key].append({
                    "role": "assistant", "content": _clean, "figure": _fig})
                st.rerun()

            # Clear button
            if _msgs:
                if st.button("🗑 Clear conversation", key="ai_clear_conv"):
                    st.session_state[_chat_key] = []
                    _bot.clear_history()
                    st.rerun()

    except ImportError:
        st.info("Chatbot module not available. Ensure chatbot/ folder is present.")
    except Exception as _e:
        st.error(f"Chat error: {_e}")


# ─────────────────────────────────────────────────────────
# NEW PAGES
# ─────────────────────────────────────────────────────────

def page_home():
    """
    Client Home — 8 animated KPI cards, interactive trend charts, summary data table.
    This is the client's personal performance dashboard.
    """
    company  = st.session_state.user_company
    comp_hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    years     = sorted(dl.get_years(_CONSOLIDATED_DF, company))

    # ── Header: Welcome + year selector + Submit Data inline ─────────────────
    h_left, h_mid, h_right = st.columns([3, 1, 1])
    with h_left:
        st.markdown(section_header_html(
            f"Welcome, {st.session_state.user_name.split()[0]} 👋",
            f"{company} · Your Performance Dashboard",
        ), unsafe_allow_html=True)
    with h_mid:
        if years:
            sel_yr = st.selectbox("", sorted(years, reverse=True),
                                  key="home_yr", label_visibility="collapsed")
        else:
            sel_yr = CURR_YEAR
    with h_right:
        if st.button("📋 Submit Data", use_container_width=True, key="home_submit_btn"):
            st.session_state.page = "entry"
            st.rerun()

    if not years:
        st.markdown(empty_state_html("📊", "No data yet",
            "Submit your first KPI report to see your dashboard.",
            "→ Submit Data"), unsafe_allow_html=True)
        return

    from formula_engine import TemplateInputs as TI, calculate as calc
    valid = {f.name for f in TI.__dataclass_fields__.values()}

    step  = dl.get_step_data(comp_hist, sel_yr)
    clean = {k: v for k, v in step.items() if k in valid}
    inp   = TI(company=company, year=sel_yr, **clean)
    out   = calc(inp)

    prev_out = None
    if sel_yr - 1 in years:
        ps = dl.get_step_data(comp_hist, sel_yr - 1)
        pc = {k: v for k, v in ps.items() if k in valid}
        prev_out = calc(TI(company=company, year=sel_yr - 1, **pc))

    # ── Submission status strip — data completeness + DSS+ verification state ───
    status_hist  = dl.get_company_hist(_CONSOLIDATED_DF, company)
    step_data_yr = dl.get_step_data(status_hist, sel_yr) if status_hist else {}

    def _has(key, min_val=1):
        v = step_data_yr.get(key, 0)
        try: return float(v) >= min_val
        except: return bool(v)

    section_done = [
        _has("total_sites"),
        _has("production"),
        _has("water_withdrawals"),
        _has("renew_elec_purchased") or _has("nonrenew_elec_purchased") or _has("nat_gas"),
        step_data_yr.get("co2_scope2_steam") is not None and _has("production"),
        _has("waste_total"),
    ]
    n_done = sum(section_done)
    pct    = n_done / 6 * 100
    sc     = GREEN if pct == 100 else (AMBER if pct >= 50 else RED)

    # Check DSS+ verification status from persistent CSV
    verif_status = "Not Submitted"
    verif_color  = "#94A3B8"
    verif_icon   = "○"
    try:
        from pathlib import Path
        vcsv = Path("data_storage/verifications.csv")
        if vcsv.exists():
            import csv
            with open(vcsv, newline="") as f:
                for row in csv.DictReader(f):
                    if row.get("Company","").strip() == company and str(row.get("Year","")).strip() == str(sel_yr):
                        vs = row.get("Status","").strip()
                        if vs == "Verified":
                            verif_status = "Verified by dss+"; verif_color = GREEN; verif_icon = "✓"
                        elif vs == "Pending":
                            verif_status = "Pending Review";   verif_color = AMBER; verif_icon = "◉"
                        elif vs == "Flagged":
                            verif_status = "Flagged — see notes"; verif_color = RED; verif_icon = "⚑"
        elif n_done > 0:
            verif_status = "Pending Review"; verif_color = AMBER; verif_icon = "◉"
    except Exception:
        pass

    st.markdown(f"""
    <div style="background:#fff;border:1px solid {BORDER};border-radius:8px;
        padding:12px 20px;margin-bottom:16px;display:flex;align-items:center;gap:16px">
      <div style="flex:1">
        <div style="font-size:12px;color:{MUTED};margin-bottom:5px">
          {sel_yr} Submission Status</div>
        <div style="background:#F1F5F9;border-radius:4px;height:6px;overflow:hidden">
          <div style="background:{sc};width:{pct:.0f}%;height:100%;border-radius:4px;
              transition:width .8s ease"></div>
        </div>
      </div>
      <div style="font-size:18px;font-weight:700;color:{sc}">{n_done}/6</div>
      <div style="font-size:12px;color:{MUTED}">sections complete</div>
      <div style="border-left:1px solid {BORDER};padding-left:16px;font-size:12px;
          color:{verif_color};font-weight:600;white-space:nowrap">
        {verif_icon} {verif_status}</div>
    </div>""", unsafe_allow_html=True)

    # ── 8 KPI cards (4 × 2) ──────────────────────────────────────────────────
    renew_tot = max(inp.renew_elec_purchased + inp.nonrenew_elec_purchased + inp.self_gen_elec, 1)
    renew_pct = inp.renew_elec_purchased / renew_tot * 100

    def _yoy(cur, prev_val, lower=True):
        if not prev_val or prev_val == 0: return "", ""
        pct = (cur - prev_val) / abs(prev_val) * 100
        good = pct <= 0 if lower else pct >= 0
        bg   = "#DCFCE7" if good else "#FEE2E2"
        col  = "#166534" if good else "#991B1B"
        arr  = "▼" if pct < 0 else "▲"
        sign = "+" if pct > 0 else ""
        chip = (f'<span style="background:{bg};color:{col};font-size:10px;font-weight:600;'
                f'padding:2px 7px;border-radius:4px">{arr}{sign}{pct:.1f}%</span>')
        return chip, bg

    p = prev_out
    cards = [
        ("CO₂ Absolute",      f"{out.total_co2:,.0f}",           "tCO₂",  *_yoy(out.total_co2, p.total_co2 if p else 0),       CAT_CO2),
        ("CO₂ Intensity",     f"{out.co2_kpi:.3f}",              "t/t",    *_yoy(out.co2_kpi,   p.co2_kpi   if p else 0),       CAT_CO2),
        ("Energy Intensity",  f"{out.energy_kpi:.2f}",           "GJ/t",   *_yoy(out.energy_kpi,p.energy_kpi if p else 0),      CAT_ENERGY),
        ("Renewable Share",   f"{renew_pct:.1f}",                "%",      *_yoy(renew_pct, 0, lower=False),                    CAT_RENEW),
        ("Water Intensity",   f"{out.water_kpi:.2f}",            "m³/t",   *_yoy(out.water_kpi, p.water_kpi if p else 0),       CAT_WATER),
        ("Water Withdrawal",  f"{inp.water_withdrawals:,.0f}",   "m³",     *_yoy(inp.water_withdrawals, 0),                     CAT_WATER),
        ("Waste Recovery",    f"{out.waste_recovery_pct*100:.1f}","%",     *_yoy(out.waste_recovery_pct*100, (p.waste_recovery_pct*100 if p else 0), lower=False), CAT_WASTE),
        ("ISO 14001",         f"{out.pct_certified*100:.0f}",    "%",      *_yoy(out.pct_certified*100, (p.pct_certified*100 if p else 0), lower=False), GREEN),
    ]
    COLORS_CARD = [CAT_CO2,CAT_CO2,CAT_ENERGY,CAT_RENEW,CAT_WATER,CAT_WATER,CAT_WASTE,GREEN]

    for row_start in [0, 4]:
        cols = st.columns(4)
        for i, (label, val_str, unit, chip_html, _, color) in enumerate(cards[row_start:row_start+4]):
            with cols[i]:
                st.markdown(f"""
                <div style="background:#fff;border:1px solid {BORDER};border-radius:10px;
                    padding:16px 18px 14px;margin-bottom:8px;height:110px;
                    display:flex;flex-direction:column;justify-content:space-between;
                    animation:tipFadeIn 400ms ease-out {i*70+row_start*30}ms both;
                    transition:box-shadow 200ms,transform 200ms"
                    onmouseover="this.style.boxShadow='0 6px 20px rgba(15,23,42,.1)';this.style.transform='translateY(-2px)'"
                    onmouseout="this.style.boxShadow='';this.style.transform=''">
                  <div style="font-size:10.5px;font-weight:600;color:{MUTED};
                      text-transform:uppercase;letter-spacing:.6px">{label}</div>
                  <div style="font-size:26px;font-weight:700;color:{color};
                      font-variant-numeric:tabular-nums;line-height:1;letter-spacing:-.5px;
                      white-space:nowrap;overflow:hidden;text-overflow:ellipsis">
                    {val_str}
                    <span style="font-size:11px;font-weight:400;color:{MUTED};margin-left:2px">{unit}</span>
                  </div>
                  <div style="margin-top:2px">{chip_html}</div>
                </div>""", unsafe_allow_html=True)

    st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

    # ── Interactive charts ────────────────────────────────────────────────────
    t1, t2, t3, t4 = st.tabs(["📈 CO₂ Trend", "⚡ Energy Mix", "💧 Water", "♻️ Waste & Fuel"])

    # Build multi-year computed KPIs for charts
    yr_kpis = {}
    for y in years:
        sd = dl.get_step_data(comp_hist, y)
        sc = {k: v for k, v in sd.items() if k in valid}
        o  = calc(TI(company=company, year=y, **sc))
        ii = TI(company=company, year=y, **sc)
        rt = max(ii.renew_elec_purchased + ii.nonrenew_elec_purchased + ii.self_gen_elec, 1)
        yr_kpis[y] = {
            "scope1": o.total_co2_scope1, "scope2": o.total_co2_scope2,
            "total_co2": o.total_co2, "co2_kpi": o.co2_kpi,
            "energy_kpi": o.energy_kpi, "water_kpi": o.water_kpi,
            "waste_pct": o.waste_recovery_pct * 100,
            "renew_pct": ii.renew_elec_purchased / rt * 100,
            "nat_gas": ii.nat_gas, "coal": ii.coal_sub, "diesel": ii.diesel,
            "biomass": ii.biomass, "renew_elec": ii.renew_elec_purchased,
            "nonrenew_elec": ii.nonrenew_elec_purchased,
            "water_m3": ii.water_withdrawals, "production": ii.production,
        }

    ys = years

    with t1:
        # Stacked area: Scope 1 + Scope 2 with intensity overlay
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=ys, y=[yr_kpis[y]["scope2"] for y in ys],
            name="Scope 2", fill="tonexty", stackgroup="co2",
            mode="none", fillcolor="rgba(71,85,105,0.25)",
            hovertemplate="<b>%{x}</b> · Scope 2<br>%{y:,.0f} tCO₂<extra></extra>",
        ))
        fig.add_trace(go.Scatter(
            x=ys, y=[yr_kpis[y]["scope1"] for y in ys],
            name="Scope 1", fill="tonexty", stackgroup="co2",
            mode="none", fillcolor="rgba(71,85,105,0.5)",
            hovertemplate="<b>%{x}</b> · Scope 1<br>%{y:,.0f} tCO₂<extra></extra>",
        ))
        # Intensity as secondary line
        fig.add_trace(go.Scatter(
            x=ys, y=[yr_kpis[y]["co2_kpi"] for y in ys],
            name="CO₂ Intensity (t/t)", yaxis="y2",
            mode="lines+markers",
            line=dict(color="#8B0000", width=2.5, dash="dot"),
            marker=dict(size=6, color="#8B0000"),
            hovertemplate="<b>%{x}</b><br>Intensity: %{y:.3f} t/t<extra></extra>",
        ))
        # Annotate best/worst year
        if len(ys) >= 2:
            best_y = min(ys, key=lambda y: yr_kpis[y]["co2_kpi"])
            fig.add_annotation(x=best_y, y=yr_kpis[best_y]["co2_kpi"], yref="y2",
                               text="Best", showarrow=True, arrowhead=2,
                               ax=0, ay=-30, font=dict(size=10, color=GREEN),
                               arrowcolor=GREEN)
        fig.update_layout(
            **chart_layout_defaults("Total CO₂ Emissions (Scope 1 + 2) with Intensity", height=320),
            yaxis=dict(title=dict(text="tCO₂", font=dict(color=CAT_CO2)), tickformat=",", showline=True, linecolor="#999", showticklabels=True, tickfont=dict(size=12, color="#1C2E3F")),
            yaxis2=dict(title=dict(text="tCO₂/t", font=dict(color="#C8102E")),
                        overlaying="y", side="right", tickformat=".3f"),
            hovermode="x unified",
        )
        apply_chart_animation(fig)
        st.plotly_chart(fig, use_container_width=True)

    with t2:
        fuel_cfg = [
            ("Renewable Elec.",    [yr_kpis[y]["renew_elec"]    for y in ys], CAT_RENEW),
            ("Non-Renew. Elec.",   [yr_kpis[y]["nonrenew_elec"] for y in ys], "#94A3B8"),
            ("Natural Gas",        [yr_kpis[y]["nat_gas"]        for y in ys], CAT_ENERGY),
            ("Coal",               [yr_kpis[y]["coal"]           for y in ys], "#475569"),
            ("Diesel",             [yr_kpis[y]["diesel"]         for y in ys], "#78716C"),
            ("Biomass",            [yr_kpis[y]["biomass"]        for y in ys], "#16A34A"),
        ]
        fig2 = go.Figure()
        for label, vals, color in fuel_cfg:
            if any(v > 0 for v in vals):
                fig2.add_trace(go.Bar(
                    name=label, x=ys, y=vals, marker_color=color,
                    marker_line_width=0,
                    hovertemplate=f"<b>{label}</b><br>%{{x}}: %{{y:,.0f}} GJ<br>%{{customdata:.1f}}% of total<extra></extra>",
                ))
        fig2.update_layout(
            barmode="stack",
            **chart_layout_defaults("Energy Mix by Source (GJ)", height=320),
            hovermode="x unified",
            bargap=0.3,
        )
        apply_chart_animation(fig2)
        st.plotly_chart(fig2, use_container_width=True)

    with t3:
        w_m3  = [yr_kpis[y]["water_m3"]  for y in ys]
        w_kpi = [yr_kpis[y]["water_kpi"] for y in ys]
        fig3  = go.Figure()
        fig3.add_trace(go.Bar(
            x=ys, y=w_m3, name="Total Withdrawals",
            marker_color=CAT_WATER, marker_line_width=0,
            opacity=0.8,
            hovertemplate="<b>%{x}</b><br>%{y:,.0f} m³<extra></extra>",
        ))
        fig3.add_trace(go.Scatter(
            x=ys, y=w_kpi, name="Intensity (m³/t)",
            yaxis="y2", mode="lines+markers",
            line=dict(color="#0E7490", width=2.5),
            marker=dict(size=7, color="#0E7490", symbol="diamond"),
            hovertemplate="<b>%{x}</b><br>Intensity: %{y:.2f} m³/t<extra></extra>",
        ))
        fig3.update_layout(
            **chart_layout_defaults("Water Withdrawals & Intensity", height=320),
            yaxis=dict(title=dict(text="m³", font=dict(color=CAT_WATER)), tickformat=",", showline=True, linecolor="#999", showticklabels=True, tickfont=dict(size=12, color="#1C2E3F")),
            yaxis2=dict(title=dict(text="m³/t (intensity)", font=dict(color="#0E7490")),
                        overlaying="y", side="right", tickformat=".2f"),
            hovermode="x unified",
        )
        apply_chart_animation(fig3)
        st.plotly_chart(fig3, use_container_width=True)

    with t4:
        w_total    = [dl.get_step_data(comp_hist, y).get("waste_total",    0) for y in ys]
        w_recovery = [dl.get_step_data(comp_hist, y).get("waste_recovery", 0) for y in ys]
        w_pcts     = [yr_kpis[y]["waste_pct"]   for y in ys]

        c1, c2 = st.columns([2, 1])
        with c1:
            fig4 = go.Figure()
            fig4.add_trace(go.Bar(
                x=ys, y=w_total, name="Total Waste",
                marker_color="#E2E8F0", marker_line_width=0, opacity=0.8,
            ))
            fig4.add_trace(go.Bar(
                x=ys, y=w_recovery, name="Recovered",
                marker_color=CAT_WASTE, marker_line_width=0,
            ))
            fig4.add_trace(go.Scatter(
                x=ys, y=w_pcts, name="Recovery %",
                yaxis="y2", mode="lines+markers",
                line=dict(color="#6D28D9", width=2.5),
                marker=dict(size=7, color="#6D28D9"),
                hovertemplate="Recovery: %{y:.1f}%<extra></extra>",
            ))
            fig4.update_layout(
                barmode="overlay",
                **chart_layout_defaults("Waste Recovery (T)", height=300),
                yaxis=dict(title="Metric t", tickformat=",", showline=True, linecolor="#999", showticklabels=True, tickfont=dict(size=12, color="#1C2E3F")),
                yaxis2=dict(title="Recovery %", overlaying="y", side="right",
                            range=[0, 110], ticksuffix="%"),
                hovermode="x unified",
            )
            apply_chart_animation(fig4)
            st.plotly_chart(fig4, use_container_width=True)

        with c2:
            # Waste recovery % trend — more useful than a static gauge
            rec_pcts = [yr_kpis[y]["waste_pct"] for y in ys]
            fig_rec = go.Figure()
            # Background zones
            fig_rec.add_hrect(y0=0,  y1=70,  fillcolor="#FEE2E2", opacity=0.25, line_width=0)
            fig_rec.add_hrect(y0=70, y1=85,  fillcolor="#FEF3C7", opacity=0.25, line_width=0)
            fig_rec.add_hrect(y0=85, y1=100, fillcolor="#DCFCE7", opacity=0.25, line_width=0)
            # 90% best-practice target line
            fig_rec.add_hline(y=90, line_dash="dot", line_color=GREEN,
                              line_width=1.5, annotation_text="Target 90%",
                              annotation_font=dict(size=9, color=GREEN))
            # Recovery trend
            fig_rec.add_trace(go.Scatter(
                x=ys, y=rec_pcts, mode="lines+markers",
                fill="tozeroy",
                fillcolor="rgba(124,58,237,0.10)",
                line=dict(color=CAT_WASTE, width=2.5),
                marker=dict(size=7, color=CAT_WASTE, symbol="circle",
                            line=dict(color="white", width=1.5)),
                hovertemplate="<b>%{x}</b><br>Recovery: %{y:.1f}%<extra></extra>",
                name="Recovery %",
            ))
            fig_rec.update_layout(
                **chart_layout_defaults("Waste Recovery Trend (%)", height=300,
                                        showlegend=False),
                yaxis=dict(range=[0, 105], ticksuffix="%", gridcolor="#F1F5F9", showline=True, linecolor="#999", showticklabels=True, tickfont=dict(size=12, color="#1C2E3F"),
                           zeroline=False),
                xaxis=dict(gridcolor="#F1F5F9"),
            )
            apply_chart_animation(fig_rec)
            st.plotly_chart(fig_rec, use_container_width=True)

    # ── Historical KPI summary table ──────────────────────────────────────────
    st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
    st.markdown(f"**Historical KPI Summary — {company}**")

    import pandas as pd
    tbl_rows = []
    table_years = sorted([y for y in years if 2014 <= y <= 2023], reverse=True)
    for y in table_years:
        sd = dl.get_step_data(comp_hist, y)
        sc = {k: v for k, v in sd.items() if k in valid}
        o  = calc(TI(company=company, year=y, **sc))
        ii = TI(company=company, year=y, **sc)
        rt = max(ii.renew_elec_purchased + ii.nonrenew_elec_purchased + ii.self_gen_elec, 1)
        tbl_rows.append({
            "Year":              y,
            "Production (MT)":   f"{ii.production/1e6:.3f}",
            "CO₂ Total (T)":     f"{o.total_co2:,.0f}",
            "CO₂ Intensity":     f"{o.co2_kpi:.3f}",
            "Energy KPI (GJ/t)": f"{o.energy_kpi:.2f}",
            "Renew. Elec. %":    f"{ii.renew_elec_purchased/rt*100:.1f}%",
            "Water KPI (m³/t)":  f"{o.water_kpi:.2f}",
            "Waste Recovery %":  f"{o.waste_recovery_pct*100:.1f}%",
        })
    tbl_df = pd.DataFrame(tbl_rows)
    st.dataframe(
        tbl_df.style
            .set_properties(**{"text-align": "right", "font-size": "12px"})
            .set_table_styles([
                {"selector": "th", "props": [
                    ("font-size","11px"), ("text-transform","uppercase"),
                    ("letter-spacing",".4px"), ("color","#64748B"),
                    ("background","#F8FAFC"), ("padding","8px 12px"),
                ]},
                {"selector": "td:first-child", "props": [
                    ("font-weight","600"), ("color","#0F172A"), ("text-align","center"),
                ]},
            ]),
        use_container_width=True, hide_index=True,
    )

def page_my_dashboard():
    """
    Client My Dashboard — Sector analysis view.
    """
    company = st.session_state.user_company
    st.markdown(section_header_html(
        "My Dashboard",
        f"TIP Sector Analysis · {company} highlighted",
    ), unsafe_allow_html=True)

    if _CONSOLIDATED_DF.empty or _SECTOR_DF.empty:
        st.info("Sector data not loaded. Run build_esg_master.py first.")
        return

    # Define has_wide locally (same as in page_analysis)
    df      = _CONSOLIDATED_DF
    has_wide = not df.empty and "Row_Label" not in df.columns

    # ── Year range selector ───────────────────────────────────────────────────
    col_yr1, col_yr2, col_toggle, _ = st.columns([1, 1, 2, 2])
    with col_yr1:
        _dash_data_yrs = sorted(
            _CONSOLIDATED_DF["Year"].dropna().unique().astype(int).tolist()
            if not _CONSOLIDATED_DF.empty else LONG_YEARS
        ) or list(LONG_YEARS)
        yr_start = st.selectbox("From", _dash_data_yrs, index=0, key="dash_yr_start")
    with col_yr2:
        yr_end   = st.selectbox("To", _dash_data_yrs[::-1], index=0, key="dash_yr_end")
    with col_toggle:
        show_company = st.toggle(f"Highlight {company.split()[0]}", value=True, key="dash_highlight")

    yr_range = [y for y in _dash_data_yrs if yr_start <= y <= yr_end]
    if not yr_range:
        yr_range = _dash_data_yrs

    # Sector data for the range
    sec_range = _SECTOR_DF[_SECTOR_DF["Year"].isin(yr_range)].sort_values("Year")

    # Company overlay data
    comp_hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    from formula_engine import TemplateInputs as TI, calculate as calc
    valid = {f.name for f in TI.__dataclass_fields__.values()}
    co_kpis = {}
    if show_company:
        for y in yr_range:
            sd = dl.get_step_data(comp_hist, y)
            sc = {k: v for k, v in sd.items() if k in valid}
            if sc:
                o = calc(TI(company=company, year=y, **sc))
                co_kpis[y] = {
                    "co2_kpi": o.co2_kpi, "energy_kpi": o.energy_kpi,
                    "water_kpi": o.water_kpi, "total_co2": o.total_co2,
                }

    # ── 4-metric summary row ──────────────────────────────────────────────────
    if not sec_range.empty:
        latest_sec = sec_range.iloc[-1]
        latest_yr  = int(latest_sec["Year"])
        prev_sec   = sec_range.iloc[-2] if len(sec_range) > 1 else None

        metric_cols = st.columns(4)
        metrics = [
            ("Sector CO₂ Intensity", "Avg_CO2_KPI",    ".3f", "tCO₂/t"),
            ("Sector Energy KPI",    "Avg_Energy_KPI", ".2f", "GJ/t"),
            ("Sector Water KPI",     "Avg_Water_KPI",  ".2f", "m³/t"),
            ("Avg Renewable %",      "Avg_Renewable_Share", ".1f", "%"),
        ]
        for i, (label, col, fmt, unit) in enumerate(metrics):
            with metric_cols[i]:
                val  = latest_sec.get(col, 0)
                prev = prev_sec.get(col, 0) if prev_sec is not None else None
                delta = (f"{(val-prev)/abs(prev)*100:+.1f}%" if prev and prev != 0 else None)
                st.metric(label, f"{val:{fmt}} {unit}", delta=delta)

    st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)

    # ── Charts ────────────────────────────────────────────────────────────────
    r1c1, r1c2 = st.columns(2, gap="medium")

    with r1c1:
        # Sector total CO₂ bar + company line overlay
        fig1 = go.Figure()
        fig1.add_trace(go.Bar(
            x=sec_range["Year"].tolist(),
            y=sec_range["Total_CO2"].tolist() if "Total_CO2" in sec_range else [],
            name="Sector Total CO₂",
            marker_color="#CBD5E1", marker_line_width=0, opacity=0.75,
            hovertemplate="<b>%{x}</b><br>Sector: %{y:,.0f} tCO₂<extra></extra>",
        ))
        if co_kpis and show_company:
            co_yrs = sorted(co_kpis.keys())
            fig1.add_trace(go.Scatter(
                x=co_yrs, y=[co_kpis[y]["total_co2"] for y in co_yrs],
                name=company.split()[0], mode="lines+markers",
                line=dict(color=CAT_CO2, width=2.5),
                marker=dict(size=7, color=CAT_CO2, symbol="circle"),
                yaxis="y2",
                hovertemplate="<b>%{x}</b><br>Your CO₂: %{y:,.0f} T<extra></extra>",
            ))
        fig1.update_layout(
            **chart_layout_defaults("Sector CO₂ vs Your Performance", height=300),
            yaxis2=dict(overlaying="y", side="right",
                        title=dict(text=f"{company.split()[0]} CO₂", font=dict(color=CAT_CO2, size=10))),
            hovermode="x unified",
        )
        apply_chart_animation(fig1)
        st.plotly_chart(fig1, use_container_width=True)

    with r1c2:
        # CO₂ intensity trend — sector average vs company
        fig2 = go.Figure()
        if "Avg_CO2_KPI" in sec_range.columns:
            fig2.add_trace(go.Scatter(
                x=sec_range["Year"].tolist(), y=sec_range["Avg_CO2_KPI"].tolist(),
                name="Sector Average", mode="lines+markers",
                line=dict(color="#94A3B8", width=2, dash="dot"),
                marker=dict(size=5, color="#94A3B8"),
                fill="tozeroy", fillcolor="rgba(148,163,184,0.08)",
                hovertemplate="Sector avg: %{y:.3f}<extra></extra>",
            ))
        if co_kpis and show_company:
            co_yrs = sorted(co_kpis.keys())
            fig2.add_trace(go.Scatter(
                x=co_yrs, y=[co_kpis[y]["co2_kpi"] for y in co_yrs],
                name=company.split()[0],
                mode="lines+markers",
                line=dict(color="#8B0000", width=3.5),
                marker=dict(size=9, color="#8B0000", symbol="diamond",
                            line=dict(color="white", width=1.5)),
                hovertemplate=f"{company.split()[0]}: %{{y:.3f}}<extra></extra>",
            ))
        fig2.update_layout(
            **chart_layout_defaults("CO₂ Intensity Trend (tCO₂/t)", height=300),
            hovermode="x unified",
        )
        apply_chart_animation(fig2)
        st.plotly_chart(fig2, use_container_width=True)

    r2c1, r2c2 = st.columns(2, gap="medium")

    with r2c1:
        # Energy & renewable share
        fig3 = go.Figure()
        if "Avg_Energy_KPI" in sec_range.columns:
            fig3.add_trace(go.Bar(
                x=sec_range["Year"].tolist(), y=sec_range["Avg_Energy_KPI"].tolist(),
                name="Sector Energy KPI", marker_color=CAT_ENERGY, marker_line_width=0, opacity=0.8,
                hovertemplate="<b>%{x}</b><br>Sector avg: %{y:.2f} GJ/t<extra></extra>",
            ))
        if co_kpis and show_company:
            co_yrs = sorted(co_kpis.keys())
            fig3.add_trace(go.Scatter(
                x=co_yrs, y=[co_kpis[y]["energy_kpi"] for y in co_yrs],
                name=company.split()[0], mode="lines+markers",
                line=dict(color="#5C2700", width=3.5),
                marker=dict(size=8, color="#5C2700",
                            line=dict(color="white", width=1.5)),
                hovertemplate=f"{company.split()[0]}: %{{y:.2f}} GJ/t<extra></extra>",
            ))
        if "Avg_Renewable_Share" in sec_range.columns:
            fig3.add_trace(go.Scatter(
                x=sec_range["Year"].tolist(), y=sec_range["Avg_Renewable_Share"].tolist(),
                name="Renewable Share %", mode="lines", yaxis="y2",
                line=dict(color=CAT_RENEW, width=1.5, dash="longdash"),
                hovertemplate="Renew. %: %{y:.1f}%<extra></extra>",
            ))
        fig3.update_layout(
            **chart_layout_defaults("Energy KPI vs Renewable Share", height=300),
            yaxis2=dict(overlaying="y", side="right", ticksuffix="%",
                        title=dict(font=dict(color=CAT_RENEW, size=10))),
            hovermode="x unified",
        )
        apply_chart_animation(fig3)
        st.plotly_chart(fig3, use_container_width=True)

    with r2c2:
        # Water intensity trend
        fig4 = go.Figure()
        if "Avg_Water_KPI" in sec_range.columns:
            fig4.add_trace(go.Scatter(
                x=sec_range["Year"].tolist(), y=sec_range["Avg_Water_KPI"].tolist(),
                name="Sector Average", mode="lines",
                line=dict(color=CAT_WATER, width=2),
                fill="tozeroy", fillcolor="rgba(8,145,178,0.10)",
                hovertemplate="Sector avg: %{y:.2f}<extra></extra>",
            ))
        if co_kpis and show_company:
            co_yrs = sorted(co_kpis.keys())
            fig4.add_trace(go.Scatter(
                x=co_yrs, y=[co_kpis[y]["water_kpi"] for y in co_yrs],
                name=company.split()[0], mode="lines+markers",
                line=dict(color="#0C4A6E", width=3.5),
                marker=dict(size=8, color="#0C4A6E", symbol="square",
                            line=dict(color="white", width=1.5)),
                hovertemplate=f"{company.split()[0]}: %{{y:.2f}} m³/t<extra></extra>",
            ))
        fig4.update_layout(
            **chart_layout_defaults("Water Intensity Trend (m³/t)", height=300),
            hovermode="x unified",
        )
        apply_chart_animation(fig4)
        st.plotly_chart(fig4, use_container_width=True)

    # ── Sector production & companies chart ───────────────────────────────────
    if "Total_Production" in sec_range.columns:
        fig5 = go.Figure()
        fig5.add_trace(go.Scatter(
            x=sec_range["Year"].tolist(),
            y=(sec_range["Total_Production"] / 1e6).tolist(),
            mode="lines+markers",
            fill="tozeroy", fillcolor="rgba(22,163,74,0.08)",
            line=dict(color=GREEN, width=2.5),
            marker=dict(size=6, color=GREEN),
            name="TIP Total Production",
            hovertemplate="<b>%{x}</b><br>%{y:.2f} million T<extra></extra>",
        ))
        fig5.update_layout(**chart_layout_defaults(
            "TIP Sector Total Production (million metric t)", height=220, showlegend=False))
        apply_chart_animation(fig5)
        st.plotly_chart(fig5, use_container_width=True,
                        key=_chart_key(company, "dash_prod"))

    # ── Additional client-facing charts ───────────────────────────────────────
    st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
    st.markdown(f"**{company.split()[0]} — Company-specific trends vs sector**")
    st.caption("Your data highlighted · Toggle 'Highlight' to show/hide company overlay")

    r3c1, r3c2 = st.columns(2, gap="medium")

    with r3c1:
        # Renewable electricity trend for this company vs sector
        sec_renew_mean, sec_renew_q25, sec_renew_q75 = {}, {}, {}
        if has_wide and "Renewable_Electricity_Share_%" in sec_range.columns:
            sec_renew_mean = sec_range.set_index("Year")["Renewable_Electricity_Share_%"].to_dict()
        if not _CONSOLIDATED_DF.empty and "Renewable_Electricity_Share_%" in _CONSOLIDATED_DF.columns:
            grp = _CONSOLIDATED_DF.groupby("Year")["Renewable_Electricity_Share_%"]
            sec_renew_q25 = grp.quantile(.25).to_dict()
            sec_renew_q75 = grp.quantile(.75).to_dict()

        fig6 = go.Figure()
        if sec_renew_q25:
            fig6.add_trace(go.Scatter(x=yr_range, y=[sec_renew_q75.get(y) for y in yr_range],
                fill=None, mode="lines", line=dict(width=0), showlegend=False))
            fig6.add_trace(go.Scatter(x=yr_range, y=[sec_renew_q25.get(y) for y in yr_range],
                fill="tonexty", mode="lines", line=dict(width=0),
                fillcolor="rgba(22,163,74,0.10)", name="Sector IQR"))
            fig6.add_trace(go.Scatter(x=yr_range, y=[sec_renew_mean.get(y) for y in yr_range],
                mode="lines", name="Sector Median",
                line=dict(color="#94A3B8", width=1.5, dash="dashdot")))
        if show_company and co_kpis:
            fig6.add_trace(go.Scatter(
                x=sorted(co_kpis.keys()),
                y=[co_kpis[y].get("renew_pct", 0) for y in sorted(co_kpis.keys())],
                mode="lines+markers", name=company.split()[0],
                line=dict(color=CAT_RENEW, width=2.5),
                marker=dict(size=7, color=CAT_RENEW),
                hovertemplate="<b>%{x}</b><br>Renewable: %{y:.1f}%<extra></extra>",
            ))
        fig6.update_layout(**chart_layout_defaults("Renewable Electricity Share (%)", height=270),
                           yaxis=dict(ticksuffix="%", gridcolor="#F1F5F9", showline=True, linecolor="#999", showticklabels=True, tickfont=dict(size=12, color="#1C2E3F")),
                           xaxis=dict(gridcolor="#F1F5F9"), hovermode="x unified")
        apply_chart_animation(fig6)
        st.plotly_chart(fig6, use_container_width=True,
                        key=_chart_key(company, sel_yr if 'sel_yr' in dir() else 0, "renew_dash"))

    with r3c2:
        # YoY CO₂ change bar chart for this company
        if co_kpis and len(sorted(co_kpis.keys())) >= 2:
            co_yrs_sorted = sorted(co_kpis.keys())
            yoy_bars  = []
            yoy_years = []
            for j in range(1, len(co_yrs_sorted)):
                y_cur  = co_yrs_sorted[j]
                y_prev = co_yrs_sorted[j-1]
                cur_v  = co_kpis.get(y_cur, {}).get("co2_kpi", 0)
                prv_v  = co_kpis.get(y_prev, {}).get("co2_kpi", 0)
                if prv_v and prv_v != 0:
                    yoy_bars.append((cur_v - prv_v) / abs(prv_v) * 100)
                    yoy_years.append(y_cur)
            if yoy_bars:
                bar_colors = [CAT_RENEW if v < 0 else RED for v in yoy_bars]
                fig7 = go.Figure(go.Bar(
                    x=yoy_years, y=yoy_bars,
                    marker_color=bar_colors, marker_line_width=0,
                    hovertemplate="<b>%{x}</b><br>YoY: %{y:+.2f}%<extra></extra>",
                ))
                fig7.add_hline(y=0, line_color="#CBD5E1", line_width=1)
                fig7.update_layout(**chart_layout_defaults(
                    f"CO₂ Intensity YoY Change (%) — {company.split()[0]}", height=270,
                    showlegend=False),
                    yaxis=dict(ticksuffix="%", gridcolor="#F1F5F9", zeroline=False, showline=True, linecolor="#999", showticklabels=True, tickfont=dict(size=12, color="#1C2E3F")),
                    xaxis=dict(gridcolor="#F1F5F9"))
                apply_chart_animation(fig7)
                st.plotly_chart(fig7, use_container_width=True,
                                key=_chart_key(company, "yoy_co2_dash"))

def page_company_data():
    """
    DSS+ Company Data — full KPI template table for a selected company.
    Accessed from Portfolio 'Open Template' button or directly from nav.
    """
    st.markdown(section_header_html(
        "Company Data",
        "Full KPI template for selected company · All historical years",
    ), unsafe_allow_html=True)

    companies_in_db = dl.get_companies(_CONSOLIDATED_DF) or COMPANIES

    # Pre-select company from portfolio if set
    pre_co = st.session_state.pop("portfolio_company", None)
    default_co = (pre_co
                  or st.session_state.get("reporting_company")
                  or companies_in_db[0])
    if default_co not in companies_in_db:
        default_co = companies_in_db[0]

    col_co, col_yr, _ = st.columns([2, 1, 3])
    with col_co:
        sel_co = st.selectbox(
            "Company", options=companies_in_db,
            index=companies_in_db.index(default_co),
            key="codata_company"
        )
    with col_yr:
        avail_years = dl.get_years(_CONSOLIDATED_DF, sel_co) or [CURR_YEAR]
        sel_yr = st.selectbox(
            "Year", options=sorted(avail_years, reverse=True),
            key="codata_year"
        )

    # Set session state so all render_*_tab() functions read the right data
    st.session_state.reporting_company  = sel_co
    st.session_state.reporting_year     = sel_yr

    # Load and populate session state with this company's data
    hist    = dl.get_company_hist(_CONSOLIDATED_DF, sel_co)
    step_data = dl.get_step_data(hist, sel_yr) if hist else {}
    valid_fields = _VALID_TEMPLATE_FIELDS

    for field, val in step_data.items():
        if field in valid_fields:
            st.session_state[field] = val

    from formula_engine import TemplateInputs as TI, calculate as calc
    valid = {f.name for f in TI.__dataclass_fields__.values()}
    clean = {k: v for k, v in step_data.items() if k in valid}
    inp   = TI(company=sel_co, year=sel_yr, **clean)
    out   = calc(inp)

    st.session_state["_codata_inp"] = inp
    st.session_state["_codata_out"] = out
    st.session_state["template_done"]       = True
    st.session_state["company_setup_done"]  = True
    st.session_state["step"]                = 6

    # ── Render all template sheets as tabs ────────────────────────────────────
    tab_main, tab_elec, tab_waste, tab_qual, tab_conv = st.tabs([
        "Main Data Input",
        "Electricity by Country",
        "Waste",
        "Qualitative Data",
        "Conversion Tables",
    ])
    with tab_main:
        render_template_table()
    with tab_elec:
        render_electricity_tab()
    with tab_waste:
        render_waste_tab()
    with tab_qual:
        render_qualitative_tab()
    with tab_conv:
        render_conversion_tab()

    st.markdown('<div style="height:12px"></div>', unsafe_allow_html=True)
    if st.button("← Back to Portfolio", key="codata_back"):
        st.session_state.page = "portfolio"
        st.rerun()



def page_reports():
    """
    Sustainability Report — one-page CSR report with company summary,
    KPI tables, trend charts, benchmarking position. Download as PDF.
    """
    from pdf_report import generate_executive_pdf, build_kpi_dict_from_outputs, REPORTLAB_OK
    from datetime import date as _date

    company   = st.session_state.user_company
    comp_hist = dl.get_company_hist(_CONSOLIDATED_DF, company)
    years     = sorted(dl.get_years(_CONSOLIDATED_DF, company) or [CURR_YEAR])

    # ── Header with Download button top-right ─────────────────────────────────
    hdr_col, btn_col = st.columns([3, 1])
    with hdr_col:
        st.markdown(f"## {company}")
        st.caption("TIP ESG Sustainability Report · Tire Industry Project")
    with btn_col:
        if years:
            sel_yr = st.selectbox("Year", sorted(years, reverse=True),
                                  key="rpt_year_sel", label_visibility="collapsed")
        else:
            sel_yr = CURR_YEAR

    # ── Load data ─────────────────────────────────────────────────────────────
    from formula_engine import TemplateInputs as TI, calculate as calc
    valid = {f.name for f in TI.__dataclass_fields__.values()}

    step  = dl.get_step_data(comp_hist, sel_yr)
    clean = {k: v for k, v in step.items() if k in valid}
    inp   = TI(company=company, year=sel_yr, **clean)
    out   = calc(inp)

    prev_out = None
    if sel_yr - 1 in years:
        ps = dl.get_step_data(comp_hist, sel_yr - 1)
        pc = {k: v for k, v in ps.items() if k in valid}
        prev_out = calc(TI(company=company, year=sel_yr - 1, **pc))

    kpi_dict = build_kpi_dict_from_outputs(inp, out, prev_out)
    rt       = max(inp.renew_elec_purchased + inp.nonrenew_elec_purchased + inp.self_gen_elec, 1)
    renew_pct = inp.renew_elec_purchased / rt * 100

    # ── Pre-compute historical table (needed by both PDF and web view) ────────
    from formula_engine import TemplateInputs as TI, calculate as calc
    valid = {f.name for f in TI.__dataclass_fields__.values()}
    tbl_yrs = sorted([y for y in years if y <= sel_yr and y >= sel_yr - 9], reverse=True)
    tbl_rows = []
    for y in tbl_yrs:
        sd_t = dl.get_step_data(comp_hist, y)
        sc_t = {k: v for k, v in sd_t.items() if k in valid}
        if not sc_t: continue
        o_t  = calc(TI(company=company, year=y, **sc_t))
        ii_t = TI(company=company, year=y, **sc_t)
        rt2  = max(ii_t.renew_elec_purchased + ii_t.nonrenew_elec_purchased + ii_t.self_gen_elec, 1)
        tbl_rows.append({
            "Year":              y,
            "Production (M T)":  f"{ii_t.production/1e6:.2f}",
            "CO₂ Total (T)":     f"{o_t.total_co2:,.0f}",
            "CO₂ Intensity":     f"{o_t.co2_kpi:.3f}",
            "Energy KPI (GJ/t)": f"{o_t.energy_kpi:.2f}",
            "Renew. Elec. %":    f"{ii_t.renew_elec_purchased/rt2*100:.1f}%",
            "Water KPI (m³/t)":  f"{o_t.water_kpi:.2f}",
            "Waste Recovery %":  f"{o_t.waste_recovery_pct*100:.1f}%",
        })

    # CO₂ trend for last 5 years
    trend_yrs  = sorted([y for y in years if y <= sel_yr])[-5:]
    co2_trend  = []
    for ty in trend_yrs:
        sd = dl.get_step_data(comp_hist, ty)
        sc = {k: v for k, v in sd.items() if k in valid}
        co2_trend.append(calc(TI(company=company, year=ty, **sc)).co2_kpi)

    # Generate PDF using matplotlib — matches the web page content
    import sys as _sys, os as _os
    _sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
    pdf_bytes = None
    try:
        import pdf_charts_v2 as pc
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units    import mm
        from reportlab.pdfgen       import canvas as rl_canvas
        from reportlab.lib.utils    import ImageReader
        import io as _io

        W, H = A4; MARGIN = 12*mm; CW = W - 2*MARGIN
        buf2 = _io.BytesIO()
        cv   = rl_canvas.Canvas(buf2, pagesize=A4)

        # --- Cover band ---
        cv.setFillColor((10/255, 34/255, 64/255))
        cv.rect(0, H - 30*mm, W, 30*mm, fill=1, stroke=0)
        cv.setFont("Helvetica-Bold", 15)
        cv.setFillColor((1,1,1))
        cv.drawString(MARGIN, H - 13*mm, company)
        cv.setFont("Helvetica", 9); cv.setFillColor((0.55,0.65,0.75))
        cv.drawString(MARGIN, H - 20*mm, f"TIP ESG Sustainability Performance Report  ·  {sel_yr}")
        cv.setFont("Helvetica-Bold", 7); cv.setFillColor((0.3,0.9,0.4))
        cv.drawString(W - MARGIN - 55, H - 15*mm, "dss+ Verified Standard")
        cv.setFont("Helvetica-Bold", 28); cv.setFillColor((1,1,1))
        cv.drawRightString(W - MARGIN, H - 17*mm, str(sel_yr))

        cursor = H - 34*mm

        def _embed_img(img_bytes, h=55*mm, w=None):
            nonlocal cursor
            if cursor - h < MARGIN + 5*mm:
                cv.showPage(); cursor = H - MARGIN
            reader = ImageReader(_io.BytesIO(img_bytes))
            use_w = w or CW
            cv.drawImage(reader, MARGIN, cursor - h, width=use_w, height=h, preserveAspectRatio=True)
            cursor -= h + 4*mm

        def _section_hdr(title, subtitle, color=(10/255,34/255,64/255)):
            nonlocal cursor
            cv.setFillColor((0.96, 0.97, 0.99))
            cv.rect(MARGIN, cursor - 10*mm, CW, 10*mm, fill=1, stroke=0)
            cv.setFillColor(color)
            cv.rect(MARGIN, cursor - 10*mm, 2.5, 10*mm, fill=1, stroke=0)
            cv.setFont("Helvetica-Bold", 9); cv.setFillColor(color)
            cv.drawString(MARGIN + 5, cursor - 6*mm, title)
            cv.setFont("Helvetica", 7); cv.setFillColor((0.4,0.4,0.4))
            cv.drawString(MARGIN + 5, cursor - 9*mm, subtitle)
            cursor -= 12*mm

        # --- 6 KPI cards row ---
        kpi_card_data = [
            ("CO₂ Intensity",     f"{out.co2_kpi:.3f}", "tCO₂/t"),
            ("Renewable Elec.",   f"{renew_pct:.1f}",   "%"),
            ("Water Intensity",   f"{out.water_kpi:.2f}","m³/t"),
            ("Waste Recovery",    f"{out.waste_recovery_pct*100:.1f}","%"),
            ("Energy Intensity",  f"{out.energy_kpi:.2f}","GJ/t"),
            ("Production",        f"{inp.production/1e6:.2f}","M T"),
        ]
        card_w = CW / 6; card_h = 18*mm
        if cursor - card_h < MARGIN: cv.showPage(); cursor = H - MARGIN
        for j, (lbl, val, unit) in enumerate(kpi_card_data):
            cx = MARGIN + j * card_w
            cv.setFillColor((0.97, 0.98, 1.0))
            cv.roundRect(cx, cursor - card_h, card_w - 1, card_h, 2, fill=1, stroke=0)
            cv.setFillColor((0.4, 0.45, 0.5))
            cv.setFont("Helvetica", 6); cv.drawCentredString(cx + card_w/2, cursor - 5*mm, lbl.upper())
            cv.setFont("Helvetica-Bold", 11); cv.setFillColor((0.06, 0.13, 0.25))
            cv.drawCentredString(cx + card_w/2, cursor - 10*mm, val)
            cv.setFont("Helvetica", 6); cv.setFillColor((0.5, 0.5, 0.5))
            cv.drawCentredString(cx + card_w/2, cursor - 14*mm, unit)
        cursor -= card_h + 6*mm

        # --- Section 1: Environmental ---
        _section_hdr("1.  Environmental Performance",
                     "CO₂ emissions, energy consumption and climate targets")

        # Build data for charts
        all_hist = {}
        for y_h in years:
            sd_h = dl.get_step_data(comp_hist, y_h)
            sc_h = {k: v for k,v in sd_h.items() if k in valid}
            if not sc_h: continue
            o_h  = calc(TI(company=company, year=y_h, **sc_h))
            ii_h = TI(company=company, year=y_h, **sc_h)
            rt_h = max(ii_h.renew_elec_purchased + ii_h.nonrenew_elec_purchased + ii_h.self_gen_elec, 1)
            all_hist[y_h] = {"co2": o_h.total_co2, "nat_gas": ii_h.nat_gas,
                              "coal": ii_h.coal_sub, "diesel": ii_h.diesel,
                              "renew_gj": ii_h.renew_elec_purchased,
                              "nonrenew_gj": ii_h.nonrenew_elec_purchased,
                              "biomass": ii_h.biomass, "water_m3": ii_h.water_withdrawals,
                              "waste_pct": o_h.waste_recovery_pct*100}
        hy = sorted(all_hist.keys())

        # CO₂ + energy mix side by side
        if hy:
            img_co2 = pc.area_line(hy, [all_hist[y]["co2"] for y in hy],
                                   "Total CO₂ Emissions (tCO₂)", color=pc.C["co2"])
            img_nrg = pc.stacked_bar(hy[-8:], {
                "Nat. Gas": [all_hist[y]["nat_gas"]    for y in hy[-8:]],
                "Renew.":   [all_hist[y]["renew_gj"]   for y in hy[-8:]],
                "Diesel":   [all_hist[y]["diesel"]      for y in hy[-8:]],
                "Coal":     [all_hist[y]["coal"]        for y in hy[-8:]],
            }, "Energy Mix by Source (GJ)",
            color_dict={"Nat. Gas":pc.C["energy"],"Renew.":pc.C["green"],
                        "Diesel":"#78716C","Coal":"#475569"})

            half_w = CW / 2 - 2*mm
            h_img  = 52*mm
            if cursor - h_img < MARGIN: cv.showPage(); cursor = H - MARGIN
            cv.drawImage(ImageReader(_io.BytesIO(img_co2)),
                         MARGIN, cursor-h_img, width=half_w, height=h_img, preserveAspectRatio=True)
            cv.drawImage(ImageReader(_io.BytesIO(img_nrg)),
                         MARGIN+half_w+2*mm, cursor-h_img, width=half_w, height=h_img, preserveAspectRatio=True)
            cursor -= h_img + 5*mm

        # --- Section 2: Resource Efficiency ---
        _section_hdr("2.  Resource Efficiency",
                     "Water withdrawals, waste management and circular economy",
                     color=(0.03,0.35,0.43))
        if hy:
            img_wat = pc.bar_chart(hy, [all_hist[y]["water_m3"]/1e6 for y in hy],
                                   "Water Withdrawals (M m³)", "M m³", color=pc.C["water"])
            img_wst = pc.area_with_target(hy, [all_hist[y]["waste_pct"] for y in hy],
                                          "Waste Recovery Rate (%)", "%",
                                          color=pc.C["waste"])
            h_img = 52*mm
            if cursor - h_img < MARGIN: cv.showPage(); cursor = H - MARGIN
            cv.drawImage(ImageReader(_io.BytesIO(img_wat)),
                         MARGIN, cursor-h_img, width=half_w, height=h_img, preserveAspectRatio=True)
            cv.drawImage(ImageReader(_io.BytesIO(img_wst)),
                         MARGIN+half_w+2*mm, cursor-h_img, width=half_w, height=h_img, preserveAspectRatio=True)
            cursor -= h_img + 5*mm

        # --- Section 3: Historical table ---
        _section_hdr("3.  Historical Performance Data",
                     f"{max(tbl_yrs[-1] if tbl_rows else sel_yr-9, sel_yr-9)}–{sel_yr}",
                     color=(0.09,0.32,0.09))
        if tbl_rows:
            cv.setFont("Helvetica-Bold", 7); cv.setFillColor((0.06,0.13,0.25))
            headers = list(tbl_rows[0].keys())
            col_w   = CW / len(headers)
            row_h   = 5.5*mm
            # header row
            if cursor - row_h < MARGIN: cv.showPage(); cursor = H - MARGIN
            cv.setFillColor((0.94,0.95,0.98))
            cv.rect(MARGIN, cursor-row_h, CW, row_h, fill=1, stroke=0)
            for j, h_txt in enumerate(headers):
                cv.setFont("Helvetica-Bold", 6); cv.setFillColor((0.3,0.35,0.4))
                cv.drawCentredString(MARGIN + (j+0.5)*col_w, cursor-4*mm, str(h_txt)[:12])
            cursor -= row_h
            for ri, row in enumerate(tbl_rows):
                if cursor - row_h < MARGIN: cv.showPage(); cursor = H - MARGIN
                if ri % 2 == 0:
                    cv.setFillColor((0.975,0.978,0.985))
                    cv.rect(MARGIN, cursor-row_h, CW, row_h, fill=1, stroke=0)
                for j, col_name in enumerate(headers):
                    cv.setFont("Helvetica", 6)
                    cv.setFillColor((0.15,0.15,0.2) if j==0 else (0.35,0.38,0.42))
                    cv.drawCentredString(MARGIN+(j+0.5)*col_w, cursor-4*mm, str(row.get(col_name,""))[:10])
                cursor -= row_h

        # --- Footer ---
        cv.setFillColor((0.95,0.96,0.98))
        cv.rect(0, 0, W, 12*mm, fill=1, stroke=0)
        cv.setFont("Helvetica", 6); cv.setFillColor((0.4,0.4,0.4))
        cv.drawString(MARGIN, 5*mm,
            "Methodology: GHG Protocol (Scope 1+2) · TIP KPI definitions v3.1 · IEA 2023 emission factors")
        from datetime import date as _ddate
        cv.drawRightString(W-MARGIN, 5*mm,
            f"Generated {_ddate.today().strftime('%d %b %Y')} · TIP ESG Platform by dss+")

        cv.save(); buf2.seek(0); pdf_bytes = buf2.read()
    except Exception as _e:
        pdf_bytes = None
        st.warning(f"PDF generation error: {_e}")
    filename = f"{company.replace(' ','_')}_Sustainability_Report_{sel_yr}.pdf"
    with btn_col:
        if pdf_bytes:
            st.download_button("⬇ Download PDF", data=pdf_bytes, file_name=filename,
                               mime="application/pdf", type="primary",
                               use_container_width=True)
        else:
            st.button("⬇ Download PDF", disabled=True, use_container_width=True)

    st.divider()

    # ══════════════════════════════════════════════════════════════════════════
    # REPORT BODY — styled as a professional one-page CSR report
    # ══════════════════════════════════════════════════════════════════════════

    # ── Cover band ────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#0A2240 0%,#164E63 100%);
        border-radius:12px;padding:28px 32px;margin-bottom:20px;
        display:flex;justify-content:space-between;align-items:center">
      <div>
        <div style="color:rgba(255,255,255,.5);font-size:11px;font-weight:500;
            text-transform:uppercase;letter-spacing:.8px;margin-bottom:6px">
          TIP ESG Platform · Tire Industry Project
        </div>
        <div style="color:#fff;font-size:28px;font-weight:800;letter-spacing:-.5px;
            margin-bottom:4px">{_html.escape(company)}</div>
        <div style="color:rgba(255,255,255,.6);font-size:14px">
          Sustainability Performance Report · {sel_yr}
        </div>
      </div>
      <div style="text-align:right">
        <div style="color:rgba(255,255,255,.4);font-size:10px;text-transform:uppercase">
          Reporting Year</div>
        <div style="color:#fff;font-size:52px;font-weight:800;line-height:1">
          {sel_yr}</div>
        <div style="color:{GREEN};font-size:11px;font-weight:600;margin-top:4px">
          ● dss+ Verified Standard</div>
      </div>
    </div>""", unsafe_allow_html=True)

    # ── Executive Summary ──────────────────────────────────────────────────────
    def _yoy_str(cur, prev_val, lower=True):
        if not prev_val or prev_val == 0: return ""
        pct = (cur - prev_val) / abs(prev_val) * 100
        good = pct <= 0 if lower else pct >= 0
        arrow = "▼" if pct < 0 else "▲"
        col   = GREEN if good else RED
        return f'<span style="color:{col};font-size:11px;font-weight:600">{arrow} {abs(pct):.1f}%</span>'

    co2_yoy    = _yoy_str(out.co2_kpi,    prev_out.co2_kpi if prev_out else None)
    energy_yoy = _yoy_str(out.energy_kpi, prev_out.energy_kpi if prev_out else None)
    water_yoy  = _yoy_str(out.water_kpi,  prev_out.water_kpi if prev_out else None)
    waste_yoy  = _yoy_str(out.waste_recovery_pct, prev_out.waste_recovery_pct if prev_out else None, lower=False)

    kpi_summary = [
        ("CO₂ Intensity",     f"{out.co2_kpi:.3f}",           "tCO₂/t", co2_yoy,    CAT_CO2),
        ("Renewable Elec.",   f"{renew_pct:.1f}",              "%",       "",          CAT_RENEW),
        ("Water Intensity",   f"{out.water_kpi:.2f}",          "m³/t",    water_yoy,  CAT_WATER),
        ("Waste Recovery",    f"{out.waste_recovery_pct*100:.1f}","%",    waste_yoy,  CAT_WASTE),
        ("Energy Intensity",  f"{out.energy_kpi:.2f}",         "GJ/t",    energy_yoy, CAT_ENERGY),
        ("Production",        f"{inp.production/1e6:.2f}",     "M T",     "",          NAVY),
    ]
    kpi_cols = st.columns(6)
    for i, (label, val, unit, yoy_h, color) in enumerate(kpi_summary):
        with kpi_cols[i]:
            st.markdown(f"""
            <div style="background:#fff;border:1px solid {BORDER};border-radius:8px;
                padding:10px 8px;text-align:center;
                height:90px;display:flex;flex-direction:column;
                justify-content:space-between;align-items:center;
                animation:tipFadeIn 400ms ease-out {i*50}ms both">
              <div style="font-size:9px;color:{MUTED};text-transform:uppercase;
                  letter-spacing:.5px;font-weight:600">{label}</div>
              <div style="font-size:19px;font-weight:800;color:{color};
                  font-variant-numeric:tabular-nums;line-height:1">{val}</div>
              <div style="font-size:9px;color:{MUTED}">{unit}</div>
              <div style="margin-top:2px;min-height:14px">{yoy_h}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)

    # ── Section 1: Environmental Performance ─────────────────────────────────
    st.markdown(f"""
    <div style="border-left:3px solid {CAT_CO2};padding:4px 0 4px 12px;margin-bottom:12px">
      <div style="font-size:14px;font-weight:700;color:{TEXT}">
        1. Environmental Performance</div>
      <div style="font-size:11px;color:{MUTED}">
        CO₂ emissions, energy consumption and climate targets</div>
    </div>""", unsafe_allow_html=True)

    r1c1, r1c2 = st.columns(2, gap="medium")
    with r1c1:
        # CO₂ trend
        co2_all = []
        for y in years:
            sd = dl.get_step_data(comp_hist, y)
            sc = {k: v for k, v in sd.items() if k in valid}
            if sc: co2_all.append((y, calc(TI(company=company, year=y, **sc)).total_co2))
        if co2_all:
            ys_c, vals_c = zip(*co2_all)
            fig_co2 = go.Figure()
            fig_co2.add_trace(go.Scatter(
                x=list(ys_c), y=list(vals_c), mode="lines+markers",
                fill="tozeroy", fillcolor="rgba(71,85,105,0.08)",
                line=dict(color=CAT_CO2, width=2.5),
                marker=dict(size=6, color=CAT_CO2),
                hovertemplate="<b>%{x}</b>: %{y:,.0f} tCO₂<extra></extra>",
            ))
            fig_co2.update_layout(**chart_layout_defaults("Total CO₂ Emissions (tCO₂)", height=220,
                                                           showlegend=False),
                                   yaxis=dict(tickformat=",", gridcolor="#F1F5F9", showline=True, linecolor="#999", showticklabels=True, tickfont=dict(size=12, color="#1C2E3F")),
                                   xaxis=dict(gridcolor="#F1F5F9"))
            apply_chart_animation(fig_co2)
            st.plotly_chart(fig_co2, use_container_width=True)

    with r1c2:
        # Energy mix stacked bar
        fuel_data = [(y, dl.get_step_data(comp_hist, y)) for y in years[-8:]]
        if fuel_data:
            fig_nrg = go.Figure()
            fuel_keys = [("Nat. Gas","nat_gas",CAT_ENERGY),
                         ("Coal","coal_sub","#475569"),
                         ("Diesel","diesel","#78716C"),
                         ("Renew. Elec","renew_elec_purchased",CAT_RENEW)]
            for lbl, fkey, fcol in fuel_keys:
                vals = [sd.get(fkey, 0) for _, sd in fuel_data]
                if any(v>0 for v in vals):
                    fig_nrg.add_trace(go.Bar(
                        x=[y for y,_ in fuel_data], y=vals,
                        name=lbl, marker_color=fcol, marker_line_width=0,
                        hovertemplate=f"<b>{lbl}</b>: %{{y:,.0f}} GJ<extra></extra>",
                    ))
            fig_nrg.update_layout(**chart_layout_defaults("Energy Mix by Source (GJ)", height=220),
                                   barmode="stack", bargap=0.2,
                                   yaxis=dict(tickformat=",", gridcolor="#F1F5F9", showline=True, linecolor="#999", showticklabels=True, tickfont=dict(size=12, color="#1C2E3F")),
                                   xaxis=dict(gridcolor="#F1F5F9"))
            apply_chart_animation(fig_nrg)
            st.plotly_chart(fig_nrg, use_container_width=True)

    # ── Section 2: Resource Efficiency ────────────────────────────────────────
    st.markdown(f"""
    <div style="border-left:3px solid {CAT_WATER};padding:4px 0 4px 12px;margin-bottom:12px">
      <div style="font-size:14px;font-weight:700;color:{TEXT}">
        2. Resource Efficiency</div>
      <div style="font-size:11px;color:{MUTED}">
        Water withdrawals, waste management and circular economy</div>
    </div>""", unsafe_allow_html=True)

    r2c1, r2c2 = st.columns(2, gap="medium")
    with r2c1:
        water_vals = [(y, dl.get_step_data(comp_hist,y).get("water_withdrawals",0)) for y in years]
        if water_vals:
            ys_w, vals_w = zip(*water_vals)
            fig_wat = go.Figure(go.Bar(
                x=list(ys_w), y=[v/1e6 for v in vals_w],
                marker_color=CAT_WATER, marker_line_width=0, opacity=0.85,
                hovertemplate="<b>%{x}</b>: %{y:.2f} M m³<extra></extra>",
            ))
            fig_wat.update_layout(**chart_layout_defaults("Water Withdrawals (M m³)", height=200,
                                                           showlegend=False),
                                   yaxis=dict(gridcolor="#F1F5F9"),
                                   xaxis=dict(gridcolor="#F1F5F9"))
            apply_chart_animation(fig_wat)
            st.plotly_chart(fig_wat, use_container_width=True)

    with r2c2:
        waste_rec_vals = []
        for y in years:
            sd = dl.get_step_data(comp_hist, y)
            sc = {k: v for k, v in sd.items() if k in valid}
            if sc:
                o = calc(TI(company=company, year=y, **sc))
                waste_rec_vals.append((y, o.waste_recovery_pct*100))
        if waste_rec_vals:
            ys_wr, vals_wr = zip(*waste_rec_vals)
            fig_wr = go.Figure()
            fig_wr.add_hline(y=90, line_dash="dot", line_color=GREEN,
                             line_width=1.5, annotation_text="Target 90%",
                             annotation_font=dict(size=9, color=GREEN))
            fig_wr.add_trace(go.Scatter(
                x=list(ys_wr), y=list(vals_wr), mode="lines+markers",
                fill="tozeroy", fillcolor="rgba(124,58,237,0.08)",
                line=dict(color=CAT_WASTE, width=2.5),
                marker=dict(size=6, color=CAT_WASTE),
                hovertemplate="<b>%{x}</b>: %{y:.1f}%<extra></extra>",
            ))
            fig_wr.update_layout(**chart_layout_defaults("Waste Recovery Rate (%)", height=200,
                                                          showlegend=False),
                                  yaxis=dict(range=[0,105], ticksuffix="%",
                                             gridcolor="#F1F5F9"),
                                  xaxis=dict(gridcolor="#F1F5F9"))
            apply_chart_animation(fig_wr)
            st.plotly_chart(fig_wr, use_container_width=True)

    # ── Section 3: Historical KPI Table ───────────────────────────────────────
    st.markdown(f"""
    <div style="border-left:3px solid {GREEN};padding:4px 0 4px 12px;margin-bottom:12px">
      <div style="font-size:14px;font-weight:700;color:{TEXT}">
        3. Historical Performance Data ({max(years)-9 if len(years)>=10 else min(years)}–{sel_yr})</div>
    </div>""", unsafe_allow_html=True)

    # tbl_rows already computed above (before PDF generation)
    if tbl_rows:
        tbl_df = pd.DataFrame(tbl_rows)
        st.dataframe(
            tbl_df.style
                .set_properties(**{"text-align":"right","font-size":"12px"})
                .set_table_styles([
                    {"selector":"th","props":[
                        ("font-size","10px"),("text-transform","uppercase"),
                        ("letter-spacing",".4px"),("color","#64748B"),
                        ("background","#F8FAFC"),("padding","8px 12px")]},
                    {"selector":"td:first-child","props":[
                        ("font-weight","600"),("color","#0F172A"),
                        ("text-align","center")]},
                ]),
            use_container_width=True, hide_index=True,
        )

    # ── Footer ────────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="background:#F8FAFC;border-radius:8px;padding:12px 20px;margin-top:16px;
        display:flex;justify-content:space-between;align-items:center;
        border:1px solid {BORDER}">
      <div style="font-size:11px;color:{MUTED}">
        Methodology: GHG Protocol (Scope 1+2) · TIP KPI definitions v3.1 ·
        Emission factors: IEA 2023</div>
      <div style="font-size:11px;color:{MUTED}">
        Generated {_date.today().strftime('%d %b %Y')} · TIP ESG Platform powered by dss+</div>
    </div>""", unsafe_allow_html=True)


def page_settings():
    """Settings page — for both client and DSS+ users."""
    st.markdown(section_header_html("Settings", "Account & preferences"),
                unsafe_allow_html=True)

    tab_acct, tab_notif, tab_about = st.tabs(["Account", "Notifications", "About"])

    with tab_acct:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Profile**")
            st.text_input("Display Name", value=st.session_state.user_name,
                          key="set_name")
            st.text_input("Email", value=st.session_state.user_email,
                          disabled=True, key="set_email")
            role = "dss+ Analyst (Internal)" if st.session_state.is_dss else "Client Company User"
            st.text_input("Role", value=role, disabled=True, key="set_role")
            if st.button("Save Profile", key="save_profile"):
                st.success("Profile updated.")
        with col2:
            st.markdown("**Security**")
            st.text_input("Current Password", type="password", key="set_cur_pw")
            st.text_input("New Password",     type="password", key="set_new_pw")
            st.text_input("Confirm Password", type="password", key="set_cfm_pw")
            if st.button("Change Password", key="change_pw"):
                st.info("Password change will be available in the full production release.")

    with tab_notif:
        st.markdown("**Email notifications**")
        st.checkbox("Submission deadline reminders",          value=True,  key="n1")
        st.checkbox("Verification status updates",            value=True,  key="n2")
        st.checkbox("Sector benchmarks published",            value=False, key="n3")
        st.checkbox("AI anomaly alerts",                      value=True,  key="n4")
        if st.button("Save notification preferences", key="save_notif"):
            st.success("Preferences saved.")

    with tab_about:
        st.markdown(f"""
        **TIP ESG Platform**

        Version 1.0 · Built for the WBCSD Tire Industry Project by dss+

        - Formula engine: GHG Protocol Scope 1 & 2
        - Benchmark data: TIP member companies 2009–{CURR_YEAR}
        - AI assistant: Local Ollama (phi3) / Azure OpenAI Enterprise
        - Storage: Local filesystem (v1) → Azure SharePoint (v2)

        *For technical support contact your dss+ account manager.*
        """)


def page_portfolio():
    """DSS+ Portfolio — 10-company grid with status chips and KPI heatmap."""
    st.markdown(section_header_html(
        "Portfolio Overview",
        f"All TIP member companies · {CURR_YEAR} reporting cycle",
        badge=f"{len(_COMPANIES)} Companies",
    ), unsafe_allow_html=True)

    if _CONSOLIDATED_DF.empty:
        st.markdown(empty_state_html("🗂️", "No data loaded",
            "Run python build_esg_master.py to load company data."),
            unsafe_allow_html=True)
        return

    # Status summary bar
    statuses = ["complete", "review", "pending"]
    status_map = {}
    for i, co in enumerate(_COMPANIES):
        # Determine status from data completeness
        hist = dl.get_company_hist(_CONSOLIDATED_DF, co)
        step = dl.get_step_data(hist, CURR_YEAR) if hist else {}
        n    = len(step)
        status_map[co] = "complete" if n >= 15 else "review" if n >= 5 else "pending"

    n_complete = sum(1 for s in status_map.values() if s == "complete")
    n_review   = sum(1 for s in status_map.values() if s == "review")
    n_pending  = sum(1 for s in status_map.values() if s == "pending")

    st.markdown(f"""
    <div style="display:flex;gap:12px;margin-bottom:20px">
      <div style="background:#DCFCE7;border-radius:8px;padding:12px 20px;flex:1;text-align:center">
        <div style="font-size:28px;font-weight:700;color:#166534">{n_complete}</div>
        <div style="font-size:11px;color:#166534;font-weight:500">Complete</div>
      </div>
      <div style="background:#FEF3C7;border-radius:8px;padding:12px 20px;flex:1;text-align:center">
        <div style="font-size:28px;font-weight:700;color:#92400E">{n_review}</div>
        <div style="font-size:11px;color:#92400E;font-weight:500">In Review</div>
      </div>
      <div style="background:#F1F5F9;border-radius:8px;padding:12px 20px;flex:1;text-align:center">
        <div style="font-size:28px;font-weight:700;color:#475569">{n_pending}</div>
        <div style="font-size:11px;color:#475569;font-weight:500">Pending</div>
      </div>
    </div>""", unsafe_allow_html=True)

    # Company grid — 2 columns
    cols = st.columns(2, gap="medium")
    from formula_engine import TemplateInputs as TI, calculate as calc
    valid = {f.name for f in TI.__dataclass_fields__.values()}

    for i, company in enumerate(_COMPANIES):
        hist  = dl.get_company_hist(_CONSOLIDATED_DF, company)
        step  = dl.get_step_data(hist, CURR_YEAR) if hist else {}
        clean = {k: v for k, v in step.items() if k in valid}
        out   = calc(TI(company=company, year=CURR_YEAR, **clean))
        kpis  = {"co2_kpi": out.co2_kpi, "energy_kpi": out.energy_kpi,
                 "water_kpi": out.water_kpi}

        with cols[i % 2]:
            st.markdown(co_card_html(
                company, status_map[company], CURR_YEAR,
                kpis, anim_delay=i * 60,
            ), unsafe_allow_html=True)
            if st.button(f"Open {company.split()[0]} Template →",
                         key=f"port_view_{i}", use_container_width=True):
                st.session_state.portfolio_company  = company
                st.session_state.reporting_company  = company
                st.session_state.dss_verif_company  = company
                st.session_state.dss_ready_company  = company
                st.session_state.dss_analy_company  = company
                st.session_state.company_setup_done = False
                st.session_state.template_done      = False
                st.session_state.step               = 0
                st.session_state.page               = "company_data"
                st.rerun()
        st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)


def page_doc_library():
    """DSS+ Document Library — upload PDFs, AI-extract KPIs."""
    st.markdown(section_header_html(
        "Document Library",
        "Upload company submissions and source documents",
    ), unsafe_allow_html=True)

    col_up, col_lib = st.columns([1, 1], gap="large")

    with col_up:
        st.markdown("**Upload Document**")
        company_sel = st.selectbox("Company", _COMPANIES, key="doclib_co")
        year_sel    = st.number_input("Year", min_value=2009,
                                      max_value=CURR_YEAR + 1,
                                      value=CURR_YEAR, step=1, key="doclib_yr")
        doc_type    = st.selectbox("Document Type",
                                   ["Annual ESG Report", "Sustainability Appendix",
                                    "GHG Inventory", "Audit Evidence", "Other"],
                                   key="doclib_type")
        uploaded    = st.file_uploader("Upload PDF or Excel",
                                        type=["pdf", "xlsx", "csv"],
                                        key="doclib_file")
        if uploaded and st.button("📤 Upload & Extract KPIs",
                                   type="primary", use_container_width=True,
                                   key="doclib_upload"):
            with st.spinner("Uploading and extracting KPIs via AI…"):
                import time; time.sleep(1.5)
            st.success(f"Uploaded {uploaded.name} for {company_sel} {year_sel}. "
                       "AI extraction queued — results appear in Verification Queue.")

    with col_lib:
        st.markdown("**Recent Documents**")
        docs = [
            ("VerdaTyres Corp",    2023, "Annual ESG Report",     "complete", "13 May 2026"),
            ("AlphaTread Ltd",     2023, "GHG Inventory",         "review",   "12 May 2026"),
            ("BetaRubber Inc",     2022, "Sustainability Appendix","complete", "10 May 2026"),
            ("DeltaGrip GmbH",     2023, "Annual ESG Report",     "pending",  "08 May 2026"),
        ]
        for co, yr, dtype, status, ts in docs:
            chip = status_chip_html(status)
            st.markdown(f"""
            <div style="background:#fff;border:1px solid {BORDER};border-radius:8px;
                padding:12px 14px;margin-bottom:6px">
              <div style="display:flex;justify-content:space-between;align-items:center">
                <div>
                  <div style="font-size:13px;font-weight:500;color:{TEXT}">{co} · {yr}</div>
                  <div style="font-size:11px;color:{MUTED}">{dtype} · {ts}</div>
                </div>
                {chip}
              </div>
            </div>""", unsafe_allow_html=True)


def page_sector_reports():
    """DSS+ Sector Reports — generate annual TIP sustainability report."""
    st.markdown(section_header_html(
        "Sector Reports",
        "Generate the annual TIP consolidated sustainability report",
    ), unsafe_allow_html=True)

    col_gen, col_prev = st.columns([1, 1], gap="large")

    with col_gen:
        st.markdown("**Generate Report**")
        rpt_year = st.selectbox("Report Year", LONG_YEARS[::-1], key="sr_year")
        rpt_scope = st.multiselect("Include Companies", _COMPANIES,
                                   default=_COMPANIES, key="sr_scope")
        rpt_format = st.radio("Format", ["PDF (Executive)", "Excel (Full Data)"],
                              key="sr_fmt", horizontal=True)
        st.markdown("**Sections to include:**")
        c1, c2 = st.columns(2)
        with c1:
            inc_co2    = st.checkbox("CO₂ & GHG",        True, key="inc_co2")
            inc_energy = st.checkbox("Energy",            True, key="inc_energy")
            inc_water  = st.checkbox("Water",             True, key="inc_water")
        with c2:
            inc_waste  = st.checkbox("Waste",             True, key="inc_waste")
            inc_bench  = st.checkbox("Benchmarking",      True, key="inc_bench")
            inc_sdg    = st.checkbox("SDG Roadmap",       False, key="inc_sdg")

        if st.button("🌍 Generate Sector Report", type="primary",
                     use_container_width=True, key="gen_sector"):
            with st.spinner(f"Generating {rpt_year} sector report…"):
                import time; time.sleep(2)
            if not _CONSOLIDATED_DF.empty:
                subset = _CONSOLIDATED_DF[
                    (_CONSOLIDATED_DF["Company"].isin(rpt_scope)) &
                    (_CONSOLIDATED_DF["Year"] == rpt_year)
                ]
                csv_bytes = subset.to_csv(index=False).encode()
                st.download_button(
                    "⬇ Download Sector Data (CSV)",
                    data=csv_bytes,
                    file_name=f"TIP_Sector_Report_{rpt_year}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            st.success(f"Sector report for {rpt_year} ready.")

    with col_prev:
        st.markdown("**Previous Reports**")
        prev_reports = [
            (2023, "TIP Annual Sustainability Report 2023", "Published"),
            (2022, "TIP Annual Sustainability Report 2022", "Published"),
            (2021, "TIP Annual Sustainability Report 2021", "Archived"),
        ]
        for yr, name, status in prev_reports:
            chip = status_chip_html("complete" if status == "Published" else "pending")
            st.markdown(f"""
            <div style="background:#fff;border:1px solid {BORDER};border-radius:8px;
                padding:12px 14px;margin-bottom:6px">
              <div style="display:flex;justify-content:space-between;align-items:center">
                <div>
                  <div style="font-size:13px;font-weight:500;color:{TEXT}">{name}</div>
                  <div style="font-size:11px;color:{MUTED}">{yr} · {status}</div>
                </div>
                {chip}
              </div>
            </div>""", unsafe_allow_html=True)


def page_admin():
    """DSS+ Admin — tenant management, RBAC, AI usage."""
    st.markdown(section_header_html(
        "Admin", "Platform administration",
        badge="DSS+ Only",
    ), unsafe_allow_html=True)

    tab_tenants, tab_users, tab_ai, tab_system = st.tabs(
        ["Tenants", "Users", "AI Usage", "System"]
    )

    with tab_tenants:
        st.markdown("**Active Tenants (TIP Member Companies)**")
        for i, co in enumerate(_COMPANIES):
            st.markdown(f"""
            <div style="background:#fff;border:1px solid {BORDER};border-radius:8px;
                padding:10px 14px;margin-bottom:4px;display:flex;
                align-items:center;justify-content:space-between">
              <div style="font-size:13px;font-weight:500;color:{TEXT}">{co}</div>
              <div style="display:flex;gap:8px">
                {status_chip_html('complete')}
                <span style="font-size:11px;color:{MUTED}">Active since 2021</span>
              </div>
            </div>""", unsafe_allow_html=True)

        st.markdown('<div style="height:12px"></div>', unsafe_allow_html=True)
        with st.expander("+ Add New Tenant"):
            st.text_input("Company Name", key="admin_co_name")
            st.text_input("Contact Email", key="admin_co_email")
            st.number_input("Joined Year", min_value=2009,
                            max_value=CURR_YEAR, value=CURR_YEAR,
                            key="admin_co_year")
            if st.button("Add Tenant", type="primary", key="admin_add_co"):
                st.info("Tenant provisioning will be available in v2 (Azure AD integration).")

    with tab_users:
        st.markdown("**Role-Based Access Control**")
        roles = {
            "Client User":    "Edit and submit own company data",
            "Client Admin":   "Manage users + approve within tenant",
            "DSS+ Analyst":   "Cross-tenant read, verification write",
            "DSS+ Admin":     "Super-user, manage all tenants and AI settings",
        }
        for role, desc in roles.items():
            st.markdown(f"""
            <div style="background:#fff;border:1px solid {BORDER};border-radius:8px;
                padding:10px 14px;margin-bottom:4px">
              <div style="font-size:13px;font-weight:600;color:{TEXT}">{role}</div>
              <div style="font-size:11px;color:{MUTED}">{desc}</div>
            </div>""", unsafe_allow_html=True)

    with tab_ai:
        st.markdown("**AI Usage Logs**")
        st.info("AI usage logs are stored in data_storage/chat_logs/ (JSONL format). "
                "Full usage analytics will be available in v2.")
        log_dir = Path("data_storage/chat_logs")
        if log_dir.exists():
            logs = list(log_dir.glob("*.jsonl"))
            st.metric("Log files this week", len(logs))
            for lf in sorted(logs, reverse=True)[:5]:
                st.markdown(f"• `{lf.name}` — {lf.stat().st_size:,} bytes")

    with tab_system:
        st.markdown("**System Information**")
        import platform, sys
        info = {
            "Python": sys.version.split()[0],
            "Platform": platform.system(),
            "Data Year Range": f"{cfg.DATA_YEAR_START}–{cfg.DATA_YEAR_END}",
            "Companies": len(_COMPANIES),
            "Master CSV rows": len(_CONSOLIDATED_DF) if not _CONSOLIDATED_DF.empty else 0,
        }
        for k, v in info.items():
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;
                padding:6px 0;border-bottom:1px solid {BG};font-size:13px">
              <span style="color:{MUTED}">{k}</span>
              <span style="color:{TEXT};font-weight:500">{v}</span>
            </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# MAIN ROUTER
# ─────────────────────────────────────────────────────────
if not st.session_state.authenticated:
    show_login()
else:
    show_sidebar()
    page = st.session_state.page

    # ── Client pages ──────────────────────────────────────
    if   page == "home":            page_home()
    elif page == "entry":           page_entry()
    elif page == "my_records":      page_my_records()
    elif page == "dashboard":       page_my_dashboard()
    elif page == "benchmarking":    page_benchmarking()
    elif page == "reports":         page_reports()
    elif page == "settings":        page_settings()

    # ── DSS+ pages ────────────────────────────────────────
    elif page == "portfolio":       page_portfolio()
    elif page == "company_data":    page_company_data()
    elif page == "verification":    page_verification()
    elif page == "analysis":        page_analysis()
    elif page == "readiness":       page_readiness()
    elif page == "doc_library":     page_doc_library()
    elif page == "sector_reports":  page_sector_reports()
    elif page == "admin":           page_admin()

    else:
        # Fallback — redirect to correct home
        st.session_state.page = "portfolio" if st.session_state.is_dss else "home"
        st.rerun()